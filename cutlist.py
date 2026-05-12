# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List
import base64
import logging

import matplotlib
matplotlib.use("Agg")  # non-interactive backend
import matplotlib.pyplot as plt
import pandas as pd
from module_templates import resolve_template
from module_rules import (
    default_shelf_count,
    dishwasher_installation_metrics,
    module_supports_adjustable_shelves,
)

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image as RLImage,
    HRFlowable,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from ui_assembly import assembly_instructions
from ui_pdf_export import (
    _friendly_part_name,
    _friendly_position_name,
    _kant_legend,
    _module_preassembly_lines,
    _module_tool_hardware_lines,
    _part_role_note,
    _pdf_t,
    _short_part_code,
)

_LOG = logging.getLogger(__name__)


def _find_font_file(filename: str) -> str | None:
    """Find font file in typical locations: ./fonts, module folder, cwd."""
    candidates: list[Path] = []
    here = Path(__file__).resolve().parent
    cwd = Path.cwd()
    candidates += [
        here / "fonts" / filename,
        here / filename,
        cwd / "fonts" / filename,
        cwd / filename,
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


def _register_fonts() -> None:
    """Register DejaVu fonts once (safe to call multiple times)."""
    try:
        pdfmetrics.getFont("DejaVuSans")
        pdfmetrics.getFont("DejaVuSans-Bold")
        return
    except Exception as ex:
        _LOG.debug("DejaVu font not registered yet or unavailable: %s", ex)

    regular = _find_font_file("DejaVuSans.ttf")
    bold = _find_font_file("DejaVuSans-Bold.ttf")
    if regular:
        pdfmetrics.registerFont(TTFont("DejaVuSans", regular))
    if bold:
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold))


MANUFACTURING_PROFILES: Dict[str, Dict[str, Any]] = {
    "EU_SRB": {
        "label": "EU / SRB - frameless (na fugu)",
        "front_gap_mm": 2.0,
        "cut_rounding_step_mm": 0.5,
        "cut_rounding_only_cut": True,
        "mounting_tolerance_total_mm": 10,  # укупно (лево+десно)
        "handle_pull_mm": 2,
        "back_inset_mm": 10,
    },
    "EU_SRB_BLUM": {
        "label": "EU / SRB - Blum",
        "front_gap_mm": 2.0,
        "cut_rounding_step_mm": 0.5,
        "cut_rounding_only_cut": True,
        "mounting_tolerance_total_mm": 10,
        "handle_pull_mm": 2,
        "back_inset_mm": 10,
    },
    "EU_SRB_HETTICH": {
        "label": "EU / SRB - Hettich",
        "front_gap_mm": 2.0,
        "cut_rounding_step_mm": 0.5,
        "cut_rounding_only_cut": True,
        "mounting_tolerance_total_mm": 10,
        "handle_pull_mm": 2,
        "back_inset_mm": 10,
    },
    "EU_SRB_GTV": {
        "label": "EU / SRB - GTV",
        "front_gap_mm": 2.0,
        "cut_rounding_step_mm": 0.5,
        "cut_rounding_only_cut": True,
        "mounting_tolerance_total_mm": 10,
        "handle_pull_mm": 2,
        "back_inset_mm": 10,
    },
}


HARDWARE_BRAND_CATALOGS: Dict[str, Dict[str, str]] = {
    "BLUM": {
        "hinge": "Blum CLIP-top 110° (71B...)",
        "slide": "Blum TANDEM plus BLUMOTION (par)",
        "liftup": "Blum AVENTOS HK-S (par)",
        "dish_hinge": "Blum 79T9550",
        "handle": "Po izboru korisnika",
        "hinge_screw": "3.5x16 mm",
        "slide_screw": "3.5x16 mm",
        "back_fix": "3x16 mm / ekser 1.4x25 mm",
        "wall_mount_screw": "5x60 mm",
        "confirmat": "7x50 mm",
        "dowel": "8x30 mm",
        "shelf_pin": "5 mm",
        "leg": "Podesiva h=100 mm",
        "hang_rail": "Nosač + šina za viseće elemente",
        "wall_anchor": "Tipl + vijak 8x80 mm",
        "cabinet_connector": "Spojnica korpusa + vijak",
        "plinth_clip": "Klipsa za soklu",
        "worktop_fix": "Vijak / ugaonik za radnu ploču",
        "anti_tip": "Anti-tip ugaonik / traka",
        "appliance_hookup": "Priključni set uređaja",
        "sliding_track":   "Hafele Slido / Accuride klizni sistem vrata (set)",
        "hanging_rod":     "Okrugla šipka za vešanje O25mm + par nosača",
        "hinge_plate":     "Blum CLIP-top Montageplatte 175H71 / 175L71",
        "handle_screw":    "M4 x 50mm vijak za ručku",
    },
    "HETTICH": {
        "hinge": "Hettich Sensys 110°",
        "slide": "Hettich Quadro (par)",
        "liftup": "Hettich Free flap (par)",
        "dish_hinge": "Hettich set za front MZS",
        "handle": "Po izboru korisnika",
        "hinge_screw": "3.5x16 mm",
        "slide_screw": "3.5x16 mm",
        "back_fix": "3x16 mm / ekser 1.4x25 mm",
        "wall_mount_screw": "5x60 mm",
        "confirmat": "7x50 mm",
        "dowel": "8x30 mm",
        "shelf_pin": "5 mm",
        "leg": "Podesiva h=100 mm",
        "hang_rail": "Nosač + šina za viseće elemente",
        "wall_anchor": "Tipl + vijak 8x80 mm",
        "cabinet_connector": "Spojnica korpusa + vijak",
        "plinth_clip": "Klipsa za soklu",
        "worktop_fix": "Vijak / ugaonik za radnu ploču",
        "anti_tip": "Anti-tip ugaonik / traka",
        "appliance_hookup": "Priključni set uređaja",
        "sliding_track":   "Hettich EKU Porta 50 klizni sistem vrata (set)",
        "hanging_rod":     "Okrugla šipka za vešanje O25mm + par nosača",
        "hinge_plate":     "",
        "handle_screw":    "M4 x 50mm vijak za ručku",
    },
    "GTV": {
        "hinge": "GTV zglobna šarka 110°",
        "slide": "GTV klizač fioke (par)",
        "liftup": "GTV lift-up mehanizam (par)",
        "dish_hinge": "GTV set za front MZS",
        "handle": "Po izboru korisnika",
        "hinge_screw": "3.5x16 mm",
        "slide_screw": "3.5x16 mm",
        "back_fix": "3x16 mm / ekser 1.4x25 mm",
        "wall_mount_screw": "5x60 mm",
        "confirmat": "7x50 mm",
        "dowel": "8x30 mm",
        "shelf_pin": "5 mm",
        "leg": "Podesiva h=100 mm",
        "hang_rail": "Nosač + šina za viseće elemente",
        "wall_anchor": "Tipl + vijak 8x80 mm",
        "cabinet_connector": "Spojnica korpusa + vijak",
        "plinth_clip": "Klipsa za soklu",
        "worktop_fix": "Vijak / ugaonik za radnu ploču",
        "anti_tip": "Anti-tip ugaonik / traka",
        "appliance_hookup": "Priključni set uređaja",
        "sliding_track":   "GTV klizni sistem kliznih vrata (set)",
        "hanging_rod":     "Okrugla šipka za vešanje O25mm + par nosača",
        "hinge_plate":     "GTV montažna ploča za šarku",
        "handle_screw":    "M4 x 50mm vijak za ručku",
    },
}

# Dubina utora za leđnu ploču (standardni EU/SRB frameless)
GROOVE_DEPTH: int = 8  # mm

# Konstante za sanduk fioke
_SLIDE_CLEARANCE_MM: float = 12.0   # oslobađanje po strani za klizač fioke [mm]
_DRAWER_SETBACK_MM:  float = 20.0   # sanduk kraći od dubine elementa [mm]
_DRAWER_BOX_REVEAL:  float = 4.0    # sanduk niži od fronta (propust/reveal) [mm]


def _round_cut(mm_val: float, step: float) -> float:
    if step <= 0:
        return float(mm_val)
    return round(float(mm_val) / step) * step


def _safe_int(val: Any, default: int = 0) -> int:
    try:
        return int(val)
    except Exception:
        return int(default)


def _position_from_deo(deo: str) -> str:
    d = str(deo or "").lower()
    if "leva" in d:
        return "LEVA"
    if "desna" in d:
        return "DESNA"
    if "dno" in d or "sokla" in d:
        return "DOLE"
    if "plafon" in d or "gornja" in d:
        return "GORE"
    if "leđ" in d or "ledj" in d or "ledn" in d:
        return "ZADNJA"
    if "front" in d or "vrata" in d:
        return "PREDNJA"
    if "srednja" in d:
        return "SREDINA"
    if "nosač" in d or "nosac" in d:
        return "GORE"
    return "-"


def _assembly_step_for(section: str, deo: str) -> str:
    s = str(section or "").lower()
    d = str(deo or "").lower()
    if s == "carcass":
        if "leva" in d or "desna" in d or ("strana" in d and "ugaon" in d):
            return "1"
        if "dno" in d or "plafon" in d or "gornja" in d:
            return "2"
        if "srednja" in d or "pregrada" in d or "polica" in d or "kutna" in d:
            return "3"
        return "2"
    if s == "backs":
        return "4"
    if s in ("fronts", "drawer_boxes"):
        return "5"
    if s in ("worktop", "plinth"):
        return "6"
    if s == "hardware":
        return "7"
    return "-"


def _section_prefix(section: str) -> str:
    return {
        "carcass": "C",
        "backs": "B",
        "fronts": "F",
        "drawer_boxes": "D",
        "worktop": "W",
        "plinth": "P",
        "hardware": "H",
    }.get(str(section or "").lower(), "X")


def _annotate_parts(df: pd.DataFrame, section: str) -> pd.DataFrame:
    """Adds production-friendly columns: PartCode, Pozicija, SklopKorak."""
    if df is None or df.empty:
        return df
    out = df.copy()
    pref = _section_prefix(section)
    counters: Dict[tuple, int] = {}
    part_codes: List[str] = []
    positions: List[str] = []
    steps: List[str] = []

    for _, row in out.iterrows():
        _rid = _safe_int(row.get("ID", 0), 0)
        mod_key = f"M{_rid}" if _rid > 0 else "M0"
        key = (mod_key, pref)
        counters[key] = counters.get(key, 0) + 1
        part_codes.append(f"{mod_key}-{pref}{counters[key]:02d}")
        _deo = str(row.get("Deo", row.get("Naziv", "")) or "")
        positions.append(_position_from_deo(_deo))
        steps.append(_assembly_step_for(section, _deo))

    out["PartCode"] = part_codes
    out["Pozicija"] = positions
    out["SklopKorak"] = steps
    return out


def _attach_wall_column(df: pd.DataFrame, modules: List[Dict[str, Any]]) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    wall_by_id = {
        int(m.get("id", 0)): f"Zid {str(m.get('wall_key', 'A') or 'A').upper()}"
        for m in (modules or [])
        if int(m.get("id", 0) or 0) > 0
    }

    def _wall_for_row(row: pd.Series) -> str:
        current = str(row.get("Zid", "") or "").strip()
        if current:
            return current
        rid = _safe_int(row.get("ID", 0), 0)
        if rid > 0:
            return wall_by_id.get(rid, "")
        modul = str(row.get("Modul", "") or "")
        if "Zid " in modul:
            idx = modul.find("Zid ")
            return modul[idx:].strip()
        return ""

    out["Zid"] = out.apply(_wall_for_row, axis=1)
    return out


def _wall_length_mm(kitchen: Dict[str, Any], wall_key: str = "A") -> int:
    wk = str(wall_key or "A").upper()
    wl_map = (kitchen or {}).get("wall_lengths_mm", {}) or {}
    if wk in wl_map:
        try:
            return int(wl_map[wk])
        except Exception:
            pass
    for _w in list(((kitchen or {}).get("walls", []) or [])):
        if str((_w or {}).get("key", "")).upper() == wk:
            try:
                return int((_w or {}).get("length_mm", 3000) or 3000)
            except Exception:
                return 3000
    return int((((kitchen or {}).get("wall", {}) or {}).get("length_mm", 3000)) or 3000)


def _purchase_worktop_length(required_mm: int, standard_lengths: List[int]) -> int:
    _std = sorted(int(x) for x in (standard_lengths or []) if int(x) > 0)
    if not _std:
        _std = [2000, 3000, 4000]
    for _cand in _std:
        if _cand >= int(required_mm):
            return int(_cand)
    _step = 1000
    return int(((int(required_mm) + _step - 1) // _step) * _step)


def _required_worktop_length_for_wall(
    wall_modules: List[Dict[str, Any]],
    *,
    reserve_mm: int = 0,
) -> int:
    if not wall_modules:
        return int(max(0, reserve_mm))
    xs = [int(m.get("x_mm", 0) or 0) for m in wall_modules]
    xe = [int(m.get("x_mm", 0) or 0) + int(m.get("w_mm", 0) or 0) for m in wall_modules]
    span_mm = max(xe) - min(xs)
    return int(max(0, span_mm + int(reserve_mm or 0)))


def _cooking_unit_partial_back_height_mm(
    module: Dict[str, Any],
    *,
    front_gap_mm: float,
    step_mm: float,
) -> float:
    params = dict((module or {}).get("params", {}) or {})
    h_mm = float((module or {}).get("h_mm", 0) or 0)
    if params.get("has_drawer", True):
        dh_list = params.get("drawer_heights", None)
        if dh_list and len(dh_list) > 0:
            _base_h = float(dh_list[0])
        else:
            _base_h = 130.0
    else:
        _base_h = min(150.0, max(120.0, h_mm * 0.18))
    _service_zone_h = max(120.0, _base_h)
    return float(_round_cut(_service_zone_h - 2 * float(front_gap_mm or 0), step_mm))


def _default_worktop_cutout(module: Dict[str, Any], wt_depth: float) -> Dict[str, Any] | None:
    _tid = str((module or {}).get("template_id", "")).upper()
    _x = int((module or {}).get("x_mm", 0) or 0)
    _w = int((module or {}).get("w_mm", 0) or 0)
    _params = dict((module or {}).get("params", {}) or {})
    if _tid == "SINK_BASE":
        _cw = max(400, min(_w - 80, 500))
        _cd = max(400, min(int(wt_depth) - 40, 480))
        _cw = int(_params.get("sink_cutout_width_mm", _cw) or _cw)
        _cd = int(_params.get("sink_cutout_depth_mm", _cd) or _cd)
        _cx = int(_params.get("sink_cutout_x_mm", int(_x + max((_w - _cw) / 2.0, 0))) or int(_x + max((_w - _cw) / 2.0, 0)))
        return {
            "type": "sink",
            "x": int(_cx),
            "width": int(_cw),
            "depth": int(_cd),
        }
    if _tid in {"BASE_COOKING_UNIT", "BASE_HOB"}:
        _cw = max(450, min(_w - 60, 560))
        _cd = max(400, min(int(wt_depth) - 40, 490))
        _cw = int(_params.get("hob_cutout_width_mm", _cw) or _cw)
        _cd = int(_params.get("hob_cutout_depth_mm", _cd) or _cd)
        _cx = int(_params.get("hob_cutout_x_mm", int(_x + max((_w - _cw) / 2.0, 0))) or int(_x + max((_w - _cw) / 2.0, 0)))
        return {
            "type": "hob",
            "x": int(_cx),
            "width": int(_cw),
            "depth": int(_cd),
        }
    return None


def _hardware_brand_from_profile(profile_key: str, mfg: Dict[str, Any]) -> str:
    brand = str((mfg or {}).get("hardware_brand", "")).strip().upper()
    if brand in HARDWARE_BRAND_CATALOGS:
        return brand
    p = str(profile_key or "").upper()
    if "HETTICH" in p:
        return "HETTICH"
    if "GTV" in p:
        return "GTV"
    return "BLUM"


def _hinges_per_door(door_h_mm: float) -> int:
    if door_h_mm >= 2000:  # vrata >= 2000mm trebaju 5 sarki
        return 5
    if door_h_mm > 1200:
        return 4
    if door_h_mm > 900:
        return 3
    return 2


def _module_spans_overlap(a: Dict[str, Any], b: Dict[str, Any]) -> bool:
    ax0 = float(a.get("x_mm", 0) or 0)
    ax1 = ax0 + float(a.get("w_mm", 0) or 0)
    bx0 = float(b.get("x_mm", 0) or 0)
    bx1 = bx0 + float(b.get("w_mm", 0) or 0)
    return max(ax0, bx0) < min(ax1, bx1)


def _manufacturing_warnings(
    modules: List[Dict[str, Any]],
    *,
    kitchen: Dict[str, Any] | None = None,
    wall_h_mm: float,
    foot_h_mm: float,
    front_gap_mm: float,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    _wall_len_map = (kitchen or {}).get("wall_lengths_mm", {}) or {}
    _room_walls = (((kitchen or {}).get("room", {}) or {}).get("walls", {}) or {})
    _valid_walls = {str(k).upper() for k in _wall_len_map.keys() if str(k).strip()}
    _valid_walls.update({str(k).upper() for k in _room_walls.keys() if str(k).strip()})
    if not _valid_walls:
        _valid_walls = {"A"}

    def _row(mid: int, zone: str, modul: str, code: str, opis: str, preporuka: str) -> Dict[str, Any]:
        return {
            "ID": mid,
            "TYPE": str(zone or "").upper(),
            "Modul": modul,
            "Kategorija": "warning",
            "Naziv": f"UPOZORENJE ({code})",
            "Tip / Šifra": "CHECK",
            "Kol.": 1,
            "Napomena": f"{opis} | Preporuka: {preporuka}",
        }

    # Global profile warning
    if front_gap_mm < 1.0 or front_gap_mm > 4.0:
        rows.append(_row(
            0, "GLOBAL", "Projekat",
            "FRONT_GAP",
            f"Front fuga je {front_gap_mm:.1f} mm (van preporuke 1.0–4.0 mm).",
            "Postavi front_gap na 2.0 mm za stabilnu montažu."
        ))

    # Per-module warnings
    for m in modules:
        mid = int(m.get("id", 0) or 0)
        zone = str(m.get("zone", "")).lower()
        lbl = str(m.get("label", "") or m.get("template_id", ""))
        w = float(m.get("w_mm", 0) or 0)
        h = float(m.get("h_mm", 0) or 0)
        d = float(m.get("d_mm", 0) or 0)
        tid = str(m.get("template_id", "")).upper()
        params = m.get("params", {}) or {}
        wk = str(m.get("wall_key", "A") or "A").upper()

        try:
            resolve_template(tid)
        except Exception:
            rows.append(_row(
                mid, zone, lbl, "MISSING_TEMPLATE",
                f"Element koristi nepoznat template_id '{tid}'.",
                "Proveri module_templates.json ili zameni element važećim template-om pre eksporta."
            ))
            continue

        if wk not in _valid_walls:
            rows.append(_row(
                mid, zone, lbl, "INVALID_WALL_REFERENCE",
                f"Element referiše na nepostojeći zid '{wk}'.",
                f"Postavi wall_key na jedan od važećih zidova: {', '.join(sorted(_valid_walls))}."
            ))

        if w <= 0 or h <= 0 or d <= 0:
            rows.append(_row(
                mid, zone, lbl, "INVALID_DIMENSIONS",
                f"Element ima nevažeće dimenzije w={w:.0f}, h={h:.0f}, d={d:.0f} mm.",
                "Proveri širinu, visinu i dubinu elementa pre eksporta."
            ))
            continue

        if zone in ("wall", "wall_upper") and d > 400:
            rows.append(_row(
                mid, zone, lbl, "WALL_DEPTH",
                f"Dubina visećeg elementa {d:.0f} mm je velika za ergonomiju.",
                "Preporuka je 300–380 mm dubine za gornje elemente."
            ))

        if zone == "base" and d < 500:
            rows.append(_row(
                mid, zone, lbl, "BASE_DEPTH",
                f"Dubina donjeg elementa {d:.0f} mm je mala za standardne uređaje/sudoperu.",
                "Preporuka je min 560 mm za kuhinjske donje module."
            ))

        if zone == "tall" and "FREESTANDING" not in tid and d < 500:
            rows.append(_row(
                mid, zone, lbl, "TALL_DEPTH",
                f"Dubina visokog elementa {d:.0f} mm je mala za stabilan kuhinjski korpus.",
                "Preporuka je min 560 mm za visoke elemente."
            ))

        if zone == "tall" and h > max(wall_h_mm - foot_h_mm - 20, 0):
            rows.append(_row(
                mid, zone, lbl, "TALL_HEIGHT",
                f"Visoki element {h:.0f} mm je blizu/iznad raspoložive visine.",
                "Smanji visinu ili proveri realnu visinu plafona i stopice."
            ))

        if zone == "wall_upper":
            _supporting_wall = next((
                mm for mm in modules
                if str(mm.get("zone", "")).lower() == "wall"
                and str(mm.get("wall_key", "A") or "A").upper() == wk
                and float(mm.get("x_mm", 0) or 0) <= float(m.get("x_mm", 0) or 0)
                and (float(mm.get("x_mm", 0) or 0) + float(mm.get("w_mm", 0) or 0))
                >= (float(m.get("x_mm", 0) or 0) + w)
            ), None)
            _supported = _supporting_wall is not None
            if _supported and _supporting_wall is not None:
                _zones = ((kitchen or {}).get("zones", {}) or {})
                _wall_gap = int((_zones.get("wall", {}) or {}).get("gap_from_base_mm", 0) or 0)
                if _wall_gap == 0:
                    _foot_mm = int(float((kitchen or {}).get("foot_height_mm", foot_h_mm) or foot_h_mm))
                    _base_h = int(float((kitchen or {}).get("base_korpus_h_mm", 720) or 720))
                    _wt = ((kitchen or {}).get("worktop", {}) or {})
                    _wt_thk_mm = int(round(float(_wt.get("thickness", 4.0) or 4.0) * 10.0))
                    _vgap = int(float((kitchen or {}).get("vertical_gap_mm", 600) or 600))
                    _wall_gap = _foot_mm + _base_h + _wt_thk_mm + _vgap
                _free_h = max(wall_h_mm - _wall_gap - float(_supporting_wall.get("h_mm", 0) or 0) - 5.0, 0.0)
                if h > (_free_h + 2.0):
                    rows.append(_row(
                        mid, zone, lbl, "WALL_UPPER_HEIGHT",
                        f"Drugi red gornjih elemenata je visok {h:.0f} mm, a raspoloziv prostor iznad nosivog gornjeg elementa je {_free_h:.0f} mm.",
                        "Smanji visinu wall_upper elementa ili proveri realan razmak do plafona."
                    ))
            if not _supported:
                rows.append(_row(
                    mid, zone, lbl, "WALL_UPPER_SUPPORT",
                    "Element drugog reda nema pun oslonac na gornjem elementu ispod sebe.",
                    "Poravnaj ga tako da cela širina leži na elementu ispod."
                ))

        if zone == "tall_top":
            _supporting_tall = next((
                mm for mm in modules
                if str(mm.get("zone", "")).lower() == "tall"
                and str(mm.get("wall_key", "A") or "A").upper() == wk
                and float(mm.get("x_mm", 0) or 0) <= float(m.get("x_mm", 0) or 0)
                and (float(mm.get("x_mm", 0) or 0) + float(mm.get("w_mm", 0) or 0))
                >= (float(m.get("x_mm", 0) or 0) + w)
            ), None)
            _supported = _supporting_tall is not None
            if _supported and _supporting_tall is not None:
                _free_h = max(wall_h_mm - foot_h_mm - float(_supporting_tall.get("h_mm", 0) or 0), 0.0)
                if h > (_free_h + 2.0):
                    rows.append(_row(
                        mid, zone, lbl, "TALL_TOP_HEIGHT",
                        f"Popuna iznad visokog je visoka {h:.0f} mm, a raspoloziv prostor iznad nosivog visokog elementa je {_free_h:.0f} mm.",
                        "Smanji visinu tall_top popune ili visinu visokog elementa ispod nje."
                    ))
            if not _supported:
                rows.append(_row(
                    mid, zone, lbl, "TALL_TOP_SUPPORT",
                    "Popuna iznad visokog nema pun oslonac na visokom elementu ispod sebe.",
                    "Poravnaj je tako da cela širina leži na elementu ispod."
                ))

        if "DRAWER" in tid and d < 450:
            rows.append(_row(
                mid, zone, lbl, "DRAWER_DEPTH",
                f"Fioke na dubini {d:.0f} mm mogu imati ograničen izbor klizača.",
                "Koristi kraće klizače (npr. 350/400) ili povećaj dubinu korpusa."
            ))

        if "LIFTUP" in tid and w > 1200:
            rows.append(_row(
                mid, zone, lbl, "LIFTUP_WIDTH",
                f"Lift-up front širine {w:.0f} mm može biti pretežak.",
                "Razdvoj front na 2 segmenta ili koristi jači/međupodupirač."
            ))

        if tid in {"BASE_DISHWASHER", "BASE_DISHWASHER_FREESTANDING"} and w < 600:
            rows.append(_row(
                mid, zone, lbl, "DISHWASHER_WIDTH",
                f"Širina za mašinu za sudove {w:.0f} mm je manja od standardnih 600 mm.",
                "Povećaj širinu na najmanje 600 mm."
            ))

        if tid in {"TALL_FRIDGE", "TALL_FRIDGE_FREEZER", "TALL_FRIDGE_FREESTANDING"} and w < 600:
            rows.append(_row(
                mid, zone, lbl, "FRIDGE_WIDTH",
                f"Širina za frižider {w:.0f} mm je manja od standardnih 600 mm.",
                "Povećaj širinu na najmanje 600 mm."
            ))

        if tid in {"BASE_COOKING_UNIT", "OVEN_HOB", "BASE_OVEN_HOB_FREESTANDING"} and w < 600:
            rows.append(_row(
                mid, zone, lbl, "COOKING_WIDTH",
                f"Širina za rernu/ploču {w:.0f} mm je manja od standardnih 600 mm.",
                "Povećaj širinu na najmanje 600 mm."
            ))

        if tid == "BASE_HOB" and w < 450:
            rows.append(_row(
                mid, zone, lbl, "HOB_WIDTH",
                f"Samostalna ploča za kuvanje širine {w:.0f} mm je manja od minimalnih 450 mm.",
                "Standardne ploče su 45 cm, 60 cm ili 90 cm — povećaj širinu."
            ))

        if tid in {"TALL_OVEN", "TALL_OVEN_MICRO"} and w < 600:
            rows.append(_row(
                mid, zone, lbl, "TALL_APPLIANCE_WIDTH",
                f"Širina visoke appliance kolone {w:.0f} mm je manja od 600 mm.",
                "Povećaj širinu na najmanje 600 mm."
            ))

        if "SINK" in tid and w < 600:
            rows.append(_row(
                mid, zone, lbl, "SINK_WIDTH",
                f"Sudoperski element širine {w:.0f} mm je suviše mali za stabilan laički workflow.",
                "Koristi najmanje 600 mm širine."
            ))

        if tid == "SINK_BASE":
            _cx = float(params.get("sink_cutout_x_mm", max((w - 500.0) / 2.0, 0.0)) or 0.0)
            _cw = float(params.get("sink_cutout_width_mm", max(400.0, min(w - 80.0, 500.0))) or 0.0)
            _cd = float(params.get("sink_cutout_depth_mm", max(400.0, min(d - 40.0, 480.0))) or 0.0)
            if _cx < 0 or _cw <= 0 or (_cx + _cw) > w or _cd <= 0 or _cd > d:
                rows.append(_row(
                    mid, zone, lbl, "CUTOUT_OUT_OF_BOUNDS",
                    f"Izrez za sudoperu izlazi van radne ploče/modula (X={_cx:.0f}, W={_cw:.0f}, D={_cd:.0f} mm; modul W={w:.0f}, D={d:.0f} mm).",
                    "Smanji izrez ili ga pomeri tako da ostane unutar širine i dubine modula."
                ))

        if tid in {"BASE_COOKING_UNIT", "BASE_HOB"}:
            _cx = float(params.get("hob_cutout_x_mm", max((w - 560.0) / 2.0, 0.0)) or 0.0)
            _cw = float(params.get("hob_cutout_width_mm", max(450.0, min(w - 60.0, 560.0))) or 0.0)
            _cd = float(params.get("hob_cutout_depth_mm", max(400.0, min(d - 40.0, 490.0))) or 0.0)
            if _cx < 0 or _cw <= 0 or (_cx + _cw) > w or _cd <= 0 or _cd > d:
                rows.append(_row(
                    mid, zone, lbl, "CUTOUT_OUT_OF_BOUNDS",
                    f"Izrez za ploču za kuvanje izlazi van radne ploče/modula (X={_cx:.0f}, W={_cw:.0f}, D={_cd:.0f} mm; modul W={w:.0f}, D={d:.0f} mm).",
                    "Smanji izrez ili ga pomeri tako da ostane unutar širine i dubine modula."
                ))

        if "CORNER" in tid and w < 800:
            rows.append(_row(
                mid, zone, lbl, "CORNER_WIDTH",
                f"Ugaoni element širine {w:.0f} mm je rizičan za stabilan raspored i pristup uglu.",
                "Povećaj širinu na najmanje 800 mm."
            ))

        if tid == "FILLER_PANEL" and w > 120:
            rows.append(_row(
                mid, zone, lbl, "FILLER_TOO_WIDE",
                f"Filer panel širine {w:.0f} mm je širok za standardnu popunu prostora.",
                "Razmotri pravi završni panel ili poseban uski modul umesto širokog filera."
            ))
        if "CORNER" in tid:
            _anchor = "desno" if wk == "A" else "levo"
            _wall_len_map = (kitchen or {}).get("wall_lengths_mm", {}) or {}
            _wall_len = float(_wall_len_map.get(wk, ((kitchen or {}).get("wall", {}) or {}).get("length_mm", 3000)) or 3000)
            _x0 = float(m.get("x_mm", 0) or 0)
            _x1 = _x0 + w
            _exp_left = 5.0
            _exp_right = 5.0
            if wk == "A":
                if abs((_wall_len - _exp_right) - _x1) > 5:
                    rows.append(_row(
                        mid, zone, lbl, "CORNER_POSITION",
                        "Ugaoni element nije naslonjen na unutrasnji ugao zida A.",
                        "Postavi ga kao poslednji element desno, uz unutrasnji ugao."
                    ))
            else:
                if abs(_x0 - _exp_left) > 5:
                    rows.append(_row(
                        mid, zone, lbl, "CORNER_POSITION",
                        f"Ugaoni element nije naslonjen na unutrasnji ugao zida {wk}.",
                        f"Postavi ga kao prvi element {_anchor}, uz unutrasnji ugao."
                    ))

        if tid in {"WALL_HOOD", "WALL_MICRO"} and w < 600:
            rows.append(_row(
                mid, zone, lbl, "WALL_APPLIANCE_WIDTH",
                f"Širina modula za napu/mikrotalasnu {w:.0f} mm je manja od 600 mm.",
                "Povećaj širinu na najmanje 600 mm."
            ))

        if tid in {"WALL_HOOD", "WALL_MICRO"} and d < 300:
            rows.append(_row(
                mid, zone, lbl, "WALL_APPLIANCE_DEPTH",
                f"Dubina modula za napu/mikrotalasnu {d:.0f} mm je mala za uređaj i ventilaciju.",
                "Povećaj dubinu na najmanje 300 mm."
            ))

        if tid in {"BASE_DISHWASHER_FREESTANDING", "BASE_OVEN_HOB_FREESTANDING"} and d < 580:
            rows.append(_row(
                mid, zone, lbl, "FREESTANDING_DEPTH",
                f"Dubina samostojećeg uređaja {d:.0f} mm je manja od praktičnog minimuma 580 mm.",
                "Povećaj dubinu na najmanje 580 mm."
            ))

        if tid == "TALL_FRIDGE_FREESTANDING" and d < 600:
            rows.append(_row(
                mid, zone, lbl, "FREESTANDING_FRIDGE_DEPTH",
                f"Dubina samostojećeg frižidera {d:.0f} mm je manja od 600 mm.",
                "Povećaj dubinu na najmanje 600 mm."
            ))

        if tid in {"TALL_FRIDGE", "TALL_FRIDGE_FREEZER", "TALL_OVEN", "TALL_OVEN_MICRO"} and d < 560:
            rows.append(_row(
                mid, zone, lbl, "TALL_APPLIANCE_DEPTH",
                f"Dubina visoke appliance kolone {d:.0f} mm je manja od 560 mm.",
                "Povećaj dubinu na najmanje 560 mm."
            ))

        if ("1DOOR" in tid or tid in {"BASE_1DOOR", "WALL_1DOOR", "WALL_NARROW", "BASE_NARROW"}) and w > 600:
            rows.append(_row(
                mid, zone, lbl, "SINGLE_DOOR_WIDTH",
                f"Jednokrilni front širine {w:.0f} mm može biti nestabilan i težak za podešavanje.",
                "Podeli element na 2 vrata ili smanji širinu."
            ))

        if ("1DOOR" in tid or tid in {"BASE_1DOOR", "WALL_1DOOR", "WALL_UPPER_1DOOR", "WALL_NARROW", "BASE_NARROW"}):
            handle_side = str(params.get("handle_side", "") or "").lower()
            left_clear = 5
            right_clear = 5
            wall = (kitchen or {}).get("wall", {}) or {}
            wall_len = float(wall.get("length_mm", 3000) or 3000)
            mfg = ((kitchen or {}).get("manufacturing", {}) or {})
            profile_key = mfg.get("profile", "EU_SRB")
            prof = MANUFACTURING_PROFILES.get(profile_key, MANUFACTURING_PROFILES.get("EU_SRB", {})) or {}
            left_clear = float(mfg.get("mounting_tolerance_left_mm", prof.get("mounting_tolerance_left_mm", 5)) or 5)
            right_clear = float(mfg.get("mounting_tolerance_right_mm", prof.get("mounting_tolerance_right_mm", 5)) or 5)
            x0 = float(m.get("x_mm", 0) or 0)
            x1 = x0 + w
            if handle_side in ("left", "right") and x0 <= left_clear + 30 and handle_side == "right":
                rows.append(_row(
                    mid, zone, lbl, "SIDE_WALL_DOOR",
                    "Jednokrilna vrata su preblizu levom zidu na strani sarke.",
                    "Promeni stranu rucke ili dodaj filer 30-50 mm uz levi zid."
                ))
            if handle_side in ("left", "right") and x1 >= wall_len - right_clear - 30 and handle_side == "left":
                rows.append(_row(
                    mid, zone, lbl, "SIDE_WALL_DOOR",
                    "Jednokrilna vrata su preblizu desnom zidu na strani sarke.",
                    "Promeni stranu rucke ili dodaj filer 30-50 mm uz desni zid."
                ))

        dh_list = params.get("drawer_heights") or []
        if dh_list:
            if any(float(x) < 80 for x in dh_list):
                rows.append(_row(
                    mid, zone, lbl, "DRAWER_FRONT_MIN",
                    "Bar jedna fioka je niža od 80 mm, što je rizično za front i rukovanje.",
                    "Povećaj najmanju fioku na barem 80 mm."
                ))
            total = sum(float(x) for x in dh_list)
            if total > h - 10:
                rows.append(_row(
                    mid, zone, lbl, "DRAWER_STACK",
                    f"Zbir visina fioka {total:.0f} mm je vrlo blizu visine modula {h:.0f} mm.",
                    "Ostavi tehničku rezervu 10–20 mm za fuge i frontove."
                ))

        if "DOOR_DRAWER" in tid:
            door_h = float(params.get("door_height", h * 0.72) or (h * 0.72))
            if door_h < 180:
                rows.append(_row(
                    mid, zone, lbl, "DOOR_DRAWER_DOOR_MIN",
                    f"Vrata u kombinaciji vrata + fioka imaju samo {door_h:.0f} mm visine.",
                    "Povećaj vrata na najmanje 180 mm."
                ))
            if door_h > h - 120:
                rows.append(_row(
                    mid, zone, lbl, "DOOR_DRAWER_DRAWER_MIN",
                    f"Vrata u kombinaciji vrata + fioka uzimaju skoro celu visinu modula ({door_h:.0f} mm).",
                    "Ostavi barem 120 mm za fioku i tehničke fuge."
                ))

        install_warns = params.get("installation_warnings") or []
        for iw in install_warns:
            rows.append(_row(
                mid, zone, lbl, "INSTALLATION_ZONE",
                str(iw),
                "Proveri pristup instalaciji i ostavi servisni prostor."
            ))

    # Worktop alignment validation: all base units on the same wall should share
    # the same effective top level before a continuous worktop is mounted.
    _base_by_wall: Dict[str, List[Dict[str, Any]]] = {}
    for m in modules:
        if str(m.get("zone", "")).lower() != "base":
            continue
        wk = str(m.get("wall_key", "A") or "A").upper()
        _base_by_wall.setdefault(wk, []).append(m)
    for wk, _wall_mods in _base_by_wall.items():
        if len(_wall_mods) < 2:
            continue
        _tops = []
        for _m in _wall_mods:
            _top = float(foot_h_mm) + float(_m.get("h_mm", 0) or 0)
            _tops.append((int(_m.get("id", 0) or 0), str(_m.get("label", "") or _m.get("template_id", "")), _top))
        _min_top = min(_t[2] for _t in _tops)
        _max_top = max(_t[2] for _t in _tops)
        if (_max_top - _min_top) > 2.0:
            _mods_txt = ", ".join(f"#{_mid} {_lbl}={_top:.0f}mm" for _mid, _lbl, _top in _tops)
            for _mid, _lbl, _top in _tops:
                rows.append(_row(
                    _mid, "base", _lbl, "BASE_ALIGNMENT_INVALID",
                    f"Donji elementi na zidu {wk} nisu u istoj ravni za kontinualnu radnu ploču. Top-level raspon je {_min_top:.0f}-{_max_top:.0f} mm. ({_mods_txt})",
                    "Poravnaj visine svih BASE elemenata pre radne ploče ili koristi poseban segment/raskid ploče."
                ))

    _wt = (kitchen or {}).get("worktop", {}) or {}
    _wt_enabled = bool(_wt.get("enabled", False))
    _wt_reserve = int(_wt.get("mounting_reserve_mm", _wt.get("reserve_mm", 20)) or 20)
    _wt_lengths = sorted(
        int(x) for x in (_wt.get("standard_lengths_mm", [2000, 3000, 4000]) or [2000, 3000, 4000])
        if int(x) > 0
    )
    if _wt_enabled and _wt_lengths:
        _max_stock = max(_wt_lengths)
        for wk, _wall_mods in _base_by_wall.items():
            if not _wall_mods:
                continue
            _required = _required_worktop_length_for_wall(
                _wall_mods,
                reserve_mm=_wt_reserve,
            )
            if _required > _max_stock:
                for _m in _wall_mods:
                    _mid = int(_m.get("id", 0) or 0)
                    _lbl = str(_m.get("label", "") or _m.get("template_id", ""))
                    rows.append(_row(
                        _mid, "base", _lbl, "WORKTOP_TOO_SHORT",
                        f"Potrebna dužina radne ploče za zid {wk} je {_required} mm, a najveća standardna nabavna dužina je {_max_stock} mm.",
                        "Dodaj spoj/raskid ploče, koristi duži komercijalni format ili naruči specijalnu radnu ploču."
                    ))

    # ── Out-of-bounds validacija — modul izlazi van duzine zida ──────────────────
    for m in modules:
        mid  = int(m.get("id", 0) or 0)
        zone = str(m.get("zone", "")).lower()
        lbl  = str(m.get("label", "") or m.get("template_id", ""))
        wk   = str(m.get("wall_key", "A") or "A").upper()
        w    = float(m.get("w_mm", 0) or 0)
        x0   = float(m.get("x_mm", 0) or 0)
        x1   = x0 + w
        _wl_map = (kitchen or {}).get("wall_lengths_mm", {}) or {}
        _wl_kt  = (kitchen or {}).get("wall", {}) or {}
        _wall_len = float(_wl_map.get(wk, _wl_kt.get("length_mm", 3000)) or 3000)
        _tol = 2.0  # dozvoljena tolerancija u mm (montažni zazor)
        if x0 < (0.0 - _tol):
            rows.append(_row(
                mid, zone, lbl, "MODULE_BEFORE_WALL_START",
                f"Element pocinje pre pocetka zida {wk}: x={x0:.0f}mm < 0mm.",
                "Pomeri element udesno tako da pocinje od nule ili ostavi samo planirani tehnicki zazor."
            ))
        if x1 > _wall_len + _tol:
            rows.append(_row(
                mid, zone, lbl, "MODULE_OUT_OF_BOUNDS",
                f"Element izlazi van zida {wk}: x={x0:.0f}mm + w={w:.0f}mm = {x1:.0f}mm > duzina zida {_wall_len:.0f}mm.",
                "Pomeri element ulevo ili smanjite mu sirinu da stane u raspored zida."
            ))

    # Overlap warnings in same zone + same wall_key
    for i in range(len(modules)):
        for j in range(i + 1, len(modules)):
            a = modules[i]
            b = modules[j]
            za = str(a.get("zone", "")).lower()
            zb = str(b.get("zone", "")).lower()
            if za != zb:
                continue
            wa = str(a.get("wall_key", "A")).upper()
            wb = str(b.get("wall_key", "A")).upper()
            if wa != wb:
                continue
            if _module_spans_overlap(a, b):
                amid = int(a.get("id", 0) or 0)
                albl = str(a.get("label", "") or a.get("template_id", ""))
                blbl = str(b.get("label", "") or b.get("template_id", ""))
                rows.append(_row(
                    amid, za, albl, "OVERLAP",
                    f"Elementi se preklapaju u zoni {za.upper()} na zidu {wa}: '{albl}' i '{blbl}'.",
                    "Pomeri elemente tako da im se X-rasponi ne preklapaju."
                ))

            atid = str(a.get("template_id", "")).upper()
            btid = str(b.get("template_id", "")).upper()
            ax0 = int(a.get("x_mm", 0) or 0)
            ax1 = ax0 + int(a.get("w_mm", 0) or 0) + int(a.get("gap_after_mm", 0) or 0)
            bx0 = int(b.get("x_mm", 0) or 0)
            bx1 = bx0 + int(b.get("w_mm", 0) or 0) + int(b.get("gap_after_mm", 0) or 0)
            touching = abs(ax1 - bx0) <= 2 or abs(bx1 - ax0) <= 2
            if touching and (("CORNER" in atid and "CORNER" not in btid) or ("CORNER" in btid and "CORNER" not in atid)):
                amid = int(a.get("id", 0) or 0)
                albl = str(a.get("label", "") or a.get("template_id", ""))
                blbl = str(b.get("label", "") or b.get("template_id", ""))
                rows.append(_row(
                    amid, za, albl, "CORNER_OPENING_CLEARANCE",
                    f"Ugaoni element i susedni modul nalezu bez servisnog razmaka: '{albl}' / '{blbl}'.",
                    "Proveri frontove i rucke u otvaranju; po potrebi dodaj filer ili tehnicki razmak."
                ))
                _single_door_tids = {"BASE_1DOOR", "WALL_1DOOR", "WALL_UPPER_1DOOR", "WALL_NARROW", "BASE_NARROW"}
                for _cur, _other, _cur_tid, _other_tid, _cur_lbl in (
                    (a, b, atid, btid, albl),
                    (b, a, btid, atid, blbl),
                ):
                    if "CORNER" in _cur_tid or "CORNER" not in _other_tid:
                        continue
                    if ("1DOOR" not in _cur_tid) and (_cur_tid not in _single_door_tids):
                        continue
                    _cur_x = int(_cur.get("x_mm", 0) or 0)
                    _other_x = int(_other.get("x_mm", 0) or 0)
                    _handle = str((_cur.get("params", {}) or {}).get("handle_side", "") or "").lower()
                    _opens_into_corner = (_handle == "left") if _cur_x < _other_x else (_handle == "right")
                    if _opens_into_corner:
                        rows.append(_row(
                            int(_cur.get("id", 0) or 0),
                            za,
                            _cur_lbl,
                            "CORNER_DOOR_SWING",
                            f"Jednokrilni sused '{_cur_lbl}' otvara vrata ka uglu pored ugaonog modula.",
                            "Promeni stranu rucke ili koristi dvokrilni/fiokar pored ugla."
                        ))
                    _max_corner_neighbor_w = 500 if za == "base" else 450
                    if int(_cur.get("w_mm", 0) or 0) > _max_corner_neighbor_w:
                        rows.append(_row(
                            int(_cur.get("id", 0) or 0),
                            za,
                            _cur_lbl,
                            "CORNER_FRONT_COLLISION",
                            f"Jednokrilni sused '{_cur_lbl}' je preširok uz ugaoni modul ({int(_cur.get('w_mm', 0) or 0)}mm).",
                            f"Koristi max {_max_corner_neighbor_w}mm, dvokrilni element, fiokar ili ostavi filer uz ugao."
                        ))
    return rows


def _kant_desc(n_edge_w: int, n_edge_h: int, edge_thk: float,
               orient: str = "") -> str:
    """
    Kreira opis kantovanih ivica.
      n_edge_w: broj ivica po W osi (za vertikalne: F/prednja; za horizontalne: L+R)
      n_edge_h: broj ivica po H osi (za vertikalne: T/gornja; za horizontalne: F/prednja)
      orient:   "vertikalna" | "horizontalna" — utiče na labele osa
    """
    if n_edge_w == 0 and n_edge_h == 0:
        return "-"
    parts = []
    if orient == "horizontalna":
        # Za dno/plafon: H os = dubina, prednja ivica = H strana → label "F"
        if n_edge_h == 1:
            parts.append("F")
        elif n_edge_h == 2:
            parts.append("F+B")
        if n_edge_w == 1:
            parts.append("L")
        elif n_edge_w == 2:
            parts.append("L+R")
    else:
        # Vertikalne ploče (stranice, frontovi): standardne labele
        if n_edge_h == 1:
            parts.append("T")
        elif n_edge_h == 2:
            parts.append("T+B")
        if n_edge_w == 1:
            parts.append("F")
        elif n_edge_w == 2:
            parts.append("L+R")
    return "+".join(parts) + f" (ABS {edge_thk:.1f}mm)"


def _carcass_piece(
    mid: int, zone: str, label: str, part: str,
    cut_w: float, cut_h: float,
    n_edge_w: int, n_edge_h: int,
    mat: str, thk: float, edge_thk: float, step: float,
    kol: int = 1, nap: str = "",
    orient: str = "",
    L1: int = 0, L2: int = 0, K1: int = 0, K2: int = 0,
    modul: str = "",   # jedinstven naziv modula za krojnu listu (#ID [zona] naziv)
) -> Dict[str, Any]:
    """
    Generiše jedan red carcass DataFrame-a sa ispravnim CUT/FIN i kantom.
      n_edge_w: broj ivica po širini (W) koje imaju kant
      n_edge_h: broj ivica po visini (H) koje imaju kant
      orient:   "vertikalna" | "horizontalna"
      L1/L2/K1/K2: boolean kant flags (za nesting software)
    CUT se zaokružuje na step; FIN = CUT - (n_edge * edge_thk)
    """
    cw = _round_cut(cut_w, step)
    ch = _round_cut(cut_h, step)
    # _round_cut primenjen i na FIN dimenzije (sprečava floating point greške)
    fw = _round_cut(cw - n_edge_w * edge_thk, step)
    fh = _round_cut(ch - n_edge_h * edge_thk, step)
    return {
        "ID": mid,
        "TYPE": zone.upper(),
        "Modul": modul if modul else label,
        "Element": label,
        "Deo": part,
        "Materijal": mat,
        "Deb.": thk,
        "Kol.": kol,
        "Kant": _kant_desc(n_edge_w, n_edge_h, edge_thk, orient=orient),
        "L1": L1, "L2": L2, "K1": K1, "K2": K2,
        "Orijentacija": orient,
        "Dužina [mm]": fw,   # FIN dimenzija (gotova mera)
        "Širina [mm]": fh,   # FIN dimenzija (gotova mera)
        "CUT_W [mm]": cw,    # sirova mera pre kanta
        "CUT_H [mm]": ch,    # sirova mera pre kanta
        "CUT (W×H)": f"{cw:.1f} × {ch:.1f}",
        "FIN (W×H)": f"{fw:.1f} × {fh:.1f}",
        "Smer goda": "V" if orient == "vertikalna" else ("H" if orient == "horizontalna" else "-"),
        "Napomena": nap,
    }


def _front_row(
    mid: int, zone: str, label: str, part_name: str,
    w: float, h: float,
    front_mat: str, front_thk: float,
    edge_thk: float = 2.0,
    step: float = 0.5,
    kol: int = 1,
    napomena: str = "",
    modul: str = "",   # jedinstven naziv modula za krojnu listu
    grain_dir: str = "V",  # smer goda fronta: V/H/N
) -> Dict[str, Any]:
    """Helper: pravi jedan red frontova. Frontovi imaju kant na sve 4 ivice (L1=L2=K1=K2=1)."""
    cw = _round_cut(w, step)
    ch = _round_cut(h, step)
    # _round_cut primenjen i na FIN dimenzije (sprečava floating point greške)
    fw = _round_cut(cw - 2 * edge_thk, step)
    fh = _round_cut(ch - 2 * edge_thk, step)
    _sg = grain_dir.upper() if grain_dir and grain_dir.upper() in ("H", "V", "N") else "V"
    return {
        "ID": mid,
        "TYPE": zone.upper(),
        "Modul": modul if modul else label,
        "Element": label,
        "Deo": part_name,
        "Materijal": front_mat,
        "Deb.": front_thk,
        "Kol.": kol,
        "Kant": _kant_desc(2, 2, edge_thk),  # sve 4 ivice = L+R + T+B
        "L1": 1, "L2": 1, "K1": 1, "K2": 1,
        "Orijentacija": "vertikalna",
        "Dužina [mm]": fw,   # FIN dimenzija (gotova mera)
        "Širina [mm]": fh,   # FIN dimenzija (gotova mera)
        "CUT_W [mm]": cw,
        "CUT_H [mm]": ch,
        "CUT (W×H)": f"{cw:.1f} × {ch:.1f}",
        "FIN (W×H)": f"{fw:.1f} × {fh:.1f}",
        "Smer goda": _sg,
        "Napomena": napomena,
    }


def _drawer_box_rows(
    mid: int, zone: str, label: str, modul: str,
    W: float, D: float,
    carcass_thk: float, edge_thk: float,
    carcass_mat: str, step: float,
    front_gap: float,
    drawer_heights: list,
    slide_mm: float = _SLIDE_CLEARANCE_MM,
    setback_mm: float = _DRAWER_SETBACK_MM,
    reveal_mm: float = _DRAWER_BOX_REVEAL,
) -> List[Dict[str, Any]]:
    """
    Generiše redove krojne liste za sanduke fioka (telo fioke, bez dekorativnog fronta).
    Sve ploče su iverica iste debljine kao korpus (carcass_thk).

    Konstrukcija sanduka (butt joint, sve iverica carcass_thk):
      • 2× Bočna ploča        : box_d × box_h  (prednja + gornja ivica kantovane)
      • 1× Prednja strana sand.: inner_w × box_h  (gornja ivica kantovana)
      • 1× Zadnja strana      : inner_w × (box_h − carcass_thk)  (bez kanta)
      • 1× Dno               : inner_w × inner_d  (prednja ivica kantovana)

    Dimenzije:
      box_w   = W − 2×carcass_thk − 2×slide_mm   (spoljna širina sanduka)
      box_d   = D − setback_mm                    (dubina sanduka)
      box_h   = dh − 2×front_gap − reveal_mm      (visina sanduka po fioci)
      inner_w = box_w − 2×carcass_thk             (unutrašnja širina)
      inner_d = box_d − 2×carcass_thk             (unutrašnja dubina dna)
    """
    rows: List[Dict[str, Any]] = []

    if not drawer_heights or len(drawer_heights) == 0:
        return rows

    box_w   = W - 2 * carcass_thk - 2 * slide_mm
    box_d   = D - setback_mm
    inner_w = box_w - 2 * carcass_thk
    inner_d = box_d - 2 * carcass_thk

    if box_w <= 0 or box_d <= 0 or inner_w <= 0 or inner_d <= 0:
        return rows  # dimenzije nevalidne za ovu konfiguraciju

    for i, dh in enumerate(drawer_heights):
        box_h = float(dh) - 2 * front_gap - reveal_mm
        if box_h <= 0:
            continue

        # ── 2× Bočna ploča: box_d (dubina) × box_h ──────────────────────
        # Vertikalna; prednja ivica (L1) + gornja ivica (K1) kantovane
        rows.append(_carcass_piece(
            mid, zone, label, f"F{i + 1} — Bočna ploča",
            cut_w=box_d, cut_h=box_h,
            n_edge_w=1, n_edge_h=1,
            mat=carcass_mat, thk=carcass_thk, edge_thk=edge_thk, step=step,
            kol=2,
            orient="vertikalna",
            L1=1, L2=0, K1=1, K2=0,
            modul=modul,
        ))

        # ── 1× Prednja strana sanduka: inner_w × box_h ───────────────────
        # Vertikalna; samo gornja ivica (K1) kantovana
        rows.append(_carcass_piece(
            mid, zone, label, f"F{i + 1} — Prednja strana sanduka",
            cut_w=inner_w, cut_h=box_h,
            n_edge_w=0, n_edge_h=1,
            mat=carcass_mat, thk=carcass_thk, edge_thk=edge_thk, step=step,
            kol=1,
            orient="vertikalna",
            L1=0, L2=0, K1=1, K2=0,
            modul=modul,
        ))

        # ── 1× Zadnja strana: inner_w × (box_h − carcass_thk) ────────────
        # Vertikalna; bez kanta (dno proviruje ispod)
        zadnja_h = box_h - carcass_thk
        if zadnja_h > 0:
            rows.append(_carcass_piece(
                mid, zone, label, f"F{i + 1} — Zadnja strana sanduka",
                cut_w=inner_w, cut_h=zadnja_h,
                n_edge_w=0, n_edge_h=0,
                mat=carcass_mat, thk=carcass_thk, edge_thk=edge_thk, step=step,
                kol=1,
                orient="vertikalna",
                L1=0, L2=0, K1=0, K2=0,
                modul=modul,
            ))

        # ── 1× Dno: inner_w × inner_d ────────────────────────────────────
        # Horizontalna; prednja ivica kantovana (n_eh=1 → "F" u horiz. orijentaciji)
        rows.append(_carcass_piece(
            mid, zone, label, f"F{i + 1} — Dno sanduka",
            cut_w=inner_w, cut_h=inner_d,
            n_edge_w=0, n_edge_h=1,
            mat=carcass_mat, thk=carcass_thk, edge_thk=edge_thk, step=step,
            kol=1,
            orient="horizontalna",
            L1=1, L2=0, K1=0, K2=0,
            modul=modul,
        ))

    return rows


def _validate_modules(
    modules: List[Dict[str, Any]],
    carcass_thk: float,
    front_gap: float,
) -> List[tuple]:
    """
    Validira module pre generisanja krojne liste.
    Vraca listu grešaka: [(mid, poruka), ...]
    """
    errors: List[tuple] = []
    for inst in modules:
        mid = inst.get("id", "?")
        w = float(inst.get("w_mm", 0))
        h = float(inst.get("h_mm", 0))
        d = float(inst.get("d_mm", 0))
        params = inst.get("params", {}) or {}

        if w <= 0 or h <= 0 or d <= 0:
            errors.append((mid, f"Dimenzija <= 0: w={w}, h={h}, d={d}"))
            continue
        if w - 2 * carcass_thk <= 0:
            errors.append((mid, f"Unutrašnja širina {w - 2*carcass_thk:.0f}mm <= 0"))
        if d - carcass_thk <= 0:
            errors.append((mid, f"Unutrašnja dubina {d - carcass_thk:.0f}mm <= 0"))
        dh_list = params.get("drawer_heights", [])
        if dh_list:
            total = sum(float(x) for x in dh_list)
            if total > h:
                errors.append((mid, f"Zbir visina fioka {total:.0f}mm > visina modula {h:.0f}mm"))
            for i, dh in enumerate(dh_list):
                if float(dh) <= 0:
                    errors.append((mid, f"Visina fioke #{i+1} = {dh} <= 0"))
    return errors


def generate_cutlist(kitchen: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """
    Production-accurate krojna lista po CUTLIST SPEC v1.0 – SRB/EU.

    Vraca sekcije:
      - carcass  (stranice, dno, plafon — WALL/TALL imaju plafon, BASE NEMA)
      - backs    (leđne ploče — HDF/iverica, groove ili overlay)
      - fronts   (frontovi / vrata / fioke, kant 4×)
      - worktop  (radna ploča, 1 segment)
      - plinth   (sokla/lajsna)
    """
    modules = kitchen.get("modules", []) or []
    mats = kitchen.get("materials", {}) or {}

    # -------------------------------------------------------
    # Manufacturing profil
    # -------------------------------------------------------
    mfg = kitchen.get("manufacturing", {}) or {}
    profile_key = mfg.get("profile", "EU_SRB")
    prof = MANUFACTURING_PROFILES.get(profile_key, MANUFACTURING_PROFILES["EU_SRB"])
    hw_brand = _hardware_brand_from_profile(profile_key, mfg)
    hwc = HARDWARE_BRAND_CATALOGS.get(hw_brand, HARDWARE_BRAND_CATALOGS["BLUM"])
    front_gap  = float(prof.get("front_gap_mm", 2.0))
    step       = float(prof.get("cut_rounding_step_mm", 0.5))
    back_inset = float(prof.get("back_inset_mm", 10))

    # -------------------------------------------------------
    # Materijali
    # -------------------------------------------------------
    edge_thk    = float(mats.get("edge_abs_thk", 2))
    carcass_mat = mats.get("carcass_material", "Iverica")
    carcass_thk = float(mats.get("carcass_thk", 18))
    front_mat   = mats.get("front_material", "MDF")
    front_thk   = float(mats.get("front_thk", 18))
    back_thk    = float(mats.get("back_thk", 8))
    # Materijal leđne ploče: HDF za tanke (<=6mm), inače isti kao korpus
    back_mat    = "HDF" if back_thk <= 6 else carcass_mat
    # Groove ili overlay: ako back_inset > 0 → groove mode
    back_mode   = "groove" if back_inset > 0 else "overlay"

    # -------------------------------------------------------
    # Validacija
    # -------------------------------------------------------
    _errors = _validate_modules(modules, carcass_thk, front_gap)
    _error_map: Dict[int, str] = {mid: msg for mid, msg in _errors}

    # -------------------------------------------------------
    # Pomocne lambda za carcass
    # -------------------------------------------------------
    # _cp se poziva unutar for petlje — modul_label se prosledjuje eksplicitno pri svakom pozivu
    def _cp(mid, zone, label, part, cut_w, cut_h, n_ew, n_eh, kol=1, nap="",
            orient="", L1=0, L2=0, K1=0, K2=0, _mlbl=""):
        """Shortcut za _carcass_piece sa fiksnim mat/thk/edge/step."""
        err = _error_map.get(mid, "")
        return _carcass_piece(
            mid, zone, label, part, cut_w, cut_h, n_ew, n_eh,
            carcass_mat, carcass_thk, edge_thk, step, kol=kol,
            nap=f"GREŠKA: {err}" if err else nap,
            orient=orient, L1=L1, L2=L2, K1=K1, K2=K2,
            modul=_mlbl,
        )

    rows_carcass:       List[Dict[str, Any]] = []
    rows_backs:         List[Dict[str, Any]] = []
    rows_fronts:        List[Dict[str, Any]] = []
    rows_drawer_boxes:  List[Dict[str, Any]] = []
    _PANEL_ONLY_TIDS = {"FILLER_PANEL", "END_PANEL"}

    # -------------------------------------------------------
    # Loop kroz module
    # -------------------------------------------------------
    _DOOR_TIDS = {
        "BASE_1DOOR", "BASE_2DOOR", "BASE_DOORS",
        "WALL_1DOOR", "WALL_2DOOR", "WALL_DOORS",
        # WALL_LIFTUP je ISKLJUČEN iz opšteg skupa — obrađuje se posebno ispod
        # jer klapna uvek ide kao 1 panel (ili 2 za w>1000mm), bez dijeljenja po >500mm logici
        "WALL_UPPER_1DOOR", "WALL_UPPER_2DOOR",
        "TALL_DOORS", "TALL_TOP_DOORS",
    }
    _GLASS_TIDS = {"WALL_GLASS", "TALL_GLASS"}

    for inst in modules:
        mid   = int(inst.get("id", 0))
        zone  = str(inst.get("zone", "")).lower().strip()
        tid   = str(inst.get("template_id", "")).upper()
        label = inst.get("label", tid)
        params = inst.get("params", {}) or {}
        grain_dir = str(params.get("grain_dir", "V") or "V").upper()
        if grain_dir not in ("H", "V", "N"):
            grain_dir = "V"

        # Jedinstveni naziv modula za krojnu listu: "#ID [zona_abbr] naziv — WxH"
        # Npr: "#3 [G] Gornji (2 vrata) — 800×720"
        _zone_abbr = {"base": "D", "wall": "G", "wall_upper": "G2", "tall": "V", "tall_top": "VT"}.get(zone, zone.upper())
        _w_lbl = int(inst.get("w_mm", 0))
        _h_lbl = int(inst.get("h_mm", 0))
        modul_label = f"#{mid} [{_zone_abbr}] {label} — {_w_lbl}×{_h_lbl}"

        w = float(inst.get("w_mm", 0))
        h = float(inst.get("h_mm", 0))
        d = float(inst.get("d_mm", 0))
        _is_corner = "CORNER" in tid

        if w <= 0 or h <= 0 or d <= 0:
            continue

        if tid in _PANEL_ONLY_TIDS:
            if tid == "FILLER_PANEL":
                _pw = w
                _ph = h
                _deo = "Filer panel"
                _nap = "Panel-only element; širina popune × visina. Vidljiva dekorativna ploča."
            else:
                _pw = d
                _ph = h
                _deo = "Završna bočna ploča"
                _nap = "Panel-only element; dubina korpusa × visina. Vidljiva završna bočna strana."

            rows_carcass.append(_carcass_piece(
                mid, zone, label, _deo,
                _pw, _ph,
                1, 2,
                front_mat, front_thk, edge_thk, step,
                kol=1,
                nap=_nap,
                orient="vertikalna",
                L1=1, L2=0, K1=1, K2=1,
                modul=modul_label,
            ))
            continue

        # Unutrašnje dimenzije
        inner_w = w - 2 * carcass_thk   # širina dna/plafona
        inner_d = d - carcass_thk        # dubina dna/plafona (pozadi ulazi leđna ploča)

        # ===================================================
        # FAZA B: CARCASS
        # Stranice:  kant T+F (gornja+prednja) → n_ew=1(F), n_eh=1(T), vertikalna
        #            L1=1(prednja/dugačka), K1=1(gornja/kratka)
        # Dno:       kant F (prednja) → n_ew=0, n_eh=1(F), horizontalna
        #            L1=1(prednja ivica dugačke = frontalna strana)
        # Plafon:    isti kao Dno — samo WALL/TALL, BASE NEMA
        #
        # UREĐAJI BEZ SOPSTVENOG KORPUSA (frižider, napa...):
        # Ovi elementi su slobodnostojeći uređaji — ne pravimo im korpus od iverice.
        # Krojna lista za njih je prazna (CARCASS i BACKS se preskakuju).
        # ===================================================
        # DISHWASHER — ugradna MZS: nema bočnih ploča/dna/leđne ploče;
        #              jedina ploča je vezna letva (gornji vez) — dodaje se posebno
        _NO_CARCASS_TIDS = {
            "BASE_DISHWASHER",
            "BASE_DISHWASHER_FREESTANDING",
            "BASE_OVEN_HOB_FREESTANDING",
            "TALL_FRIDGE_FREESTANDING",
        }
        _skip_carcass = tid in _NO_CARCASS_TIDS

        if not _skip_carcass:
            if _is_corner:
                # ===================================================
                # UGAONI MODUL — L-oblik korpusa
                # Geometrija: oba kraka jednaka, sirina w, dubina d
                #   Krak 1 ide uz zid A (sirina w, dubina d)
                #   Krak 2 ide uz zid B (sirina w, dubina d, simetricno)
                # Delovi:
                #   Strana 1 + 2 (spoljna krajnja ploča): h × d  (kao standardna strana)
                #   Kutna vertikala (unutrašnji separator, 2×): h × d
                #   Dno Krak 1 (kvadratni ugao + produženje Kraka 1): (w-2t) × (d-t)
                #   Dno Krak 2 (produženje Kraka 2 van kvadrata): (d-2t) × (w-d-t)
                #   Za WALL zone: isti Plafon komadi kao Dno
                # ===================================================
                _corner_nap = "L-ugaoni korpus"
                # Spoljna strana 1 i 2 (krajnje bočne ploče ugaonog elementa)
                rows_carcass.append(_cp(mid, zone, label, "Strana 1 (ugaona)",
                                        d, h, n_ew=1, n_eh=1,
                                        orient="vertikalna", L1=1, L2=0, K1=1, K2=0,
                                        _mlbl=modul_label, nap=_corner_nap))
                rows_carcass.append(_cp(mid, zone, label, "Strana 2 (ugaona)",
                                        d, h, n_ew=1, n_eh=1,
                                        orient="vertikalna", L1=1, L2=0, K1=1, K2=0,
                                        _mlbl=modul_label, nap=_corner_nap))
                # Kutna vertikala — unutrašnji separator (2×, po jedna za svaki krak)
                rows_carcass.append(_cp(mid, zone, label, "Kutna vertikala",
                                        d, h, n_ew=1, n_eh=1,
                                        kol=2, orient="vertikalna", L1=1, L2=0, K1=1, K2=0,
                                        _mlbl=modul_label,
                                        nap=f"{_corner_nap}; unutrašnji pregradni zid (2×)"))
                # Dno Krak 1 — pokriva kvadratni ugao + produženje Kraka 1
                rows_carcass.append(_cp(mid, zone, label, "Dno — Krak 1",
                                        inner_w, inner_d, n_ew=0, n_eh=1,
                                        orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                        _mlbl=modul_label, nap=_corner_nap))
                # Dno Krak 2 — produženje drugog kraka van ugaonog kvadrata
                _dk2_w = d - 2 * carcass_thk
                _dk2_h = w - d - carcass_thk
                if _dk2_w > 0 and _dk2_h > 0:
                    rows_carcass.append(_cp(mid, zone, label, "Dno — Krak 2",
                                            _dk2_w, _dk2_h, n_ew=0, n_eh=1,
                                            orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                            _mlbl=modul_label,
                                            nap=f"{_corner_nap}; produženje drugog kraka"))
                # Plafon — SAMO za WALL zone (gornji ugaoni element)
                if zone in ("wall", "wall_upper", "tall", "tall_top"):
                    rows_carcass.append(_cp(mid, zone, label, "Plafon — Krak 1",
                                            inner_w, inner_d, n_ew=0, n_eh=1,
                                            orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                            _mlbl=modul_label, nap=_corner_nap))
                    if _dk2_w > 0 and _dk2_h > 0:
                        rows_carcass.append(_cp(mid, zone, label, "Plafon — Krak 2",
                                                _dk2_w, _dk2_h, n_ew=0, n_eh=1,
                                                orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                                _mlbl=modul_label, nap=_corner_nap))
            else:
                # Stranice (iste za sve zone)
                rows_carcass.append(_cp(mid, zone, label, "Leva strana",  d, h, n_ew=1, n_eh=1,
                                        orient="vertikalna",   L1=1, L2=0, K1=1, K2=0, _mlbl=modul_label))
                rows_carcass.append(_cp(mid, zone, label, "Desna strana", d, h, n_ew=1, n_eh=1,
                                        orient="vertikalna",   L1=1, L2=0, K1=1, K2=0, _mlbl=modul_label))
                # Dno — kant samo prednja (L1)
                rows_carcass.append(_cp(mid, zone, label, "Dno", inner_w, inner_d, n_ew=0, n_eh=1,
                                        orient="horizontalna", L1=1, L2=0, K1=0, K2=0, _mlbl=modul_label))
                # Plafon — SAMO za WALL, WALL_UPPER, TALL, TALL_TOP (BASE nema — tu je radna ploča)
                if zone in ("wall", "wall_upper", "tall", "tall_top"):
                    rows_carcass.append(_cp(mid, zone, label, "Plafon", inner_w, inner_d, n_ew=0, n_eh=1,
                                            orient="horizontalna", L1=1, L2=0, K1=0, K2=0, _mlbl=modul_label))
                # Srednja vertikalna ploča — za 2-vrata KRILNA elemente
                # Klizna vrata (SLIDING) nemaju srednju vertikalu — vrata klize po cijeloj širini
                if "2DOOR" in tid and "SLIDING" not in tid:
                    rows_carcass.append(_cp(mid, zone, label, "Srednja vertikala",
                                            d, h, n_ew=1, n_eh=1,
                                            orient="vertikalna", L1=1, L2=0, K1=1, K2=0,
                                            _mlbl=modul_label))

        # Police — jedinstvena logika:
        # - otvoreni/pantry moduli imaju pametne podrazumevane police
        # - standardni ormari dobijaju podrazumevane police
        # - korisnik uvek moze da override-uje preko params["n_shelves"]
        _n_sh = default_shelf_count(
            tid,
            zone=zone,
            h_mm=h,
            params=params,
            features={},
        )
        if not _skip_carcass and _n_sh > 0 and module_supports_adjustable_shelves(tid, features={}):
            _shelf_part = "Polica (podesiva)" if "n_shelves" in (params or {}) else "Polica"
            rows_carcass.append(_cp(mid, zone, label, _shelf_part,
                                    inner_w, inner_d, n_ew=0, n_eh=1,
                                    orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                    kol=_n_sh, _mlbl=modul_label))

        # Horizontalne pregrade za TALL aparat module (odvajaju zone: fioka / rerna / mikrovalna)
        # TALL_OVEN_MICRO mora biti provjeren PRIJE TALL_OVEN (sadrži isti substring)
        if not _skip_carcass:
            if "TALL_OVEN_MICRO" in tid:
                rows_carcass.append(_cp(mid, zone, label, "Horizontalna pregrada",
                                        inner_w, inner_d, n_ew=0, n_eh=1,
                                        orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                        kol=2, _mlbl=modul_label,
                                        nap="Pregrada između zona (rerna / mikrovalna)"))
            elif "TALL_OVEN" in tid:
                rows_carcass.append(_cp(mid, zone, label, "Horizontalna pregrada",
                                        inner_w, inner_d, n_ew=0, n_eh=1,
                                        orient="horizontalna", L1=1, L2=0, K1=0, K2=0,
                                        _mlbl=modul_label,
                                        nap="Pregrada između zona (iznad rerne)"))

        # ===================================================
        # FAZA C: LEĐNE PLOČE (backs)
        # groove: back_W = w - 2*GROOVE_DEPTH, back_H = h - GROOVE_DEPTH
        # overlay: back_W = w, back_H = h
        # Nema kanta; FIN = CUT
        # Uređaji bez korpusa (npr. MZS slot) nemaju ni leđnu ploču
        # ===================================================
        err_nap = f"GREŠKA: {_error_map[mid]}" if mid in _error_map else ""
        if not _skip_carcass:
            if _is_corner:
                # Ugaoni modul: 2 leđne ploče — po jedna za svaki krak produženja
                # Kvadratni ugaoni prostor (d×d) naslonjen je na zidove — leđna ploča tamo nije potrebna.
                # Svaki krak produženja ima sopstvenu leđnu ploču širine (w - d).
                _arm_bw_raw = w - d   # sirina krakova van ugaonog kvadrata
                if back_mode == "groove":
                    _bw_arm = _round_cut(_arm_bw_raw - 2 * GROOVE_DEPTH, step)
                    _bh_arm = _round_cut(h - GROOVE_DEPTH, step)
                    _back_nap_arm = f"utor {GROOVE_DEPTH}mm; leđna ploča ugaonog kraka"
                else:
                    _bw_arm = _round_cut(_arm_bw_raw, step)
                    _bh_arm = _round_cut(h, step)
                    _back_nap_arm = "overlay; leđna ploča ugaonog kraka"
                for _arm_lbl in ("Leđna ploča — Krak 1", "Leđna ploča — Krak 2"):
                    rows_backs.append({
                        "ID":          mid,
                        "TYPE":        zone.upper(),
                        "Modul":       modul_label,
                        "Element":     label,
                        "Deo":         _arm_lbl,
                        "Materijal":   back_mat,
                        "Deb.":        back_thk,
                        "Kol.":        1,
                        "Kant":        "-",
                        "L1": 0, "L2": 0, "K1": 0, "K2": 0,
                        "Orijentacija": "vertikalna",
                        "Dužina [mm]": _bw_arm,
                        "Širina [mm]": _bh_arm,
                        "CUT_W [mm]":  _bw_arm,
                        "CUT_H [mm]":  _bh_arm,
                        "CUT (W×H)":   f"{_bw_arm:.1f} × {_bh_arm:.1f}",
                        "FIN (W×H)":   f"{_bw_arm:.1f} × {_bh_arm:.1f}",
                        "Napomena":    err_nap if err_nap else _back_nap_arm,
                    })
            else:
                if back_mode == "groove":
                    bw = _round_cut(w - 2 * GROOVE_DEPTH, step)
                    bh = _round_cut(h - GROOVE_DEPTH, step)
                    back_nap = f"utor {GROOVE_DEPTH}mm"
                else:
                    bw = _round_cut(w, step)
                    bh = _round_cut(h, step)
                    back_nap = "overlay"
                _back_part_name = "Leđna ploča"
                if tid in {"BASE_COOKING_UNIT", "OVEN_HOB"}:
                    bh = _cooking_unit_partial_back_height_mm(inst, front_gap_mm=front_gap, step_mm=step)
                    _back_part_name = "Parcijalna leđna ploča"
                    back_nap = (
                        f"{back_nap}; parcijalna leđa samo u donjoj servisnoj zoni,"
                        " gornja zona iza rerne ostaje otvorena radi ventilacije i priključaka"
                    )
                if "HOOD" in tid:
                    back_nap = f"{back_nap}; ostaviti otvor/prolaz za odvod nape prema modelu uređaja"
                elif "MICRO" in tid:
                    back_nap = f"{back_nap}; ostaviti otvor za kabl i ventilaciju mikrotalasne"
                elif "SINK" in tid:
                    back_nap = f"{back_nap}; ostaviti otvor za dovod/odvod vode prema poziciji instalacija"
                rows_backs.append({
                    "ID":          mid,
                    "TYPE":        zone.upper(),
                    "Modul":       modul_label,
                    "Element":     label,
                    "Deo":         _back_part_name,
                    "Materijal":   back_mat,
                    "Deb.":        back_thk,
                    "Kol.":        1,
                    "Kant":        "-",
                    "L1": 0, "L2": 0, "K1": 0, "K2": 0,
                    "Orijentacija": "vertikalna",
                    "Dužina [mm]": bw,   # FIN = CUT (nema kanta)
                    "Širina [mm]": bh,   # FIN = CUT (nema kanta)
                    "CUT_W [mm]":  bw,
                    "CUT_H [mm]":  bh,
                    "CUT (W×H)":   f"{bw:.1f} × {bh:.1f}",
                    "FIN (W×H)":   f"{bw:.1f} × {bh:.1f}",
                    "Napomena":    err_nap if err_nap else back_nap,
                })

        # ===================================================
        # FAZA E: FRONTOVI (po tipu modula)
        # Frontovi uvek imaju kant 4× (sve 4 ivice)
        # fw = širina fronta = w - 2 * front_gap
        # ===================================================
        fw = w - 2 * front_gap

        def fr(part_name, fw_, fh_, kol_=1, nap_=""):
            """Shortcut za _front_row sa fiksnim mat/thk/edge/step."""
            return _front_row(mid, zone, label, part_name, fw_, fh_,
                              front_mat, front_thk, edge_thk, step, kol=kol_, napomena=nap_,
                              modul=modul_label, grain_dir=grain_dir)

        # ── Lift-up klapna (WALL_LIFTUP) ───────────────────────────────────────
        # Klapna se UVEK seca kao 1 panel pune širine (ne deli se po >500mm logici).
        # Za w ≤ 1000mm: 1 kom, 1 lift-up mehanizam (AVENTOS HK-S ili ekvivalent).
        # Za w > 1000mm: 2 panela (svaki pola), 2 mehanizma — usklađeno sa hw sekcijom.
        # Hardware sekcija (faza F, _pair_cnt) je već ispravna i NE menja se.
        if "LIFTUP" in tid:
            if w <= 1000:
                rows_fronts.append(fr(
                    "Front klapne", fw, h - 2 * front_gap, kol_=1,
                    nap_="Lift-up klapna — 1 panel pune širine; 1× lift-up mehanizam (AVENTOS HK-S ili ekviv.)"
                ))
            else:
                _liftup_half_w = (fw - front_gap) / 2
                rows_fronts.append(fr(
                    "Front klapne (polovina)", _liftup_half_w, h - 2 * front_gap, kol_=2,
                    nap_=(
                        f"Lift-up klapna š={w:.0f} mm > 1000 mm — front podeljen na 2 panela; "
                        "2× lift-up mehanizam (AVENTOS HK-S ili ekviv.)"
                    )
                ))

        # Vrata (1 ili 2 krila) — BASE, WALL, TALL_DOORS + ormarni ugaoni (CORNER)
        # WARDROBE+SLIDING i 2DOOR+SLIDING moraju biti isključeni — obrađuju se posebno ispod
        elif (tid in _DOOR_TIDS or "1DOOR" in tid
                or ("2DOOR" in tid and "SLIDING" not in tid)
                or ("WARDROBE" in tid and "CORNER" in tid and "DRAWERS" not in tid and "SLIDING" not in tid)):
            # Za TALL_DOORS: door_count iz params ima prioritet
            _p_door_count = int((params or {}).get("door_count", 0) or 0)
            if _p_door_count in (1, 2):
                num_doors = _p_door_count
            else:
                num_doors = 2 if (w > 500 and "1DOOR" not in tid) else 1
            door_w = (fw - front_gap) / num_doors if num_doors > 1 else fw
            door_h = h - 2 * front_gap
            rows_fronts.append(fr("Vrata", door_w, door_h, kol_=num_doors))

        # Uski bazni izvlačni modul — jedan front, unutra ide gotov uski izvlačni set
        elif "BASE_NARROW" in tid:
            rows_fronts.append(fr(
                "Front uskog izvlačnog modula", fw, h - 2 * front_gap, kol_=1,
                nap_="Uski modul za flaše/ulja/začine; unutrašnjost je kupovni izvlačni program"
            ))

        # Uski gornji zidni modul — standardna uska vrata
        elif "WALL_NARROW" in tid:
            rows_fronts.append(fr(
                "Vrata", fw, h - 2 * front_gap, kol_=1,
                nap_="Uski zidni element; proveriti smer otvaranja prema položaju na zidu"
            ))

        # Kuhinjski ugaoni elementi — L-korpus; front zavisi od mehanizma (lazy Susan / blind-corner)
        elif "CORNER" in tid:
            corner_front_w = fw
            corner_front_h = h - 2 * front_gap
            _is_diag_corner = "DIAGONAL" in tid
            _corner_front_nap = (
                "Dijagonalni ugaoni element; 2 manja fronta za krilna vrata ugla — dimenziju uskladiti sa mehanizmom"
                if _is_diag_corner
                else "L-ugaoni element; 2 manja fronta ugla (lazy Susan okretna ili blind-corner pull-out) — dimenziju i mehanizam proveriti prema stvarnom rasporedu"
            )
            rows_fronts.append(fr(
                "Ugaoni front", corner_front_w, corner_front_h, kol_=1,
                nap_=_corner_front_nap
            ))

        # Američki plakar — klizna vrata (TALL_WARDROBE_AMERICAN)
        elif "WARDROBE_AMERICAN" in tid:
            # Broj kliznih panela: ~1 panel na svakih 700mm širine
            _ap_n = max(2, int(round(w / 700)))
            _ap_w = (fw - (_ap_n - 1) * front_gap) / _ap_n
            _ap_h = h - 2 * front_gap
            rows_fronts.append(fr(
                f"Klizna vrata — američki plakar", _ap_w, _ap_h, kol_=_ap_n,
                nap_=f"{_ap_n} klizna panela × {_ap_w:.0f}mm"
            ))

        # Klizni ormari — TALL_WARDROBE_2DOOR_SLIDING i TALL_WARDROBE_CORNER_SLIDING
        # Pažnja: NE smeju pasti u "2DOOR" granu iznad (imaju šarnirska vrata umesto kliznih)
        elif "WARDROBE" in tid and "SLIDING" in tid:
            _sl_n = max(2, int(round(w / 700)))
            _sl_w = (fw - (_sl_n - 1) * front_gap) / _sl_n
            _sl_h = h - 2 * front_gap
            rows_fronts.append(fr(
                "Klizna vrata — ormar", _sl_w, _sl_h, kol_=_sl_n,
                nap_=f"{_sl_n} klizna panela × {_sl_w:.0f}mm; klizni sistem po celoj širini"
            ))

        # Staklena vrata
        elif tid in _GLASS_TIDS or "GLASS" in tid:
            num_doors = 2 if w > 450 else 1
            door_w = (fw - front_gap) / num_doors if num_doors > 1 else fw
            door_h = h - 2 * front_gap
            rows_fronts.append(fr("Staklena vrata", door_w, door_h, kol_=num_doors, nap_="Staklo"))

        # Kombi: vrata + fioka (BASE_DOOR_DRAWER) — i vrata i fioka front
        # MORA biti PRIJE opšteg "DRAWER" check-a jer TID sadrži "DRAWER"!
        elif "DOOR_DRAWER" in tid:
            _default_door_h = min(550.0, max(180.0, h - 170.0))
            _raw_door_h = float(params.get("door_height", _default_door_h))
            # Validacija: door_height ne sme biti veci od h - 80mm (minimum za fioku)
            _min_drawer_h = 80.0
            if _raw_door_h > h - _min_drawer_h:
                _raw_door_h = _default_door_h
            _door_h_dd = _raw_door_h - 2 * front_gap
            rows_fronts.append(fr("Vrata", fw, max(1.0, _door_h_dd)))
            _dh_list_dd = params.get("drawer_heights")
            if _dh_list_dd and len(_dh_list_dd) > 0:
                for _i, _dh in enumerate(_dh_list_dd):
                    rows_fronts.append(fr(f"Front fioke {_i + 1}", fw,
                                         max(1.0, float(_dh) - 2 * front_gap)))
                _dh_list_box = list(_dh_list_dd)
            else:
                _fdh_dd = h - _raw_door_h - 2 * front_gap
                rows_fronts.append(fr("Front fioke", fw, max(1.0, _fdh_dd)))
                _dh_list_box = [max(1.0, h - _raw_door_h)]
            rows_drawer_boxes.extend(_drawer_box_rows(
                mid=mid, zone=zone, label=label, modul=modul_label,
                W=w, D=d,
                carcass_thk=carcass_thk, edge_thk=edge_thk,
                carcass_mat=carcass_mat, step=step,
                front_gap=front_gap,
                drawer_heights=_dh_list_box,
            ))

        # Fioke (BASE_DRAWERS, BASE_NARROW_DRAWERS, itd.)
        elif "DRAWER" in tid and "DOORS" not in tid and "OVEN" not in tid:
            drawer_heights = params.get("drawer_heights", None)
            if drawer_heights and len(drawer_heights) > 0:
                for _i, _dh in enumerate(drawer_heights):
                    fh_i = float(_dh) - 2 * front_gap
                    rows_fronts.append(fr(f"Front fioke {_i + 1}", fw, fh_i))
                _dh_list_box = list(drawer_heights)
            else:
                n_d = max(1, int(params.get("n_drawers", 3 if h > 600 else 2)))
                fh_uniform = (h / n_d) - 2 * front_gap
                rows_fronts.append(fr("Front fioke (unif.)", fw, fh_uniform, kol_=n_d))
                _dh_list_box = [float(h) / n_d] * n_d
            # Sanduk fioke — telo (iverica, sve dimenzije)
            if _dh_list_box:
                rows_drawer_boxes.extend(_drawer_box_rows(
                    mid=mid, zone=zone, label=label, modul=modul_label,
                    W=w, D=d,
                    carcass_thk=carcass_thk, edge_thk=edge_thk,
                    carcass_mat=carcass_mat, step=step,
                    front_gap=front_gap,
                    drawer_heights=_dh_list_box,
                ))

        # Kombinovani: vrata + fioka (BASE_DOORS_DRAWERS)
        elif "DOORS_DRAWERS" in tid:
            door_h = (h * 0.72) - 2 * front_gap
            rows_fronts.append(fr("Vrata", fw, door_h))
            drawer_h_dd = (h * 0.28) - 2 * front_gap
            rows_fronts.append(fr("Front fioke", fw, drawer_h_dd))

        # BASE_COOKING_UNIT — kuhinjska jedinica (rerna + HOB ploča za kuvanje)
        # Rerna IMA SOPSTVENI PANEL — ne generiše se ovde!
        # HOB (ploča) nije u krojnoj listi. SAMO front fioke (ako postoji).
        elif "COOKING_UNIT" in tid:
            if params.get("has_drawer", True):  # default: fioka je standard
                dh_list = params.get("drawer_heights", None)
                if dh_list and len(dh_list) > 0:
                    fh_drawer = float(dh_list[0]) - 2 * front_gap
                    _dh_list_box = [float(dh_list[0])]
                else:
                    fh_drawer = 130.0 - 2 * front_gap  # standardna visina fioke kuhinjske jedinice
                    _dh_list_box = [130.0]
                rows_fronts.append(fr("Front fioke (kuhinjska jedinica)", fw, fh_drawer,
                                      nap_="Fioka ispod rerne; rerna ima sopstveni panel"))
                rows_drawer_boxes.extend(_drawer_box_rows(
                    mid=mid, zone=zone, label=label, modul=modul_label,
                    W=w, D=d,
                    carcass_thk=carcass_thk, edge_thk=edge_thk,
                    carcass_mat=carcass_mat, step=step,
                    front_gap=front_gap,
                    drawer_heights=_dh_list_box,
                ))

        # BASE_OVEN_HOB (legacy) — isto kao COOKING_UNIT, zadrzano radi kompatibilnosti
        elif tid == "OVEN_HOB":
            if params.get("has_drawer", True):
                dh_list = params.get("drawer_heights", None)
                if dh_list and len(dh_list) > 0:
                    fh_drawer = float(dh_list[0]) - 2 * front_gap
                    _dh_list_box = [float(dh_list[0])]
                else:
                    fh_drawer = 150.0 - 2 * front_gap
                    _dh_list_box = [150.0]
                rows_fronts.append(fr("Front fioke (ispod rerne)", fw, fh_drawer,
                                      nap_="Fioka ispod rerne; rerna ima sopstveni panel"))
                rows_drawer_boxes.extend(_drawer_box_rows(
                    mid=mid, zone=zone, label=label, modul=modul_label,
                    W=w, D=d,
                    carcass_thk=carcass_thk, edge_thk=edge_thk,
                    carcass_mat=carcass_mat, step=step,
                    front_gap=front_gap,
                    drawer_heights=_dh_list_box,
                ))

        # Samostojeći šporet — nema fronta ni korpusa u ovoj aplikaciji
        elif tid == "BASE_OVEN_HOB_FREESTANDING":
            _LOG.debug("Skipping fronts for freestanding cooker module id=%s", mid)

        # BASE_HOB — ploča za kuvanje je uređaj na radnoj ploči; ispod ide standardni ormar sa vratima
        elif "HOB" in tid:
            num_d_hob = 2 if w > 500 else 1
            door_w_hob = (fw - front_gap) / num_d_hob if num_d_hob > 1 else fw
            door_h_hob = h - 2 * front_gap
            rows_fronts.append(fr(
                "Vrata (ispod ploče za kuvanje)", door_w_hob, door_h_hob, kol_=num_d_hob,
                nap_="Ploča za kuvanje se nabavlja posebno; ispod je standardni bazni ormar bez police"
            ))

        # BASE_TRASH — izvlačni sortirnik sa jednom frontom na mehanizmu
        elif "TRASH" in tid:
            rows_fronts.append(fr(
                "Front sortirnika", fw, h - 2 * front_gap, kol_=1,
                nap_="Jedna fronta na izvlačnom sistemu za kante"
            ))

        # TALL_OVEN — appliance kolona sa donjim servisnim frontom
        elif tid == "TALL_OVEN":
            _service_front_h = max(120.0, min(360.0, h * 0.18) - 2 * front_gap)
            rows_fronts.append(fr(
                "Donji servisni front", fw, _service_front_h, kol_=1,
                nap_="Kolona za ugradnu rernu; gornju appliance zonu zatvara sam uređaj, a ovaj front pokriva donji servisni / skladišni prostor"
            ))

        # TALL_OVEN_MICRO — appliance kolona sa donjim servisnim frontom
        elif tid == "TALL_OVEN_MICRO":
            _service_front_h = max(120.0, min(320.0, h * 0.16) - 2 * front_gap)
            rows_fronts.append(fr(
                "Donji servisni front", fw, _service_front_h, kol_=1,
                nap_="Kolona za ugradnu mikrotalasnu i rernu; appliance zone zatvaraju uređaji, a ovaj front pokriva donji servisni / skladišni prostor"
            ))

        # Rerna (BASE_OVEN) — ima vrata
        elif "OVEN" in tid:
            oven_door_h = h - 2 * front_gap
            rows_fronts.append(fr("Vrata rerne", fw, oven_door_h))

        # Aspirator/napa
        elif "HOOD" in tid:
            _LOG.debug("Skipping fronts for hood module id=%s", mid)

        # Sudopera — vrata ispod
        elif "SINK" in tid:
            door_h = h - 2 * front_gap
            num_d = 2 if w > 600 else 1
            door_w2 = (fw - front_gap) / num_d if num_d > 1 else fw
            rows_fronts.append(fr("Vrata (ispod sudopere)", door_w2, door_h, kol_=num_d))

        # ── Ugradna mašina za sudove ──────────────────────────────────────────
        # Nema sopstveni korpus — bočne ploče obezbeđuju susedni elementi.
        # Jedina korpusna ploča: VEZNA LETVA (gornji vez ispod radne ploče).
        # Opciono: integrisani front panel.
        elif tid == "BASE_DISHWASHER":
            _dish = dishwasher_installation_metrics(kitchen, inst)
            _front_h = float(_dish.get("dishwasher_front_height", 720))
            _lower_fill_h = float(_dish.get("dishwasher_lower_filler_height", 0))
            _platform_h = float(_dish.get("dishwasher_platform_height", 0))
            _raised = bool(_dish.get("dishwasher_raised_mode", False))
            _lower_fill_cut_h = max(0.0, _lower_fill_h - 2 * front_gap)
            _lower_fill_fin_h = max(0.0, _lower_fill_cut_h - 2 * edge_thk)

            # 1. Vezna letva — jedina ploča MZS slota
            #    Širina  = širina slota (w), dubina = 100mm, debljina = carcass_thk
            #    Kant    = samo prednja ivica (L1, frontalna strana)
            _VEZNA_H = 100.0   # standardna visina/dubina vezne letve [mm]
            rows_carcass.append(_cp(
                mid, zone, label,
                "Vezna letva — MZS (ugradna)",
                w,           # CUT_W = puna širina slota
                _VEZNA_H,    # CUT_H = dubina letve (front-back)
                n_ew=0, n_eh=1,   # 1 kant na H smeru = prednja (F) ivica
                orient="horizontalna",
                L1=1, L2=0, K1=0, K2=0,
                nap="Gornji vez ugradne MZS — prednja ivica kantovana; "
                    "oslanja se na susedne korpuse",
                _mlbl=modul_label,
            ))
            if _raised and _platform_h > 0:
                rows_carcass.append(_cp(
                    mid, zone, label,
                    "Postolje / nosač — MZS",
                    w,
                    max(100.0, d - 40.0),
                    n_ew=0, n_eh=0,
                    orient="horizontalna",
                    L1=0, L2=0, K1=0, K2=0,
                    nap=(
                        f"Konstruktivna platforma za podizanje ugradne MZS; "
                        f"visina podizanja = {_platform_h:.0f} mm. "
                        f"Sokla ostaje kontinuirana, maska zatvara prostor iznad sokle."
                    ),
                    _mlbl=modul_label,
                ))
            # 2. Front panel (integrisana MZS — vrata u ravni sa ostalim frontovima)
            rows_fronts.append(fr(
                "Front — mašina za sudove",
                fw, max(0.0, _front_h - 2 * front_gap),
                nap_=(
                    "Integrisani front ugradne MZS"
                    if not _raised else
                    f"Integrisani front ugradne MZS; raised mode, front ostaje standardne visine {_front_h:.0f} mm"
                ),
            ))
            if _raised and _lower_fill_h > 0:
                rows_fronts.append(fr(
                    "Donja maska — MZS",
                    fw, _lower_fill_cut_h,
                    nap_=(
                        f"Donja maska iznad sokle za raised dishwasher; "
                        f"gotova visina maske = {_lower_fill_fin_h:.0f} mm, "
                        f"CUT visina = {_lower_fill_cut_h:.0f} mm; "
                        f"sokla ostaje zasebna i kontinuirana."
                    ),
                ))

        # Pantry/ostava
        elif "PANTRY" in tid:
            num_d = 2 if w > 500 else 1
            door_w3 = (fw - front_gap) / num_d if num_d > 1 else fw
            door_h3 = h - 2 * front_gap
            rows_fronts.append(fr("Vrata", door_w3, door_h3, kol_=num_d))

        # Integrisani frižider — ima appliance front/frontove
        elif tid == "TALL_FRIDGE":
            rows_fronts.append(fr(
                "Front integrisanog frižidera", fw, h - 2 * front_gap, kol_=1,
                nap_="Front se vezuje na vrata ugradnog frižidera prema fabričkom setu"
            ))

        elif tid == "TALL_FRIDGE_FREEZER":
            upper_h = (h * 0.60) - 2 * front_gap
            lower_h = (h * 0.40) - 2 * front_gap
            rows_fronts.append(fr(
                "Gornji front frižidera", fw, upper_h, kol_=1,
                nap_="Gornji appliance front za integrisani frižider/zamrzivač"
            ))
            rows_fronts.append(fr(
                "Donji front zamrzivača", fw, lower_h, kol_=1,
                nap_="Donji appliance front za integrisani frižider/zamrzivač"
            ))

        # Samostojeći frižider — bez fronta
        elif tid == "TALL_FRIDGE_FREESTANDING":
            _LOG.debug("Skipping fronts for freestanding fridge module id=%s", mid)

        # Otvoreni elementi — nema fronta
        elif "OPEN" in tid:
            _LOG.debug("Skipping fronts for open module id=%s", mid)

    # -------------------------------------------------------
    # FAZA D: SOKLA (plinth)
    # -------------------------------------------------------
    rows_plinth: List[Dict[str, Any]] = []
    foot_h = int(kitchen.get("foot_height_mm", 100))
    base_mods = [m for m in modules if str(m.get("zone", "")).lower().strip() == "base"]
    _base_by_wall: Dict[str, List[Dict[str, Any]]] = {}
    for _bm in base_mods:
        _bwk = str(_bm.get("wall_key", "A") or "A").upper()
        _base_by_wall.setdefault(_bwk, []).append(_bm)
    if base_mods and foot_h > 0:
        for _bwk, _bmods in sorted(_base_by_wall.items()):
            x0 = min(int(m.get("x_mm", 0)) for m in _bmods)
            x1 = max(int(m.get("x_mm", 0)) + int(m.get("w_mm", 0)) for m in _bmods)
            rows_plinth.append({
                "ID":          "-",
                "Modul":       f"Sokla - Zid {_bwk}",
                "Zid":         f"Zid {_bwk}",
                "Deo":         "Sokla (lajsna)",
                "Materijal":   "MDF",
                "Deb. [mm]":   16,
                "Dužina [mm]": x1 - x0,
                "Širina [mm]":  foot_h,
                "CUT_W [mm]":   x1 - x0,
                "CUT_H [mm]":   foot_h,
                "Visina [mm]": foot_h,
                "Kol.":        1,
                "Napomena":    f"Segment zida {_bwk}, raspon {x0}–{x1}mm",
            })

    # -------------------------------------------------------
    # Radna ploca — kontinualni segment po geometriji zida
    # -------------------------------------------------------
    rows_worktop: List[Dict[str, Any]] = []
    wt = kitchen.get("worktop", {}) or {}
    wt_thk_mm = int(round(float(wt.get("thickness", 0.0)) * 10.0))
    if wt_thk_mm > 0:
        if base_mods:
            wt_depth = float(wt.get("width", 600.0))
            wt_reserve_mm = int(wt.get("mounting_reserve_mm", wt.get("reserve_mm", 20)) or 20)
            wt_front_overhang = int(wt.get("front_overhang_mm", 20) or 20)
            wt_field_cut = bool(wt.get("field_cut", True))
            wt_edge_protection = bool(wt.get("edge_protection", True))
            wt_edge_protection_type = str(wt.get("edge_protection_type", "silikon / vodootporni premaz") or "silikon / vodootporni premaz")
            wt_purchase_lengths = list(wt.get("standard_lengths_mm", [2000, 3000, 4000]) or [2000, 3000, 4000])
            _is_l_kitchen = str((kitchen or {}).get("kitchen_layout", "")) == "l_oblik"
            for _bwk, _bmods in sorted(_base_by_wall.items()):
                xs = [int(m.get("x_mm", 0)) for m in _bmods]
                xe = [int(m.get("x_mm", 0)) + int(m.get("w_mm", 0)) for m in _bmods]
                _base_zone_w = max(xe) - min(xs)
                _wall_len = _wall_length_mm(kitchen, _bwk)
                _required_len = _required_worktop_length_for_wall(
                    _bmods,
                    reserve_mm=wt_reserve_mm,
                )
                _purchase_len = _purchase_worktop_length(_required_len, wt_purchase_lengths)
                _joint_type = str(wt.get("joint_type", "STRAIGHT") or "STRAIGHT").upper()
                _cutout_specs: List[Dict[str, Any]] = []
                for _m in _bmods:
                    _spec = _default_worktop_cutout(_m, wt_depth)
                    if _spec:
                        _cutout_specs.append(_spec)
                _joint_desc = ""
                if _is_l_kitchen:
                    _joint_desc = f" Spoj u uglu: {_joint_type}."
                _cutout_desc = ""
                _cutout_export = ""
                if _cutout_specs:
                    _cutout_parts = []
                    for _c in _cutout_specs:
                        _type_lbl = "sudopera" if _c["type"] == "sink" else "ploča za kuvanje"
                        _cutout_parts.append(
                            f"{_type_lbl} [X={int(_c['x'])} mm, W={int(_c['width'])} mm, D={int(_c['depth'])} mm]"
                        )
                    _cutout_export = "; ".join(_cutout_parts)
                    _cutout_desc = " Izrezi na licu mesta: " + _cutout_export + "."
                _protection_desc = ""
                if wt_edge_protection:
                    _protection_desc = f" Zaštita isečenih ivica: {wt_edge_protection_type}."
                _field_cut_desc = " Sečenje na licu mesta." if wt_field_cut else ""
                _wt_note = (
                    f"Kontinualni segment po zidu {_bwk}. "
                    f"Geometrija zida: {_wall_len} mm; base oslonac: {_base_zone_w} mm. "
                    f"Potrebna dužina za ugradnju: {_required_len} mm "
                    f"(rezerva +{wt_reserve_mm} mm). "
                    f"Nabavna dužina: {_purchase_len} mm. "
                    f"Prepust napred: {wt_front_overhang} mm."
                    f"{_field_cut_desc}{_protection_desc}{_cutout_desc}{_joint_desc}"
                )
                rows_worktop.append({
                    "ID":           "-",
                    "Modul":        f"Radna ploča - Zid {_bwk}",
                    "Zid":          f"Zid {_bwk}",
                    "Deo":          "Radna ploča",
                    "Materijal":    wt.get("material", "Radna ploča"),
                    "Deb.":         wt_thk_mm,
                    "Kol.":         1,
                    "Kant":         "-",
                    "L1": 1, "L2": 0, "K1": 0, "K2": 0,
                    "Orijentacija": "horizontalna",
                    "Dužina [mm]":  float(_required_len),
                    "Širina [mm]":  float(wt_depth),
                    "CUT_W [mm]":   float(_required_len),
                    "CUT_H [mm]":   float(wt_depth),
                    "Required length [mm]": float(_required_len),
                    "Purchase length [mm]": float(_purchase_len),
                    "Wall length [mm]": float(_wall_len),
                    "Field cut": "TRUE" if wt_field_cut else "FALSE",
                    "Joint type": _joint_type if _is_l_kitchen else "",
                    "Front overhang [mm]": float(wt_front_overhang),
                    "Edge protection": "TRUE" if wt_edge_protection else "FALSE",
                    "Edge protection type": wt_edge_protection_type if wt_edge_protection else "",
                    "Cutouts": _cutout_export,
                    "Napomena":     _wt_note,
                })
            # Nosači radne ploče: 2 daske po BASE elementu (100mm × širina_elementa × carcass_thk)
            # Služe za pričvršćivanje radne ploče na gornji deo korpusa
            # DISHWASHER — nema korpus; gornji vez (vezna letva) dodat posebno u carcass
            _NO_NOSAC_KW = ("DISHWASHER",)
            for _bm in base_mods:
                _bm_tid = str(_bm.get("template_id", "")).upper()
                if any(_kw in _bm_tid for _kw in _NO_NOSAC_KW):
                    continue
                _bw = int(_bm.get("w_mm", 600))
                _blbl = str(_bm.get("label", ""))
                _nosac_w = float(_bw) - 2 * carcass_thk  # unutrašnja širina
                _nosac_h = 96.0  # standardna visina nosača radne ploče
                rows_worktop.append({
                    "ID":           int(_bm.get("id", 0)),
                    "Modul":        _blbl,
                    "Deo":          "Nosač radne ploče",
                    "Materijal":    carcass_mat,
                    "Deb.":         carcass_thk,
                    "Kol.":         2,
                    "Kant":         "F (ABS 0.5mm)",
                    "L1": 1, "L2": 0, "K1": 0, "K2": 0,
                    "Orijentacija": "horizontalna",
                    "Dužina [mm]":  _nosac_w,
                    "Širina [mm]":  _nosac_h,
                    "CUT_W [mm]":   _nosac_w + edge_thk,
                    "CUT_H [mm]":   _nosac_h,
                    "Napomena":     "Gornji vez — pričvršćuje radnu ploču (prednja ivica kantovana)",
                })

    # -------------------------------------------------------
    # FAZA F: OKOVI (šarke, klizači, liftup mehanizmi)
    # Kalkulacija po tipu elementa i visini vrata:
    #   ≤ 900mm  → 2 šarke/vrata
    #   901–1200 → 3 šarke/vrata
    #   > 1200   → 4 šarke/vrata
    # -------------------------------------------------------
    rows_hardware: List[Dict[str, Any]] = []

    def _hw(
        mid_: int, zone_: str, lbl_: str, naziv: str, tip: str, kol: int,
        nap: str = "", kategorija: str = "okov",
    ) -> Dict[str, Any]:
        return {
            "ID":        mid_,
            "TYPE":      zone_.upper(),
            "Modul":     lbl_,
            "Kategorija": kategorija,
            "Naziv":     naziv,
            "Tip / Šifra": tip,
            "Kol.":      kol,
            "Napomena":  nap,
        }

    # Aparati bez okova — preskačemo
    _NO_HW = ()

    for _hm in modules:
        _hmid   = int(_hm.get("id", 0))
        _hzone  = str(_hm.get("zone", "base")).lower()
        _hlbl   = str(_hm.get("label", ""))
        _htid   = str(_hm.get("template_id", "") or "").upper()
        _hh     = float(_hm.get("h_mm", 720))
        _hw_mm  = float(_hm.get("w_mm", 600))
        _hp     = _hm.get("params") or {}
        _hmlbl  = _hlbl or _htid

        if _htid in _PANEL_ONLY_TIDS:
            _panel_qty = 4 if _htid == "FILLER_PANEL" else 6
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Vijak / spojnica za panel",
                "4x30 mm",
                _panel_qty,
                "Za pričvršćenje dekorativnog panela uz susedni korpus ili zidnu letvu",
                kategorija="potrosni",
            ))
            continue

        if any(_k in _htid for _k in _NO_HW):
            continue

        _is_liftup   = "LIFTUP" in _htid
        _is_drawer_m = "DRAWER" in _htid and "DOOR" not in _htid
        _is_trash_m  = "TRASH" in _htid
        _is_hob_m    = "HOB" in _htid
        _is_hood_m   = "HOOD" in _htid
        _is_micro_m  = "MICRO" in _htid
        _is_narrow_m = "BASE_NARROW" in _htid
        _is_wall_narrow_m = "WALL_NARROW" in _htid
        _is_glass_m  = "GLASS" in _htid
        _is_cooking_m = _htid in {"BASE_COOKING_UNIT", "OVEN_HOB"}
        _is_fridge_tall_m = _htid in {"TALL_FRIDGE", "TALL_FRIDGE_FREEZER"}
        _is_freestanding_fridge_m = _htid == "TALL_FRIDGE_FREESTANDING"
        _is_freestanding_dish_m = _htid == "BASE_DISHWASHER_FREESTANDING"
        _is_freestanding_cooker_m = _htid == "BASE_OVEN_HOB_FREESTANDING"
        _is_tall_oven_m = _htid in {"TALL_OVEN", "TALL_OVEN_MICRO"}
        _is_open_m   = any(_k in _htid for k in ("OPEN", "GLASS", "NARROW") for _k in [k])
        _is_door_m   = (
            (any(_k in _htid for _k in ("DOOR", "PANTRY", "SINK", "OVEN")) or _is_hob_m)
            and not _is_drawer_m
            and not _is_open_m
        )
        _is_dishw    = _htid == "BASE_DISHWASHER"
        _is_ddraw    = "DOORS_DRAWERS" in _htid  # kombi: vrata + fioka
        _is_wardrobe_sliding  = "WARDROBE" in _htid and "SLIDING" in _htid
        _is_wardrobe_american = _htid == "TALL_WARDROBE_AMERICAN"
        _is_wardrobe_hang     = _htid == "TALL_WARDROBE_INT_HANG"

        # ── Liftup mehanizam ────────────────────────────────
        if _is_liftup:
            _pair_cnt = 1 if _hw_mm <= 1000 else 2
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Lift-up mehanizam",
                hwc.get("liftup", ""),
                _pair_cnt,
                f"Liftup vrata š={_hw_mm:.0f} × v={_hh:.0f}mm",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Nosač / vijci za lift-up front",
                "Set za montažu fronta",
                1,
                "Prateći montažni set uz lift-up mehanizam",
                kategorija="potrosni",
            ))

        # ── Klizači za fioke ─────────────────────────────────
        elif _is_drawer_m:
            _n_dr = max(1, int(_hp.get("n_drawers", 3 if _hh > 600 else 2)))
            _dh_list = _hp.get("drawer_heights")
            if _dh_list:
                _n_dr = len(_dh_list)
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Klizač za fioku",
                hwc.get("slide", ""),
                _n_dr,
                f"{_n_dr} × 1 par klizača (levi + desni)",
            ))

        elif _is_trash_m:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Izvlačni mehanizam sortirnika",
                "Sortirnik / pull-out set",
                1,
                "Komplet klizača/nosača za kante za otpad",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Kante za otpad",
                "2x kanta + nosac",
                1,
                "Kupuje se kao gotov set; ne izrađuje se iz pločastog materijala",
            ))

        elif _is_cooking_m:
            if _hp.get("has_drawer", True):
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Klizač za fioku",
                    hwc.get("slide", ""),
                    1,
                    "1 fioka ispod rerne × 1 par klizača",
                ))
            if _hp.get("has_drawer", True):
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Nosač fronta fioke",
                    "Set nosača / ugaonika",
                    1,
                    "Za front fioke ispod rerne",
                ))

        elif _is_fridge_tall_m:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Ventilaciona rešetka / set",
                "Ulaz/izlaz vazduha",
                1,
                "Obezbediti propisan dovod i odvod vazduha za rashladni uređaj",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Set za vezu appliance fronta",
                "Klizni / door-on-door set",
                2 if _htid == "TALL_FRIDGE_FREEZER" else 1,
                "Set za spajanje fronta/frontova sa vratima integrisanog uređaja",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Anti-tip set",
                "Zidni ugaonik / traka",
                1,
                "Obavezno učvršćenje visoke kolone za zid zbog bezbednosti",
                kategorija="potrosni",
            ))

        elif _is_tall_oven_m:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Ventilacioni razmak / maska",
                "Distancer / ventilacioni set",
                1,
                "Za appliance kolonu i zaštitu od pregrevanja uređaja",
            ))

        elif "SINK" in _htid:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Sudopera",
                "Ugradna sudopera",
                1,
                "Kupuje se kao gotov proizvod; otvor se reže u radnoj ploči prema šablonu",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Sifon i odvodni set",
                "Sifon + preliv + spojnice",
                1,
                "Kupovni vodoinstalaterski set za sudoperu",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Slavina",
                "Baterija za sudoperu",
                1,
                "Kupuje se posebno prema izboru korisnika",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Sanitarni silikon",
                "Transparentni / sivi",
                1,
                "Za zaptivanje ruba sudopere i spojeva uz radnu ploču",
                kategorija="potrosni",
            ))
            # Šarke za vrata ispod sudopere — prethodno nedostajale jer SINK blok
            # hvata element pre nego što stigne do opšteg _is_door_m bloka.
            _n_vrata_sink = 2 if _hw_mm > 600 else 1
            _dh_sink = _hh - 4.0
            _sharke_sink = _hinges_per_door(_dh_sink)
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka (CLIP-top 110°)",
                hwc.get("hinge", ""),
                _n_vrata_sink * _sharke_sink,
                f"{_n_vrata_sink} vrata × {_sharke_sink} šarke (v vrata≈{_dh_sink:.0f}mm); "
                "bušenje: Ø35mm dubina 13mm, 22.5mm od ruba",
            ))

        elif _is_freestanding_dish_m:
            continue

        elif _is_freestanding_cooker_m:
            continue

        elif _is_freestanding_fridge_m:
            continue

        elif _is_narrow_m:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Uski izvlačni mehanizam",
                "Bottle pull-out / cargo set",
                1,
                "Gotov uski izvlačni program za flaše, ulja i začine",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Korpe / nosači uskog izvlačnog modula",
                "2 nivoa / set",
                1,
                "Kupuje se kao gotov set; broj i tip korpi proveriti prema širini modula",
            ))

        elif _is_wall_narrow_m:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka (CLIP-top 110°)",
                hwc.get("hinge", ""),
                2,
                "Uski zidni element sa jednim krilom",
            ))

        elif _is_glass_m:
            _n_vrata_gl = 2 if _hw_mm > 450 else 1
            _dh_gl = _hh - 4.0
            _sharke_gl = _hinges_per_door(_dh_gl)
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka za staklena vrata",
                hwc.get("hinge", ""),
                _n_vrata_gl * _sharke_gl,
                f"{_n_vrata_gl} staklena vrata × {_sharke_gl} šarki; proveriti tačan tip prema sistemu vitrine",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Magnetni zatvarač vitrine",
                "Magnet / odbojnik",
                1 if _n_vrata_gl == 1 else 2,
                "Za mekše naleganje i sigurno zatvaranje staklenih vrata",
            ))

        elif _is_wardrobe_sliding or _is_wardrobe_american:
            # ── Klizni sistem za orman (klizna vrata) ──────────────────────────
            _n_panels = 3 if _is_wardrobe_american else 2
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Klizni sistem vrata (gornja šina + donja vodilica)",
                hwc.get("sliding_track", "Klizni sistem kliznih vrata"),
                1,
                f"Komplet za {_n_panels} klizna krila; dužina šine = širina otvora + preklapanje",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Set točkića / nosača kliznih krila",
                "2 gornja + 2 donja točkića po krilu",
                _n_panels,
                f"{_n_panels} seta (1 set po krilu = 4 točkića)",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Gumeni stop / krajnji odbojnik",
                "Odbojnik za kraj šine",
                2,
                "Na oba kraja šine — sprečava udar krila na krajnjoj poziciji",
                kategorija="potrosni",
            ))

        elif "CORNER" in _htid:
            _hinge_qty = 2 if _hzone == "wall" else 2
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka za ugaoni front",
                hwc.get("hinge", ""),
                _hinge_qty,
                "Osnovni set za ugaoni front; tačan tip i ugao proveriti prema izvedbi",
            ))

        elif _is_hood_m:
            continue

        elif _is_micro_m:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Ventilacioni distancer / razmak",
                "Distancer / resetka",
                1,
                "Za sigurnu ventilaciju i provod kabla kod ugradne mikrotalasne",
                kategorija="potrosni",
            ))

        # ── Šarke za vrata ───────────────────────────────────
        elif _is_door_m:
            _n_vrata = 2 if (_hw_mm > 500 and "1DOOR" not in _htid) else 1
            # BASE_DOOR_DRAWER — uvek 1 vrata (fioka je zasebna, ne računa se kao vrata)
            if "DOOR_DRAWER" in _htid and "DOORS" not in _htid:
                _n_vrata = 1
            # Napomena: SINK ne može ovde stići (uhvati ga prethodni elif blok);
            # šarke za SINK su dodane direktno u SINK bloku gore.
            # Visina vrata ≈ visina modula (front gap ~4mm obostrano)
            _dh = _hh - 4.0
            _sharke_po = _hinges_per_door(_dh)
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka (CLIP-top 110°)",
                hwc.get("hinge", ""),
                _n_vrata * _sharke_po,
                f"{_n_vrata} vrata × {_sharke_po} šarke (v vrata={_dh:.0f}mm); bušenje: Ø35mm dubina 13mm, 22.5mm od ruba vrata",
            ))

        # ── Kombi: vrata + fioka ─────────────────────────────
        if _is_ddraw:
            # Šarke za donja vrata
            _dh2 = _hh * 0.72 - 4.0
            _sharke2 = _hinges_per_door(_dh2)
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka (CLIP-top 110°)",
                hwc.get("hinge", ""),
                _sharke2,
                f"Vrata donjeg dijela (v={_dh2:.0f}mm)",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Klizač za fioku",
                hwc.get("slide", ""),
                1,
                "1 fioka × 1 par klizača",
            ))

        # ── Klizač za fioku BASE_DOOR_DRAWER (šarke su dodane via _is_door_m) ──
        if "DOOR_DRAWER" in _htid:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Klizač za fioku",
                hwc.get("slide", ""),
                1,
                "1 fioka (vrata + fioka kombi)",
            ))

        # ── Šarke za integrisanu MZS (ugradnu mašinu za sudove) ─
        if _is_dishw:
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šarka za front MZS",
                hwc.get("dish_hinge", ""),
                2,
                "Integrisani front ugradne mašine za sudove",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Montažni set za front MZS",
                "Nosači + vijci + šablon",
                1,
                "Prateći set za vezu fronta sa vratima mašine",
                kategorija="potrosni",
            ))

        # ── Vešalica (šipka za vešanje odeće) — wardrobe hang sekcija ───────
        if _is_wardrobe_hang:
            _n_sip = int(_hp.get("hanger_sections", 1))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Šipka za vešanje odeće",
                hwc.get("hanging_rod", "Okrugla šipka O25mm + par nosača"),
                _n_sip,
                f"{_n_sip} šipka/e × 1 par nosača; dužina = širina sekcije - 5mm",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Nosači šipke za vešanje",
                "Ugradbeni nosač O25mm",
                _n_sip * 2,
                f"{_n_sip} šipke × 2 nosača = {_n_sip * 2} kom ukupno",
            ))

        # ── Ručke / pulls ────────────────────────────────────────────────
        # Svaki front (vrata, front fioke, liftup) dobija jednu ručku.
        # MZS i napa (bez kabinet fronta) ne dobijaju ručku.
        # BASE_HOB ima vrata ispod ploče za kuvanje — dobija ručke normalno (_is_hob_m se ne isključuje ovde).
        if not _skip_carcass and not _is_dishw and not _is_hood_m:
            _n_handles = 0
            if _is_liftup:
                _n_handles = 1
            elif _is_drawer_m:
                _n_dr_rucke = max(1, int(_hp.get("n_drawers", 3 if _hh > 600 else 2)))
                if _hp.get("drawer_heights"):
                    _n_dr_rucke = len(_hp.get("drawer_heights"))
                _n_handles = _n_dr_rucke
            elif _is_glass_m:
                _n_handles = 2 if _hw_mm > 450 else 1
            elif _is_wardrobe_sliding or _is_wardrobe_american:
                # Svako klizno krilo dobija 1 ručku (J-pull ili lajsna)
                _n_handles = 3 if _is_wardrobe_american else 2
            elif _is_door_m:
                if "DOOR_DRAWER" in _htid:
                    _n_handles = 1  # uvek 1 vrata (gornji deo kombinovanog elem.)
                else:
                    _n_handles = 2 if (_hw_mm > 500 and "1DOOR" not in _htid) else 1
                    if "SINK" in _htid:
                        _n_handles = 2 if _hw_mm > 600 else 1
            if _is_ddraw:
                _n_handles += 2  # donja vrata + front fioke
            elif "DOOR_DRAWER" in _htid:
                _n_handles += 1  # front fioke (vrata je vec u _is_door_m)
            if _n_handles > 0:
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Ručka / pull",
                    hwc.get("handle", "Po izboru korisnika"),
                    _n_handles,
                    f"{_n_handles} × 1 ručka — odabrati po projektu",
                ))

        # ── Bazni montažni potrošni materijal (po modulu) ─────────────────
        # Dodajemo za sve module koji imaju korpus (nije sam uređaj).
        if not _skip_carcass and not any(_kw in _htid for _kw in ("DISHWASHER",)):
            _n_shelves = default_shelf_count(
                _htid,
                zone=_hzone,
                h_mm=_hh,
                params=_hp,
                features={},
            )
            _conf_qty = 20 if _hzone == "tall" else 16
            _dowel_qty = 12 if _hzone == "tall" else 8
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Konfirmat vijak",
                hwc.get("confirmat", ""),
                _conf_qty,
                "Standardno za spoj korpusa (bočne+dno+plafon)",
                kategorija="potrosni",
            ))
            rows_hardware.append(_hw(
                _hmid, _hzone, _hmlbl,
                "Drvena tipla",
                hwc.get("dowel", ""),
                _dowel_qty,
                "Pomoćni spoj i poravnanje",
                kategorija="potrosni",
            ))
            if _n_shelves > 0:
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Nosač police (klin)",
                    hwc.get("shelf_pin", ""),
                    _n_shelves * 4,
                    f"{_n_shelves} polica × 4 klina",
                ))
            if _hzone in ("base", "tall"):
                _legs = 6 if _hw_mm >= 900 else 4
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Stopica (nogica)",
                    hwc.get("leg", ""),
                    _legs,
                    "Osnovni set po elementu",
                ))
            if _hzone in ("wall", "wall_upper"):
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Zidni nosac elementa",
                    hwc.get("hang_rail", ""),
                    1,
                    "Set za kačenje visećeg elementa",
                ))
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Anker za zid",
                    hwc.get("wall_anchor", ""),
                    2 if _hw_mm <= 800 else 3,
                    "Preporučena količina po modulu",
                    kategorija="potrosni",
                ))
            if _hzone == "tall" and not (_is_fridge_tall_m or _is_tall_oven_m or _is_freestanding_fridge_m):
                rows_hardware.append(_hw(
                    _hmid, _hzone, _hmlbl,
                    "Anti-tip set",
                    hwc.get("anti_tip", "Anti-tip ugaonik / traka"),
                    1,
                    "Preporučeno učvršćenje visokog elementa za zid zbog bezbednosti",
                    kategorija="potrosni",
                ))

    # Projektni potrosni materijal koji se ne vidi uvek na nivou jednog elementa.
    _mods_by_wall = {}
    for _m in modules:
        _wk = str(_m.get("wall_key", "A") or "A").upper()
        _mods_by_wall.setdefault(_wk, []).append(_m)

    _connector_pairs = 0
    for _wk, _mods in _mods_by_wall.items():
        for _zone_key in ("base", "wall", "wall_upper", "tall", "tall_top"):
            _zone_mods = [
                _m for _m in _mods
                if str(_m.get("zone", "")).lower() == _zone_key
                and "FREESTANDING" not in str(_m.get("template_id", "")).upper()
                and str(_m.get("template_id", "")).upper() != "BASE_DISHWASHER"
            ]
            _zone_mods.sort(key=lambda _m: (int(_m.get("x_mm", 0)), int(_m.get("id", 0))))
            for _i in range(len(_zone_mods) - 1):
                _a = _zone_mods[_i]
                _b = _zone_mods[_i + 1]
                _a_end = int(_a.get("x_mm", 0)) + int(_a.get("w_mm", 0)) + int(_a.get("gap_after_mm", 0))
                _b_x = int(_b.get("x_mm", 0))
                if abs(_a_end - _b_x) <= 2:
                    _connector_pairs += 1

    if _connector_pairs > 0:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Spojnica susednih korpusa",
            hwc.get("cabinet_connector", "Spojnica korpusa + vijak"),
            _connector_pairs * 2,
            "Osnovno 2 spojnice po spoju susednih korpusa u istoj zoni",
            kategorija="potrosni",
        ))

    _base_tall_mods = [
        _m for _m in modules
        if str(_m.get("zone", "")).lower() in ("base", "tall")
        and "FREESTANDING" not in str(_m.get("template_id", "")).upper()
        and str(_m.get("template_id", "")).upper() != "BASE_DISHWASHER"
    ]
    _leg_total = 0
    for _m in _base_tall_mods:
        _mw = float(_m.get("w_mm", 600) or 600)
        _leg_total += 6 if _mw >= 900 else 4
    if _leg_total > 0:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Klipsa za soklu",
            hwc.get("plinth_clip", "Klipsa za soklu"),
            _leg_total,
            "Po jedna klipsa po stopici za prihvat sokle",
            kategorija="potrosni",
        ))

    if rows_worktop:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Vijak / ugaonik za radnu ploču",
            hwc.get("worktop_fix", "Vijak / ugaonik za radnu ploču"),
            max(4, len(base_mods) * 2),
            "Osnovni set za pričvršćenje radne ploče na nosače i korpuse",
            kategorija="potrosni",
        ))
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Zaptivna lajsna / silikon uz zid",
            "Zidna lajsna ili sanitarni silikon",
            1,
            "Završno zaptivanje spoja radne ploče prema zidu po izboru korisnika",
            kategorija="potrosni",
        ))

    # Broji SAMO CLIP-top šarke — MZS šarke (Blum 79T9550) su drugačiji tip
    # i ne koriste iste vijke, zato se isključuju iz ovog zbira.
    _total_hinge_units = sum(
        int(_r.get("Kol.", 0) or 0)
        for _r in rows_hardware
        if ("Šarka" in str(_r.get("Naziv", "")) or "Sarka" in str(_r.get("Naziv", "")))
        and "MZS" not in str(_r.get("Naziv", ""))
    )
    if _total_hinge_units > 0:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Vijak za šarku",
            hwc.get("hinge_screw", "3.5x16 mm"),
            _total_hinge_units * 8,
            "Osnovno 8 vijaka po šarki ako nisu uključeni u set",
            kategorija="potrosni",
        ))

    _total_slide_pairs = sum(
        int(_r.get("Kol.", 0) or 0)
        for _r in rows_hardware
        if str(_r.get("Naziv", "")) == "Klizač za fioku"
    )
    if _total_slide_pairs > 0:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Vijak za klizač",
            hwc.get("slide_screw", "3.5x16 mm"),
            _total_slide_pairs * 12,
            "Osnovno 12 vijaka po paru klizača ako nisu uključeni u set",
            kategorija="potrosni",
        ))

    _back_fix_qty = 0
    for _row in rows_backs:
        try:
            _back_fix_qty += int(_row.get("Kol.", 0) or 0) * 20
        except Exception:
            continue
    if _back_fix_qty > 0:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Vijak / ekser za leđa",
            hwc.get("back_fix", "3x16 mm / ekser 1.4x25 mm"),
            _back_fix_qty,
            "Osnovno pričvršćenje leđa po korpusu; broj proveriti prema standardu radionice",
            kategorija="potrosni",
        ))

    _wall_mount_modules = [
        _m for _m in modules
        if str(_m.get("zone", "")).lower() in ("wall", "wall_upper", "tall")
        and "FREESTANDING" not in str(_m.get("template_id", "")).upper()
    ]
    if _wall_mount_modules:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Vijak za zidni nosač / šinu",
            hwc.get("wall_mount_screw", "5x60 mm"),
            len(_wall_mount_modules) * 4,
            "Za pričvršćenje nosača, šine ili anti-tip seta; tip vijka prilagoditi zidu",
            kategorija="potrosni",
        ))

    # ── Montažne ploče za šarke (samo BLUM CLIP-top sistem) ───────────────
    # Svaka šarka zahteva posebnu montažnu ploču (prodaje se odvojeno)
    _hinge_plate_brand = hwc.get("hinge_plate", "")
    if _hinge_plate_brand:
        # Broji SAMO CLIP-top šarke — MZS šarke koriste drugačiju montažnu ploču
        # i ne idu na CLIP-top sistem, zato se isključuju iz ovog zbira.
        _total_hinges_for_plate = sum(
            int(_r.get("Kol.", 0) or 0)
            for _r in rows_hardware
            if "arka" in str(_r.get("Naziv", ""))
            and "MZS" not in str(_r.get("Naziv", ""))
            and str(_r.get("ID", 0)) != "0"   # preskoci projektne redove
        )
        if _total_hinges_for_plate > 0:
            rows_hardware.append(_hw(
                0, "project", "Projekat",
                "Montažna ploča za šarku",
                _hinge_plate_brand,
                _total_hinges_for_plate,
                "1 montažna ploča po šarki (BLUM CLIP-top sistem: ploča + šarki odvojeni)",
            ))

    # ── Vijci za ručke ─────────────────────────────────────────────────────
    # Svaka ručka se montira sa 2 vijka M4 kroz vrata/front fioke.
    _total_handles_proj = sum(
        int(_r.get("Kol.", 0) or 0)
        for _r in rows_hardware
        if "pull" in str(_r.get("Naziv", "")).lower()
        and str(_r.get("ID", 0)) != "0"
    )
    if _total_handles_proj > 0:
        rows_hardware.append(_hw(
            0, "project", "Projekat",
            "Vijak za ručku",
            hwc.get("handle_screw", "M4 x 50mm vijak za ručku"),
            _total_handles_proj * 2,
            "2 vijka M4 po ručki; dužinu vijka prilagoditi debljini fronta",
            kategorija="potrosni",
        ))

    # Manufacturing warnings (dimenzije, preklapanja, montažni rizici)
    wall_h_mm = float((kitchen.get("wall", {}) or {}).get("height_mm", 2600) or 2600)
    foot_h_mm = float(kitchen.get("foot_height_mm", 100) or 100)
    rows_hardware.extend(_manufacturing_warnings(
        modules,
        kitchen=kitchen,
        wall_h_mm=wall_h_mm,
        foot_h_mm=foot_h_mm,
        front_gap_mm=front_gap,
    ))
    _sections = {
        "carcass":       pd.DataFrame(rows_carcass),
        "backs":         pd.DataFrame(rows_backs),
        "fronts":        pd.DataFrame(rows_fronts),
        "drawer_boxes":  pd.DataFrame(rows_drawer_boxes),
        "worktop":       pd.DataFrame(rows_worktop),
        "plinth":        pd.DataFrame(rows_plinth),
        "hardware":      pd.DataFrame(rows_hardware),
    }
    # Production-friendly numeracija i pozicije po svim sekcijama.
    for _skey, _sdf in list(_sections.items()):
        _sections[_skey] = _attach_wall_column(_annotate_parts(_sdf, _skey), modules)
    return _sections


# --- Compatibility: main.py у неким варијантама очекује овај назив ---
def build_cutlist_sections(kitchen: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """Compatibility wrapper: alias за generate_cutlist()."""
    return generate_cutlist(kitchen)


def build_project_header(kitchen: Dict[str, Any], lang: str = "sr") -> pd.DataFrame:
    """Project header for workshop / end-user packet."""
    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)
    wall = kitchen.get("wall", {}) or {}
    meta = kitchen.get("project", {}) or {}
    room = str(meta.get("room", "") or kitchen.get("room_name", "") or _t("Kuhinja", "Kitchen"))
    project_name = str(meta.get("name", "") or kitchen.get("project_name", "") or ("Krojna Lista PRO" if _lang == "sr" else "CabinetCut PRO"))
    version = str(meta.get("version", "") or kitchen.get("version", "") or "v1")
    customer = str(meta.get("customer", "") or kitchen.get("customer_name", "") or "-")
    wall_name = str(meta.get("wall_name", "") or wall.get("name", "") or _t("Zid A", "Wall A"))
    measured_by = str(meta.get("measured_by", "") or "-")
    designed_by = str(meta.get("designed_by", "") or "-")
    workshop_note = str(
        meta.get("workshop_note", "")
        or _t(
            "U servisu raditi sečenje i kantovanje po CUT merama. Otvore i posebne obrade proveriti po napomenama.",
            "In the workshop, do cutting and edging strictly by CUT dimensions. Verify openings and special machining against the notes.",
        )
    )
    if _lang not in {"sr", "en"}:
        room = str(_translate_export_text(room, _lang, "Vrednost") or room)
        project_name = str(_translate_export_text(project_name, _lang, "Vrednost") or project_name)
        wall_name = str(_translate_export_text(wall_name, _lang, "Wall") or wall_name)
        workshop_note = str(_translate_export_text(workshop_note, _lang, "Napomena za servis") or workshop_note)
    today = datetime.now().strftime("%d.%m.%Y %H:%M")

    return pd.DataFrame([
        {"Polje": _t("Projekat", "Project"), "Vrednost": project_name},
        {"Polje": _t("Kupac", "Customer"), "Vrednost": customer},
        {"Polje": _t("Prostorija", "Room"), "Vrednost": room},
        {"Polje": _t("Zid", "Wall"), "Vrednost": wall_name},
        {"Polje": _t("Mera zida", "Wall size"), "Vrednost": f"{wall.get('length_mm', 0)} x {wall.get('height_mm', 0)} mm"},
        {"Polje": _t("Verzija", "Version"), "Vrednost": version},
        {"Polje": _t("Generisano", "Generated"), "Vrednost": today},
        {"Polje": _t("Merio", "Measured by"), "Vrednost": measured_by},
        {"Polje": _t("Crtano", "Designed by"), "Vrednost": designed_by},
        {"Polje": _t("Napomena za servis", "Workshop note"), "Vrednost": workshop_note},
    ])


def _apply_ordered_replacements(txt: str, replacements: list[tuple[str, str]]) -> str:
    out = str(txt or "")
    for src, dst in sorted(replacements, key=lambda item: len(item[0]), reverse=True):
        out = out.replace(src, dst)
    return out


def _translate_export_text_ptbr(value: Any, column: str = "") -> str:
    txt = str(value or "")
    col = str(column or "")

    if col in {"Zid", "Wall"}:
        return txt.replace("Zid ", "Parede ").replace("Wall ", "Parede ")

    if col in {"Pozicija", "Position"}:
        mapping = {
            "LEVA": "ESQUERDA",
            "DESNA": "DIREITA",
            "LEVO": "ESQUERDA",
            "DESNO": "DIREITA",
            "GORE": "ACIMA",
            "DOLE": "ABAIXO",
            "CENTAR": "CENTRO",
            "SREDINA": "CENTRO",
            "PREDNJA": "FRENTE",
            "ZADNJA": "TRASEIRA",
            "NAPRED": "FRENTE",
            "POZADI": "TRASEIRA",
            "LEFT": "ESQUERDA",
            "RIGHT": "DIREITA",
            "TOP": "ACIMA",
            "BOTTOM": "ABAIXO",
            "CENTER": "CENTRO",
            "FRONT": "FRENTE",
            "BACK": "TRASEIRA",
        }
        return mapping.get(txt.upper(), txt)

    if col in {"Kategorija", "Category"}:
        mapping = {
            "okov": "ferragens",
            "potrosni": "consumíveis",
            "consumable": "consumíveis",
            "hardware": "ferragens",
            "warning": "alerta",
        }
        return mapping.get(txt.lower(), txt)

    common: list[tuple[str, str]] = [
        ("Krojna Lista PRO", "Lista de Corte PRO"),
        ("Krojna lista PRO", "Lista de Corte PRO"),
        ("Krojna lista", "Lista de Corte"),
        ("Cut List PRO", "Lista de Corte PRO"),
        ("Cut List", "Lista de Corte"),
        ("Project", "Projeto"),
        ("Customer", "Cliente"),
        ("Room", "Ambiente"),
        ("Kitchen", "Cozinha"),
        ("Wall size", "Tamanho da parede"),
        ("Version", "Versão"),
        ("Generated", "Gerado"),
        ("Measured by", "Medido por"),
        ("Designed by", "Desenhado por"),
        ("Workshop note", "Nota para a marcenaria"),
        ("Wall A", "Parede A"),
        ("Wall B", "Parede B"),
        ("Wall C", "Parede C"),
        ("Zid A", "Parede A"),
        ("Zid B", "Parede B"),
        ("Zid C", "Parede C"),
        ("Radna ploča - Parede", "Bancada - Parede"),
        ("Radna ploča - Zid", "Bancada - Parede"),
        ("Sokla - Zid", "Rodapé - Parede"),
        ("Sokla - Parede", "Rodapé - Parede"),
        ("Donji (kuhinjska jedinica: rerna + ploča za kuvanje)", "Inferior (forno + cooktop)"),
        ("Donji (kuhinjska jedinica: rerna + placa za kuvanje)", "Inferior (forno + cooktop)"),
        ("Donji (kuhinjska jedinica: rerna + ploca za kuvanje)", "Inferior (forno + cooktop)"),
        ("Rerna+ploča", "Forno+cooktop"),
        ("Rerna+ploca", "Forno+cooktop"),
        ("Rerna + ploča", "Forno + cooktop"),
        ("Rerna + ploca", "Forno + cooktop"),
        ("Donji (sudopera)", "Inferior (pia)"),
        ("Donji (2 vrata)", "Inferior (2 portas)"),
        ("Donji (1 vrata)", "Inferior (1 porta)"),
        ("Donji (vrata + fioka)", "Inferior (portas + gaveta)"),
        ("Donji (fioke)", "Inferior (gavetas)"),
        ("Gornji (aspirator / napa)", "Superior (depurador / coifa)"),
        ("Gornji (2 vrata)", "Superior (2 portas)"),
        ("Gornji (1 vrata)", "Superior (1 porta)"),
        ("Gornji", "Superior"),
        ("Donji", "Inferior"),
        ("Iverica Korpus", "Aglomerado corpo"),
        ("Iverica Front", "Aglomerado frente"),
        ("Iverica Leđa", "Aglomerado fundo"),
        ("Iverica Leđa", "Aglomerado fundo"),
        ("Iverica", "Aglomerado"),
        ("Radna ploča Korpus", "Bancada"),
        ("Radna ploča", "Bancada"),
        ("Leđna ploča", "Painel traseiro"),
        ("Parcijalna leđna ploča", "Painel traseiro parcial"),
        ("Back panel / opening", "Painel traseiro / passagem"),
        ("Leđa / prolaz", "Painel traseiro / passagem"),
        ("Leva strana", "Lateral esquerda"),
        ("Desna strana", "Lateral direita"),
        ("Srednja vertikala", "Divisória vertical central"),
        ("Polica (podesiva)", "Prateleira ajustável"),
        ("Polica", "Prateleira"),
        ("Plafon", "Tampo superior do corpo"),
        ("Dno", "Painel inferior"),
        ("Dno sanduka fioke", "Fundo da caixa da gaveta"),
        ("Prednja strana sanduka fioke", "Frente da caixa da gaveta"),
        ("Zadnja strana sanduka fioke", "Traseira da caixa da gaveta"),
        ("Bočna stranica sanduka fioke", "Lateral da caixa da gaveta"),
        ("Dno sanduka", "Fundo da caixa"),
        ("Prednja strana sanduka", "Frente da caixa"),
        ("Zadnja strana sanduka", "Traseira da caixa"),
        ("Bočna ploča", "Lateral da gaveta"),
        ("Front fioke (kuhinjska jedinica)", "Frente da gaveta (forno + cooktop)"),
        ("Front fioke", "Frente da gaveta"),
        ("Vrata (ispod sudopere)", "Portas sob a pia"),
        ("Vrata (ispod ploče za kuvanje)", "Portas sob o cooktop"),
        ("Vrata", "Portas"),
        ("Nosač radne ploče", "Suporte da bancada"),
        ("Sokla (lajsna)", "Rodapé (perfil)"),
        ("Sokla", "Rodapé"),
        ("Filer panel", "Painel de preenchimento"),
        ("Završna bočna ploča", "Painel lateral de acabamento"),
        ("Leđa / prolaz", "Fundo / passagem"),
        ("Zona priključka", "Zona de conexão"),
        ("Ventilacija kolone", "Ventilação da coluna"),
        ("Šarka", "Dobradiça"),
        ("Ručka / pull", "Puxador"),
        ("Klizač za fioku", "Corrediça da gaveta"),
        ("Nosač police (klin)", "Suporte de prateleira"),
        ("Nosač fronta fioke", "Suporte da frente da gaveta"),
        ("Set nosača / ugaonika", "Kit de suportes / cantoneiras"),
        ("Stopica (nogica)", "Pé regulável"),
        ("Konfirmat vijak", "Parafuso confirmat"),
        ("Drvena tipla", "Cavilha de madeira"),
        ("Sanitarni silikon", "Silicone sanitário"),
        ("Sudopera", "Pia"),
        ("Slavina", "Torneira"),
        ("Sifon i odvodni set", "Sifão e kit de escoamento"),
        ("Baterija za sudoperu", "Torneira da pia"),
        ("Ugradna sudopera", "Pia de embutir"),
        ("Klipsa za soklu", "Clipe para rodapé"),
        ("Montažna ploča za šarku", "Placa de montagem da dobradiça"),
        ("Spojnica susednih korpusa", "Conector entre corpos vizinhos"),
        ("Vijak / ekser za leđa", "Parafuso / prego para fundo"),
        ("Vijak / ugaonik za radnu ploču", "Parafuso / cantoneira para bancada"),
        ("Vijak za klizač", "Parafuso da corrediça"),
        ("Vijak za ručku", "Parafuso do puxador"),
        ("Zaptivna lajsna / silikon uz zid", "Perfil de vedação / silicone junto à parede"),
        ("Zidna lajsna ili sanitarni silikon", "Perfil de parede ou silicone sanitário"),
        ("Nosač + šina za viseće elemente", "Suporte + trilho para módulos superiores"),
        ("Tipl + vijak", "Bucha + parafuso"),
        ("Wall plugs / anchors", "Buchas / ancoragens de parede"),
        ("Po izboru korisnika", "À escolha do usuário"),
        ("Podesiva h=100 mm", "Regulável h=100 mm"),
        ("Sifon + preliv + spojnice", "Sifão + ladrão + conexões"),
        ("Kupuje se kao gotov proizvod", "Comprado como produto pronto"),
        ("Kupuje se posebno", "Comprado separadamente"),
        ("Kupovni vodoinstalaterski set za sudoperu", "Kit hidráulico comprado para a pia"),
        ("Po jedna klipsa po stopici za prihvat sokle", "Um clipe por pé para fixar o rodapé"),
        ("Osnovno pričvršćenje leđa po korpusu; broj proveriti prema standardu radionice", "Fixação básica do fundo por corpo; conferir a quantidade conforme o padrão da marcenaria"),
        ("Osnovni set za pričvršćenje radne ploče na nosače i korpuse", "Kit básico para fixar a bancada aos suportes e corpos"),
        ("Osnovni set po elementu", "Kit básico por módulo"),
        ("2 vrata × 2 šarke", "2 portas × 2 dobradiças"),
        ("1 polica × 4 klina", "1 prateleira × 4 suportes"),
        ("1 par klizača", "1 par de corrediças"),
        ("otvor se reže u bancada prema gabarito", "o recorte é feito na bancada conforme o gabarito"),
        ("prema izboru korisnika", "conforme a escolha do usuário"),
        ("2 vijka M4 po ručki; dužinu vijka prilagoditi debljini fronta", "2 parafusos M4 por puxador; ajustar o comprimento à espessura da frente"),
        ("Za zaptivanje ruba sudopere i spojeva uz radnu ploču", "Para vedar a borda da pia e as juntas junto à bancada"),
        ("Standardno za spoj korpusa (bočne+dno+plafon)", "Padrão para unir o corpo (laterais+fundo+tampo superior)"),
        ("Pomoćni spoj i poravnanje", "União auxiliar e alinhamento"),
        ("Za front fioke ispod rerne", "Para a frente da gaveta sob o forno"),
        ("Fioka ispod rerne", "Gaveta sob o forno"),
        ("fioka ispod rerne", "gaveta sob o forno"),
        ("rerna ima sopstveni panel", "o forno tem seu próprio painel"),
        ("prednja ivica kantovana", "borda frontal com fita de borda"),
        ("Gornji vez", "Travessa superior"),
        ("pričvršćuje radnu ploču", "fixa a bancada"),
        ("Segment zida", "Segmento da parede"),
        ("raspon", "intervalo"),
        ("utor 8mm", "canal de 8 mm"),
        ("ostaviti otvor za dovod/odvod vode prema poziciji instalacija", "deixar abertura para entrada/saída de água conforme a posição das instalações"),
        ("parcijalna leđa samo u donjoj servisnoj zoni, gornja zona iza rerne ostaje otvorena radi ventilacije i priključaka", "fundo parcial apenas na zona inferior de serviço; a zona superior atrás do forno fica aberta para ventilação e conexões"),
        ("ostaviti otvor/prolaz za odvod nape prema modelu uređaja", "deixar abertura/passagem para o duto da coifa conforme o modelo do aparelho"),
        ("Otvor za sudoperu", "Recorte para pia"),
        ("Otvor za ploču", "Recorte para cooktop"),
        ("Utor za leđa", "Canal do fundo"),
        ("Ventilacija / otvor", "Ventilação / abertura"),
        ("Posebna obrada", "Usinagem especial"),
        ("Servis + lice mesta", "Marcenaria + obra"),
        ("Kuća / lice mesta", "Casa / obra"),
        ("Servis", "Marcenaria"),
        ("Service pass-through", "Passagem de instalação"),
        ("Po šablonu proizvođača", "Conforme o gabarito do fabricante"),
        ("Po meri iz projekta", "Conforme as medidas do projeto"),
        ("Po geometriji zida i šablonu", "Conforme a geometria da parede e o gabarito"),
        ("Priprema i finalni rez", "Preparação e corte final"),
        ("Zidna mera / finished dimension", "Medida da parede / dimensão final"),
        ("CUT osnova / purchase segment", "Base CUT / segmento de compra"),
        ("Servis radi isključivo po CUT merama", "A marcenaria trabalha estritamente pelas medidas CUT"),
        ("Finalni rez na licu mesta", "Corte final na obra"),
        ("Izrezi", "Recortes"),
        ("sudopera", "pia"),
        ("ploča za kuvanje", "cooktop"),
        ("radnoj ploči", "bancada"),
        ("radnu ploču", "bancada"),
        ("radne ploče", "bancada"),
        ("radna ploča", "bancada"),
        ("šablonu", "gabarito"),
        ("proizvođača", "fabricante"),
    ]

    note_replacements: list[tuple[str, str]] = [
        ("In the workshop, do cutting and edging strictly by CUT dimensions. Verify openings and special machining against the notes.", "Na marcenaria, faça corte e aplicação de borda estritamente pelas medidas CUT. Confira aberturas e usinagens especiais pelas notas."),
        ("In the workshop, cut strictly by CUT dimensions and verify edging in the separate table", "Na marcenaria, corte estritamente pelas medidas CUT e confira a aplicação de borda na tabela separada"),
        ("Cut strictly by the CUT dimensions from the workshop packet.", "Corte estritamente pelas medidas CUT do pacote da marcenaria."),
        ("Apply edging only to the edges marked in the edging table and verify the ABS type.", "Aplique fita de borda apenas nas bordas marcadas na tabela e confira o tipo de ABS."),
        ("Make special openings, grooves and ventilation cuts only where they are explicitly specified.", "Faça aberturas especiais, canais e recortes de ventilação apenas onde estiverem indicados explicitamente."),
        ("Cut the sink or hob opening according to the manufacturer's template, not by estimation.", "Faça o recorte da pia ou do cooktop conforme o gabarito do fabricante, não por estimativa."),
        ("The 'Operations' column clearly shows whether the workshop performs the job or whether it must be confirmed on site.", "A coluna 'Operações' mostra claramente se a marcenaria executa o serviço ou se ele deve ser confirmado na obra."),
        ("The 'Execution basis' column shows whether the work follows project dimensions or the appliance manufacturer's template.", "A coluna 'Base de execução' mostra se o trabalho segue as medidas do projeto ou o gabarito do fabricante do aparelho."),
        ("If the service positions differ from the project, confirm all machining before cutting or on site.", "Se as posições das instalações forem diferentes do projeto, confirme todas as usinagens antes do corte ou na obra."),
        ("Before delivery, verify the piece count, part labels and that all panels/fronts are included.", "Antes da entrega, confira a quantidade de peças, as etiquetas e se todos os painéis/frentes estão incluídos."),
        ("Review the project before ordering", "Revise o projeto antes de encomendar"),
        ("Confirm the wall, dimensions, version and that this is the latest revision", "Confirme a parede, as dimensões, a versão e se esta é a revisão mais recente"),
        ("Take only cutting, edging and machining to the workshop", "Leve à marcenaria apenas corte, aplicação de borda e usinagem"),
        ("The workshop should work only from the workshop packet tables", "A marcenaria deve trabalhar apenas pelas tabelas do pacote da marcenaria"),
        ("Purchase ready-made appliances, hardware and tools separately", "Compre separadamente os eletrodomésticos prontos, ferragens e ferramentas"),
        ("These items are not part of cutting and must be sourced separately", "Esses itens não fazem parte do corte e precisam ser comprados separadamente"),
        ("Sort parts by unit on site", "Separe as peças por módulo na obra"),
        ("Separate carcass parts first, then fronts, then hardware", "Separe primeiro as peças do corpo, depois as frentes e as ferragens"),
        ("Assemble the carcasses first, then doors, drawers and appliances", "Monte primeiro os corpos, depois portas, gavetas e eletrodomésticos"),
        ("Always secure tall and wall units to the wall", "Sempre fixe módulos altos e superiores na parede"),
        ("Check that the material and thickness values are correct for carcass, fronts and backs", "Confira se material e espessuras estão corretos para corpo, frentes e fundos"),
        ("Check that all CUT dimensions and quantities are listed", "Confira se todas as medidas CUT e quantidades estão listadas"),
        ("Check edge banding on every edge before sending to the workshop", "Confira a aplicação de borda em cada borda antes de enviar à marcenaria"),
        ("Check that fronts, backs, plinths and special panels are included", "Confira se frentes, fundos, rodapés e painéis especiais estão incluídos"),
        ("Check that all ready-made purchased items are separated from cut parts", "Confira se todos os itens prontos comprados estão separados das peças de corte"),
        ("Check all openings and special machining for services and ventilation", "Confira todas as aberturas e usinagens especiais para instalações e ventilação"),
        ("Check that the shopping list includes hardware, consumables and tools", "Confira se a lista de compras inclui ferragens, consumíveis e ferramentas"),
        ("Count all cut panels and compare them against the list", "Conte todos os painéis cortados e compare com a lista"),
        ("Sort parts by unit before starting assembly", "Separe as peças por módulo antes de iniciar a montagem"),
        ("Check that all hardware, appliances and tools have been purchased", "Confira se todas as ferragens, eletrodomésticos e ferramentas foram comprados"),
        ("Assemble the carcasses first, then doors and drawers, and only then the appliances", "Monte primeiro os corpos, depois portas e gavetas, e só então os eletrodomésticos"),
        ("Cut the sink opening in the worktop according to the sink manufacturer's template.", "Faça o recorte da pia na bancada conforme o gabarito do fabricante da pia."),
        ("Cut the hob opening in the worktop according to the manufacturer's template.", "Faça o recorte do cooktop na bancada conforme o gabarito do fabricante."),
        ("Provide the opening and duct path for the hood according to the model and the installation axis.", "Faça a abertura e a passagem do duto da coifa conforme o modelo e o eixo de instalação."),
        ("Wall requirement / finished dimension", "Necessidade da parede / dimensão final"),
        ("CUT basis / purchase segment", "Base CUT / segmento de compra"),
        ("Workshop works strictly by CUT dimensions.", "A marcenaria trabalha estritamente pelas medidas CUT."),
        ("Final cut is done on site.", "O corte final é feito na obra."),
        ("According to project dimensions", "Conforme as medidas do projeto"),
        ("According to the manufacturer's template", "Conforme o gabarito do fabricante"),
        ("According to wall geometry and template", "Conforme a geometria da parede e o gabarito"),
        ("Workshop + on site", "Marcenaria + obra"),
        ("Workshop", "Marcenaria"),
        ("On site", "Obra"),
        ("Back panel groove", "Canal do fundo"),
        ("Sink cut-out", "Recorte da pia"),
        ("Hob cut-out", "Recorte do cooktop"),
        ("Ventilation / opening", "Ventilação / abertura"),
        ("Preparation and final cut", "Preparação e corte final"),
        ("Worktop", "Bancada"),
        ("WORKTOP SPECIFICATION", "ESPECIFICAÇÃO DA BANCADA"),
        ("Cutting", "Corte"),
        ("Edging", "Aplicação de borda"),
        ("Machining", "Usinagem"),
        ("Who performs it", "Quem executa"),
        ("Execution basis", "Base de execução"),
        ("Services", "Instalações"),
        ("Final check", "Conferência final"),
        ("Purchased ready-made", "Comprado pronto"),
        ("Project consumables", "Consumíveis do projeto"),
        ("Hardware and mechanisms by unit", "Ferragens e mecanismos por módulo"),
        ("Consumables by unit", "Consumíveis por módulo"),
        ("Wall installation", "Instalação na parede"),
        ("Tools needed on site", "Ferramentas necessárias na obra"),
        ("For wall installation", "Para instalação na parede"),
        ("Cordless drill / driver", "Parafusadeira sem fio"),
        ("Tape measure + spirit level", "Trena + nível"),
        ("Clamps", "Sargentos"),
        ("Needed for cabinet assembly and hardware installation", "Necessário para montar os módulos e instalar as ferragens"),
        ("Needed to verify dimensions, plumb and levelling", "Necessário para conferir medidas, prumo e nivelamento"),
        ("They help keep parts aligned during assembly", "Ajudam a manter as peças alinhadas durante a montagem"),
        ("Choose according to the actual wall type on site", "Escolha conforme o tipo real da parede na obra"),
        ("Ušinagem", "Usinagem"),
        ("ušinagem", "usinagem"),
        ("ušinagens", "usinagens"),
    ]

    if col in {"Napomena", "Napomena za servis", "Instrukcija", "Stavka", "Sta radis", "Šta radiš", "Polje", "Vrednost", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena", "Grupa", "Tip / Šifra", "Note", "Instruction", "Item", "Field", "Value", "Processing type", "Operations", "Execution basis", "Processing / note", "Group", "Type / Code"}:
        txt = _apply_ordered_replacements(txt, note_replacements)

    return _apply_ordered_replacements(txt, common)


def _translate_export_text_es(value: Any, column: str = "") -> str:
    txt = str(value or "")
    col = str(column or "")

    if col in {"Zid", "Wall"}:
        return txt.replace("Zid ", "Pared ").replace("Wall ", "Pared ")

    if col in {"Pozicija", "Position"}:
        mapping = {
            "LEVA": "IZQUIERDA", "LEVO": "IZQUIERDA", "LEFT": "IZQUIERDA",
            "DESNA": "DERECHA", "DESNO": "DERECHA", "RIGHT": "DERECHA",
            "GORE": "ARRIBA", "TOP": "ARRIBA",
            "DOLE": "ABAJO", "BOTTOM": "ABAJO",
            "CENTAR": "CENTRO", "SREDINA": "CENTRO", "CENTER": "CENTRO",
            "PREDNJA": "FRENTE", "NAPRED": "FRENTE", "FRONT": "FRENTE",
            "ZADNJA": "TRASERA", "POZADI": "TRASERA", "BACK": "TRASERA",
        }
        return mapping.get(txt.upper(), txt)

    if col in {"Kategorija", "Category"}:
        mapping = {
            "okov": "herrajes",
            "potrosni": "consumibles",
            "consumable": "consumibles",
            "hardware": "herrajes",
            "warning": "aviso",
        }
        return mapping.get(txt.lower(), txt)

    common: list[tuple[str, str]] = [
        ("Krojna Lista PRO", "Lista de Corte PRO"),
        ("Krojna lista PRO", "Lista de Corte PRO"),
        ("Krojna lista", "Lista de Corte"),
        ("Cut List PRO", "Lista de Corte PRO"),
        ("Cut List", "Lista de Corte"),
        ("Project", "Proyecto"), ("Customer", "Cliente"), ("Room", "Ambiente"),
        ("Kitchen", "Cocina"), ("Wall size", "Medida de pared"), ("Version", "Versión"),
        ("Generated", "Generado"), ("Measured by", "Medido por"), ("Designed by", "Diseñado por"),
        ("Workshop note", "Nota para el taller"),
        ("Wall A", "Pared A"), ("Wall B", "Pared B"), ("Wall C", "Pared C"),
        ("Zid A", "Pared A"), ("Zid B", "Pared B"), ("Zid C", "Pared C"),
        ("Donji (kuhinjska jedinica: rerna + ploča za kuvanje)", "Bajo (horno + placa)"),
        ("Donji (kuhinjska jedinica: rerna + ploca za kuvanje)", "Bajo (horno + placa)"),
        ("Rerna+ploča", "Horno+placa"), ("Rerna+ploca", "Horno+placa"),
        ("Rerna + ploča", "Horno + placa"), ("Rerna + ploca", "Horno + placa"),
        ("Donji (sudopera)", "Bajo (fregadero)"), ("Donji (2 vrata)", "Bajo (2 puertas)"),
        ("Donji (1 vrata)", "Bajo (1 puerta)"), ("Donji (vrata + fioka)", "Bajo (puertas + cajón)"),
        ("Donji (fioke)", "Bajo (cajones)"), ("Donji", "Bajo"),
        ("Gornji (aspirator / napa)", "Alto (campana)"), ("Gornji (2 vrata)", "Alto (2 puertas)"),
        ("Gornji (1 vrata)", "Alto (1 puerta)"), ("Gornji", "Alto"),
        ("Iverica Korpus", "Tablero cuerpo"), ("Iverica Front", "Tablero frente"),
        ("Iverica Leđa", "Tablero trasera"), ("Iverica", "Tablero"),
        ("Radna ploča Korpus", "Encimera"), ("Radna ploča", "Encimera"),
        ("Leđna ploča", "Panel trasero"), ("Parcijalna leđna ploča", "Panel trasero parcial"),
        ("Back panel / opening", "Panel trasero / paso"), ("Leđa / prolaz", "Panel trasero / paso"),
        ("Leva strana", "Lateral izquierda"), ("Desna strana", "Lateral derecha"),
        ("Srednja vertikala", "División vertical central"), ("Polica (podesiva)", "Estante ajustable"),
        ("Polica", "Estante"), ("Plafon", "Panel superior"), ("Dno", "Panel inferior"),
        ("Dno sanduka fioke", "Fondo de la caja del cajón"),
        ("Prednja strana sanduka fioke", "Frente de la caja del cajón"),
        ("Zadnja strana sanduka fioke", "Trasera de la caja del cajón"),
        ("Bočna stranica sanduka fioke", "Lateral de la caja del cajón"),
        ("Dno sanduka", "Fondo de la caja"), ("Prednja strana sanduka", "Frente de la caja"),
        ("Zadnja strana sanduka", "Trasera de la caja"), ("Bočna ploča", "Lateral del cajón"),
        ("Front fioke (kuhinjska jedinica)", "Frente de cajón (horno + placa)"),
        ("Front fioke", "Frente de cajón"), ("Fioka ispod rerne", "Cajón bajo el horno"),
        ("fioka ispod rerne", "cajón bajo el horno"), ("rerna ima sopstveni panel", "el horno tiene su propio panel"),
        ("Vrata (ispod sudopere)", "Puertas bajo el fregadero"),
        ("Vrata (ispod ploče za kuvanje)", "Puertas bajo la placa"), ("Vrata", "Puertas"),
        ("Nosač radne ploče", "Soporte de encimera"), ("Sokla (lajsna)", "Zócalo (perfil)"),
        ("Sokla", "Zócalo"), ("Radna ploča - Zid", "Encimera - Pared"), ("Sokla - Zid", "Zócalo - Pared"),
        ("Filer panel", "Panel de relleno"), ("Završna bočna ploča", "Panel lateral de remate"),
        ("Šarka", "Bisagra"), ("Ručka / pull", "Tirador"), ("Klizač za fioku", "Guía de cajón"),
        ("Nosač police (klin)", "Soporte de estante"), ("Nosač fronta fioke", "Soporte del frente del cajón"),
        ("Set nosača / ugaonika", "Juego de soportes / escuadras"), ("Stopica (nogica)", "Pata regulable"),
        ("Konfirmat vijak", "Tornillo confirmat"), ("Drvena tipla", "Espiga de madera"),
        ("Sanitarni silikon", "Silicona sanitaria"), ("Sudopera", "Fregadero"), ("Slavina", "Grifo"),
        ("Sifon i odvodni set", "Sifón y kit de desagüe"), ("Baterija za sudoperu", "Grifo de fregadero"),
        ("Ugradna sudopera", "Fregadero encastrado"), ("Tipl + vijak", "Taco + tornillo"),
        ("Wall plugs / anchors", "Tacos / anclajes de pared"), ("Po izboru korisnika", "A elección del usuario"),
        ("Podesiva h=100 mm", "Regulable h=100 mm"), ("Sifon + preliv + spojnice", "Sifón + rebosadero + conectores"),
        ("Kupuje se kao gotov proizvod", "Se compra como producto terminado"),
        ("Kupuje se posebno", "Se compra por separado"), ("prema izboru korisnika", "según la elección del usuario"),
        ("Osnovni set po elementu", "Kit básico por módulo"), ("1 par klizača", "1 par de guías"),
        ("1 polica × 4 klina", "1 estante × 4 soportes"), ("2 vrata × 2 šarke", "2 puertas × 2 bisagras"),
        ("Za front fioke ispod rerne", "Para el frente del cajón bajo el horno"),
        ("Standardno za spoj korpusa (bočne+dno+plafon)", "Estándar para unir el cuerpo (laterales+inferior+superior)"),
        ("Pomoćni spoj i poravnanje", "Unión auxiliar y alineación"),
        ("Gornji vez", "Travesaño superior"), ("pričvršćuje radnu ploču", "fija la encimera"),
        ("Segment zida", "Segmento de pared"), ("raspon", "rango"), ("utor 8mm", "ranura de 8 mm"),
        ("Otvor za sudoperu", "Hueco para fregadero"), ("Otvor za ploču", "Hueco para placa"),
        ("Utor za leđa", "Ranura del panel trasero"), ("Ventilacija / otvor", "Ventilación / abertura"),
        ("Posebna obrada", "Mecanizado especial"), ("Servis + lice mesta", "Taller + obra"),
        ("Kuća / lice mesta", "Casa / obra"), ("Servis", "Taller"),
        ("parcijalna leđa samo u donjoj servisnoj zoni, gornja zona iza rerne ostaje otvorena radi ventilacije i priključaka", "panel trasero parcial solo en la zona inferior de servicio; la zona superior detrás del horno queda abierta para ventilación y conexiones"),
        ("Wall installation", "Instalación en pared"), ("Purchased separately", "Compra por separado"),
        ("Tools needed on site", "Herramientas necesarias en obra"),
        ("Anker za zid", "Anclaje de pared"), ("Vijak za zidni nosač / šinu", "Tornillo para soporte / riel de pared"),
        ("Zidni nosač elementa", "Soporte de pared del módulo"), ("Nosač + šina za viseće elemente", "Soporte + riel para módulos colgantes"),
        ("Preporučena količina po modulu", "Cantidad recomendada por módulo"),
        ("Za pričvršćenje nosača, šine ili anti-tip seta; tip vijka prilagoditi zidu", "Para fijar el soporte, el riel o el kit antivuelco; adaptar el tipo de tornillo a la pared"),
        ("Set za kačenje visećeg elementa", "Kit para colgar el módulo alto"),
        ("2 × 1 ručka — odabrati po projektu", "2 × 1 tirador — elegir según el proyecto"),
        ("Po šablonu proizvođača", "Según la plantilla del fabricante"),
        ("Po meri iz projekta", "Según las medidas del proyecto"),
        ("Po geometriji zida i šablonu", "Según la geometría de la pared y la plantilla"),
        ("Priprema i finalni rez", "Preparación y corte final"),
        ("Zidna mera / finished dimension", "Medida de pared / dimensión final"),
        ("CUT osnova / purchase segment", "Base CUT / segmento de compra"),
        ("Servis radi isključivo po CUT merama", "El taller trabaja estrictamente con medidas CUT"),
        ("Finalni rez na licu mesta", "Corte final en obra"), ("Izrezi", "Recortes"),
        ("sudopera", "fregadero"), ("ploča za kuvanje", "placa"), ("radnoj ploči", "encimera"),
        ("radnu ploču", "encimera"), ("radne ploče", "encimera"), ("radna ploča", "encimera"),
        ("šablonu", "plantilla"), ("proizvođača", "fabricante"),
    ]
    note_replacements: list[tuple[str, str]] = [
        ("In the workshop, do cutting and edging strictly by CUT dimensions. Verify openings and special machining against the notes.", "En el taller, cortar y cantear estrictamente según las medidas CUT. Verificar huecos y mecanizados especiales según las notas."),
        ("In the workshop, cut strictly by CUT dimensions and verify edging in the separate table", "En el taller, cortar estrictamente según las medidas CUT y verificar el canteado en la tabla separada"),
        ("Take only cutting, edging and machining to the workshop", "Lleva al taller solo corte, canteado y mecanizado"),
        ("The workshop should work only from the workshop packet tables", "El taller debe trabajar solo con las tablas del paquete de taller"),
        ("Purchase ready-made appliances, hardware and tools separately", "Compra por separado electrodomésticos, herrajes y herramientas"),
        ("These items are not part of cutting and must be sourced separately", "Estos elementos no forman parte del corte y deben comprarse por separado"),
        ("Back panel groove", "Ranura del panel trasero"), ("Sink cut-out", "Hueco para fregadero"),
        ("Hob cut-out", "Hueco para placa"), ("Ventilation / opening", "Ventilación / abertura"),
        ("Preparation and final cut", "Preparación y corte final"), ("Workshop + on site", "Taller + obra"),
        ("Workshop", "Taller"), ("On site", "Obra"), ("According to project dimensions", "Según las medidas del proyecto"),
        ("According to the manufacturer's template", "Según la plantilla del fabricante"),
        ("According to wall geometry and template", "Según la geometría de la pared y la plantilla"),
        ("Cut the sink opening in the worktop according to the sink manufacturer's template.", "Cortar el hueco del fregadero en la encimera según la plantilla del fabricante."),
        ("Cut the hob opening in the worktop according to the manufacturer's template.", "Cortar el hueco de la placa en la encimera según la plantilla del fabricante."),
        ("Provide the opening and duct path for the hood according to the model and the installation axis.", "Preparar la abertura y el paso del conducto de la campana según el modelo y el eje de instalación."),
        ("Wall requirement / finished dimension", "Necesidad de pared / dimensión final"),
        ("CUT basis / purchase segment", "Base CUT / segmento de compra"),
        ("Workshop works strictly by CUT dimensions.", "El taller trabaja estrictamente con medidas CUT."),
        ("Final cut is done on site.", "El corte final se realiza en obra."),
    ]
    if col in {"Napomena", "Napomena za servis", "Instrukcija", "Stavka", "Sta radis", "Šta radiš", "Polje", "Vrednost", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena", "Grupa", "Tip / Šifra", "Note", "Instruction", "Item", "Field", "Value", "Processing type", "Operations", "Execution basis", "Processing / note", "Group", "Type / Code"}:
        txt = _apply_ordered_replacements(txt, note_replacements)
    return _apply_ordered_replacements(txt, common)


def _translate_export_text_ru(value: Any, column: str = "") -> str:
    txt = str(value or "")
    col = str(column or "")

    if col in {"Zid", "Wall"}:
        return txt.replace("Zid ", "Стена ").replace("Wall ", "Стена ")

    if col in {"Pozicija", "Position"}:
        mapping = {
            "LEVA": "ЛЕВАЯ", "LEVO": "ЛЕВАЯ", "LEFT": "ЛЕВАЯ",
            "DESNA": "ПРАВАЯ", "DESNO": "ПРАВАЯ", "RIGHT": "ПРАВАЯ",
            "GORE": "ВЕРХ", "TOP": "ВЕРХ",
            "DOLE": "НИЗ", "BOTTOM": "НИЗ",
            "CENTAR": "ЦЕНТР", "SREDINA": "ЦЕНТР", "CENTER": "ЦЕНТР",
            "PREDNJA": "ПЕРЕД", "NAPRED": "ПЕРЕД", "FRONT": "ПЕРЕД",
            "ZADNJA": "ЗАДНЯЯ", "POZADI": "ЗАДНЯЯ", "BACK": "ЗАДНЯЯ",
        }
        return mapping.get(txt.upper(), txt)

    if col in {"Kategorija", "Category"}:
        mapping = {
            "okov": "фурнитура",
            "potrosni": "расходники",
            "consumable": "расходники",
            "hardware": "фурнитура",
            "warning": "предупреждение",
        }
        return mapping.get(txt.lower(), txt)

    common: list[tuple[str, str]] = [
        ("Krojna Lista PRO", "Карта раскроя PRO"), ("Krojna lista PRO", "Карта раскроя PRO"),
        ("Krojna lista", "Карта раскроя"), ("Cut List PRO", "Карта раскроя PRO"), ("Cut List", "Карта раскроя"),
        ("Project", "Проект"), ("Customer", "Клиент"), ("Room", "Помещение"), ("Kitchen", "Кухня"),
        ("Wall size", "Размер стены"), ("Version", "Версия"), ("Generated", "Создано"),
        ("Measured by", "Замерил"), ("Designed by", "Спроектировал"), ("Workshop note", "Примечание для мастерской"),
        ("Wall A", "Стена A"), ("Wall B", "Стена B"), ("Wall C", "Стена C"),
        ("Zid A", "Стена A"), ("Zid B", "Стена B"), ("Zid C", "Стена C"),
        ("Donji (kuhinjska jedinica: rerna + ploča za kuvanje)", "Нижний модуль (духовка + варочная панель)"),
        ("Donji (kuhinjska jedinica: rerna + ploca za kuvanje)", "Нижний модуль (духовка + варочная панель)"),
        ("Rerna+ploča", "Духовка+панель"), ("Rerna+ploca", "Духовка+панель"),
        ("Rerna + ploča", "Духовка + панель"), ("Rerna + ploca", "Духовка + панель"),
        ("Donji (sudopera)", "Нижний модуль (мойка)"), ("Donji (2 vrata)", "Нижний модуль (2 двери)"),
        ("Donji (1 vrata)", "Нижний модуль (1 дверь)"), ("Donji (vrata + fioka)", "Нижний модуль (двери + ящик)"),
        ("Donji (fioke)", "Нижний модуль (ящики)"), ("Donji", "Нижний модуль"),
        ("Gornji (aspirator / napa)", "Верхний модуль (вытяжка)"), ("Gornji (2 vrata)", "Верхний модуль (2 двери)"),
        ("Gornji (1 vrata)", "Верхний модуль (1 дверь)"), ("Gornji", "Верхний модуль"),
        ("Iverica Korpus", "ЛДСП корпус"), ("Iverica Front", "ЛДСП фасад"), ("Iverica Leđa", "ХДФ задник"),
        ("Iverica", "ЛДСП"), ("Lesonit Leđa", "ХДФ задник"), ("Leđa", "Задник"),
        ("Radna ploča Korpus", "Столешница"), ("Radna ploča", "Столешница"),
        ("Leđna ploča", "Задняя панель"), ("Parcijalna leđna ploča", "Частичная задняя панель"),
        ("Back panel / opening", "Задняя панель / проём"), ("Leđa / prolaz", "Задняя панель / проём"),
        ("Leva strana", "Левая боковина"), ("Desna strana", "Правая боковина"),
        ("Srednja vertikala", "Средняя вертикаль"), ("Polica (podesiva)", "Регулируемая полка"),
        ("Polica", "Полка"), ("Plafon", "Верхняя панель"), ("Dno", "Нижняя панель"),
        ("Dno sanduka fioke", "Дно ящика"), ("Prednja strana sanduka fioke", "Передняя стенка ящика"),
        ("Zadnja strana sanduka fioke", "Задняя стенка ящика"), ("Bočna stranica sanduka fioke", "Боковина ящика"),
        ("Dno sanduka", "Дно ящика"), ("Prednja strana sanduka", "Передняя стенка ящика"),
        ("Zadnja strana sanduka", "Задняя стенка ящика"), ("Bočna ploča", "Боковина ящика"),
        ("Front fioke (kuhinjska jedinica)", "Фасад ящика (духовка + панель)"), ("Front fioke", "Фасад ящика"),
        ("Fioka ispod rerne", "Ящик под духовкой"), ("fioka ispod rerne", "ящик под духовкой"),
        ("rerna ima sopstveni panel", "у духовки собственная панель"),
        ("Vrata (ispod sudopere)", "Двери под мойкой"), ("Vrata (ispod ploče za kuvanje)", "Двери под варочной панелью"),
        ("Vrata", "Двери"), ("Nosač radne ploče", "Опора столешницы"), ("Sokla (lajsna)", "Цоколь (планка)"),
        ("Sokla", "Цоколь"), ("Radna ploča - Zid", "Столешница - Стена"), ("Sokla - Zid", "Цоколь - Стена"),
        ("Filer panel", "Доборная панель"), ("Završna bočna ploča", "Торцевая боковая панель"),
        ("Šarka", "Петля"), ("Ručka / pull", "Ручка"), ("Klizač za fioku", "Направляющая ящика"),
        ("Nosač police (klin)", "Полкодержатель"), ("Nosač fronta fioke", "Крепление фасада ящика"),
        ("Set nosača / ugaonika", "Комплект креплений / уголков"), ("Stopica (nogica)", "Регулируемая ножка"),
        ("Konfirmat vijak", "Конфирмат"), ("Drvena tipla", "Деревянный шкант"),
        ("Sanitarni silikon", "Санитарный силикон"), ("Sudopera", "Мойка"), ("Slavina", "Смеситель"),
        ("Sifon i odvodni set", "Сифон и сливной комплект"), ("Baterija za sudoperu", "Смеситель для мойки"),
        ("Ugradna sudopera", "Встраиваемая мойка"), ("Tipl + vijak", "Дюбель + шуруп"),
        ("Wall plugs / anchors", "Дюбели / анкеры"), ("Po izboru korisnika", "По выбору пользователя"),
        ("Podesiva h=100 mm", "Регулируемая h=100 мм"), ("Sifon + preliv + spojnice", "Сифон + перелив + соединители"),
        ("Kupuje se kao gotov proizvod", "Покупается готовым изделием"), ("Kupuje se posebno", "Покупается отдельно"),
        ("prema izboru korisnika", "по выбору пользователя"), ("Osnovni set po elementu", "Базовый комплект на модуль"),
        ("1 par klizača", "1 пара направляющих"), ("1 polica × 4 klina", "1 полка × 4 держателя"),
        ("2 vrata × 2 šarke", "2 двери × 2 петли"), ("Za front fioke ispod rerne", "Для фасада ящика под духовкой"),
        ("Standardno za spoj korpusa (bočne+dno+plafon)", "Стандартно для соединения корпуса (боковины+низ+верх)"),
        ("Pomoćni spoj i poravnanje", "Вспомогательное соединение и выравнивание"),
        ("Gornji vez", "Верхняя стяжка"), ("pričvršćuje radnu ploču", "крепит столешницу"),
        ("Segment zida", "Сегмент стены"), ("raspon", "диапазон"), ("utor 8mm", "паз 8 мм"),
        ("Otvor za sudoperu", "Вырез под мойку"), ("Otvor za ploču", "Вырез под варочную панель"),
        ("Utor za leđa", "Паз задней панели"), ("Ventilacija / otvor", "Вентиляция / проём"),
        ("Posebna obrada", "Специальная обработка"), ("Servis + lice mesta", "Мастерская + на месте"),
        ("Kuća / lice mesta", "Дом / на месте"), ("Servis", "Мастерская"),
        ("Wall installation", "Монтаж на стену"), ("Purchased separately", "Покупается отдельно"),
        ("Tools needed on site", "Инструменты на месте"), ("Anker za zid", "Стеновой анкер"),
        ("Vijak za zidni nosač / šinu", "Шуруп для настенного крепления / рейки"),
        ("Zidni nosač elementa", "Настенный крепёж модуля"), ("Nosač + šina za viseće elemente", "Крепёж + рейка для навесных модулей"),
        ("Preporučena količina po modulu", "Рекомендуемое количество на модуль"),
        ("Za pričvršćenje nosača, šine ili anti-tip seta; tip vijka prilagoditi zidu", "Для крепления держателя, рейки или антиопрокидывателя; тип шурупа подобрать под стену"),
        ("Set za kačenje visećeg elementa", "Комплект для навесного модуля"),
        ("Po šablonu proizvođača", "По шаблону производителя"), ("Po meri iz projekta", "По размерам проекта"),
        ("Po geometriji zida i šablonu", "По геометрии стены и шаблону"),
        ("Priprema i finalni rez", "Подготовка и финальный рез"),
        ("Zidna mera / finished dimension", "Размер стены / финальный размер"),
        ("CUT osnova / purchase segment", "База CUT / покупной сегмент"),
        ("Servis radi isključivo po CUT merama", "Мастерская работает строго по размерам CUT"),
        ("Finalni rez na licu mesta", "Финальная подрезка на месте"), ("Izrezi", "Вырезы"),
        ("sudopera", "мойка"), ("ploča za kuvanje", "варочная панель"),
        ("radnoj ploči", "столешнице"), ("radnu ploču", "столешницу"), ("radne ploče", "столешницы"),
        ("radna ploča", "столешница"), ("šablonu", "шаблону"), ("proizvođača", "производителя"),
        ("parcijalna leđa samo u donjoj servisnoj zoni, gornja zona iza rerne ostaje otvorena radi ventilacije i priključaka", "частичная задняя панель только в нижней сервисной зоне; верхняя зона за духовкой остаётся открытой для вентиляции и подключений"),
    ]
    note_replacements: list[tuple[str, str]] = [
        ("In the workshop, do cutting and edging strictly by CUT dimensions. Verify openings and special machining against the notes.", "В мастерской выполнять раскрой и кромление строго по размерам CUT. Проёмы и специальную обработку сверять с примечаниями."),
        ("In the workshop, cut strictly by CUT dimensions and verify edging in the separate table", "В мастерской резать строго по размерам CUT и проверить кромление в отдельной таблице"),
        ("Take only cutting, edging and machining to the workshop", "В мастерскую передать только раскрой, кромление и обработку"),
        ("The workshop should work only from the workshop packet tables", "Мастерская должна работать только по таблицам мастерского пакета"),
        ("Purchase ready-made appliances, hardware and tools separately", "Готовую технику, фурнитуру и инструменты купить отдельно"),
        ("These items are not part of cutting and must be sourced separately", "Эти позиции не относятся к раскрою и покупаются отдельно"),
        ("Back panel groove", "Паз задней панели"), ("Sink cut-out", "Вырез под мойку"),
        ("Hob cut-out", "Вырез под варочную панель"), ("Ventilation / opening", "Вентиляция / проём"),
        ("Preparation and final cut", "Подготовка и финальный рез"), ("Workshop + on site", "Мастерская + на месте"),
        ("Workshop", "Мастерская"), ("On site", "На месте"), ("According to project dimensions", "По размерам проекта"),
        ("According to the manufacturer's template", "По шаблону производителя"),
        ("According to wall geometry and template", "По геометрии стены и шаблону"),
        ("Cut the sink opening in the worktop according to the sink manufacturer's template.", "Вырезать отверстие под мойку в столешнице по шаблону производителя мойки."),
        ("Cut the hob opening in the worktop according to the manufacturer's template.", "Вырезать отверстие под варочную панель в столешнице по шаблону производителя."),
        ("Provide the opening and duct path for the hood according to the model and the installation axis.", "Подготовить проём и трассу воздуховода для вытяжки по модели и оси установки."),
        ("Wall requirement / finished dimension", "Требование стены / финальный размер"),
        ("CUT basis / purchase segment", "База CUT / покупной сегмент"),
        ("Workshop works strictly by CUT dimensions.", "Мастерская работает строго по размерам CUT."),
        ("Final cut is done on site.", "Финальная подрезка выполняется на месте."),
    ]
    if col in {"Napomena", "Napomena za servis", "Instrukcija", "Stavka", "Sta radis", "Šta radiš", "Polje", "Vrednost", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena", "Grupa", "Tip / Šifra", "Note", "Instruction", "Item", "Field", "Value", "Processing type", "Operations", "Execution basis", "Processing / note", "Group", "Type / Code"}:
        txt = _apply_ordered_replacements(txt, note_replacements)
    return _apply_ordered_replacements(txt, common)


def _translate_export_text_zhcn(value: Any, column: str = "") -> str:
    txt = str(value or "")
    col = str(column or "")

    if col in {"Zid", "Wall"}:
        return txt.replace("Zid ", "墙 ").replace("Wall ", "墙 ")

    if col in {"Pozicija", "Position"}:
        mapping = {
            "LEVA": "左", "LEVO": "左", "LEFT": "左",
            "DESNA": "右", "DESNO": "右", "RIGHT": "右",
            "GORE": "上", "TOP": "上",
            "DOLE": "下", "BOTTOM": "下",
            "CENTAR": "中", "SREDINA": "中", "CENTER": "中",
            "PREDNJA": "前", "NAPRED": "前", "FRONT": "前",
            "ZADNJA": "后", "POZADI": "后", "BACK": "后",
        }
        return mapping.get(txt.upper(), txt)

    common: list[tuple[str, str]] = [
        ("Krojna Lista PRO", "裁切清单 PRO"), ("Krojna lista PRO", "裁切清单 PRO"),
        ("Krojna lista", "裁切清单"), ("Cut List PRO", "裁切清单 PRO"), ("Cut List", "裁切清单"),
        ("Project", "项目"), ("Customer", "客户"), ("Room", "房间"), ("Kitchen", "厨房"),
        ("Wall size", "墙面尺寸"), ("Version", "版本"), ("Generated", "生成时间"),
        ("Measured by", "测量人"), ("Designed by", "设计人"), ("Workshop note", "车间备注"),
        ("Wall A", "墙 A"), ("Wall B", "墙 B"), ("Wall C", "墙 C"),
        ("Zid A", "墙 A"), ("Zid B", "墙 B"), ("Zid C", "墙 C"),
        ("Donji", "下柜"), ("Gornji", "吊柜"),
        ("Iverica Korpus", "刨花板 柜体"), ("Iverica Front", "刨花板 门板"), ("Iverica Leđa", "HDF 背板"),
        ("Iverica", "刨花板"), ("Lesonit Leđa", "HDF 背板"), ("Leđa", "背板"),
        ("Radna ploča Korpus", "台面"), ("Radna ploča", "台面"),
        ("Leđna ploča", "背板"), ("Parcijalna leđna ploča", "局部背板"),
        ("Back panel / opening", "背板 / 开口"), ("Leđa / prolaz", "背板 / 开口"),
        ("Leva strana", "左侧板"), ("Desna strana", "右侧板"),
        ("Srednja vertikala", "中立板"), ("Polica (podesiva)", "可调层板"),
        ("Polica", "层板"), ("Plafon", "顶板"), ("Dno", "底板"),
        ("Dno sanduka fioke", "抽屉底板"), ("Prednja strana sanduka fioke", "抽屉前板"),
        ("Zadnja strana sanduka fioke", "抽屉后板"), ("Bočna stranica sanduka fioke", "抽屉侧板"),
        ("Front fioke", "抽屉门板"), ("Vrata", "门板"), ("Sokla", "踢脚板"),
        ("Filer panel", "补板"), ("Završna bočna ploča", "收口侧板"),
        ("Šarka", "铰链"), ("Ručka / pull", "拉手"), ("Klizač za fioku", "抽屉滑轨"),
        ("Konfirmat vijak", "确认钉"), ("Drvena tipla", "木榫"),
        ("Sanitarni silikon", "防霉硅胶"), ("Sudopera", "水槽"), ("Slavina", "龙头"),
        ("Ugradna sudopera", "嵌入式水槽"), ("Wall plugs / anchors", "膨胀塞 / 锚栓"),
        ("Purchased separately", "需单独购买"), ("Tools needed on site", "现场所需工具"),
        ("Wall installation", "墙面安装"), ("Po šablonu proizvođača", "按厂家模板"),
        ("Po meri iz projekta", "按项目尺寸"), ("Po geometriji zida i šablonu", "按墙面尺寸与模板"),
        ("Priprema i finalni rez", "准备与最终切割"), ("Otvor za sudoperu", "水槽开孔"),
        ("Otvor za ploču", "灶具开孔"), ("Utor za leđa", "背板槽"),
        ("Ventilacija / otvor", "通风 / 开口"), ("Posebna obrada", "特殊加工"),
        ("Servis", "车间"), ("Kuća / lice mesta", "现场"),
    ]
    note_replacements: list[tuple[str, str]] = [
        ("In the workshop, do cutting and edging strictly by CUT dimensions. Verify openings and special machining against the notes.", "车间必须严格按 CUT 尺寸切割和封边，并按备注核对开孔和特殊加工。"),
        ("In the workshop, cut strictly by CUT dimensions and verify edging in the separate table", "车间必须严格按 CUT 尺寸切割，并在单独表格中核对封边"),
        ("Take only cutting, edging and machining to the workshop", "带到车间的只有切割、封边和加工部分"),
        ("The workshop should work only from the workshop packet tables", "车间只能按车间资料包中的表格执行"),
        ("Purchase ready-made appliances, hardware and tools separately", "成品电器、五金和工具需单独购买"),
        ("These items are not part of cutting and must be sourced separately", "这些项目不属于板材切割，需单独采购"),
        ("Back panel groove", "背板槽"), ("Sink cut-out", "水槽开孔"),
        ("Hob cut-out", "灶具开孔"), ("Ventilation / opening", "通风 / 开口"),
        ("Preparation and final cut", "准备与最终切割"), ("Workshop + on site", "车间 + 现场"),
        ("Workshop", "车间"), ("On site", "现场"),
        ("According to wall geometry and template", "按墙面尺寸与模板"),
        ("Wall requirement / finished dimension", "墙面要求 / 成品尺寸"),
        ("CUT basis / purchase segment", "CUT 基准 / 采购长度"),
        ("Workshop works strictly by CUT dimensions.", "车间严格按 CUT 尺寸执行。"),
        ("Final cut is done on site.", "最终切割在现场完成。"),
    ]
    if col in {"Napomena", "Napomena za servis", "Instrukcija", "Stavka", "Sta radis", "Šta radiš", "Polje", "Vrednost", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena", "Grupa", "Tip / Šifra", "Note", "Instruction", "Item", "Field", "Value", "Processing type", "Operations", "Execution basis", "Processing / note", "Group", "Type / Code"}:
        txt = _apply_ordered_replacements(txt, note_replacements)
    return _apply_ordered_replacements(txt, common)


def _translate_export_text_hi(value: Any, column: str = "") -> str:
    txt = str(value or "")
    col = str(column or "")

    if col in {"Zid", "Wall"}:
        return txt.replace("Zid ", "दीवार ").replace("Wall ", "दीवार ")

    if col in {"Pozicija", "Position"}:
        mapping = {
            "LEVA": "बायां", "LEVO": "बायां", "LEFT": "बायां",
            "DESNA": "दायां", "DESNO": "दायां", "RIGHT": "दायां",
            "GORE": "ऊपर", "TOP": "ऊपर",
            "DOLE": "नीचे", "BOTTOM": "नीचे",
            "CENTAR": "मध्य", "SREDINA": "मध्य", "CENTER": "मध्य",
            "PREDNJA": "सामने", "NAPRED": "सामने", "FRONT": "सामने",
            "ZADNJA": "पीछे", "POZADI": "पीछे", "BACK": "पीछे",
        }
        return mapping.get(txt.upper(), txt)

    common: list[tuple[str, str]] = [
        ("Krojna Lista PRO", "कट लिस्ट PRO"), ("Krojna lista PRO", "कट लिस्ट PRO"),
        ("Krojna lista", "कट लिस्ट"), ("Cut List PRO", "कट लिस्ट PRO"), ("Cut List", "कट लिस्ट"),
        ("Project", "प्रोजेक्ट"), ("Customer", "ग्राहक"), ("Room", "कमरा"), ("Kitchen", "रसोई"),
        ("Wall size", "दीवार माप"), ("Version", "संस्करण"), ("Generated", "जनरेट किया गया"),
        ("Measured by", "माप लिया"), ("Designed by", "डिज़ाइन किया"), ("Workshop note", "वर्कशॉप नोट"),
        ("Wall A", "दीवार A"), ("Wall B", "दीवार B"), ("Wall C", "दीवार C"),
        ("Zid A", "दीवार A"), ("Zid B", "दीवार B"), ("Zid C", "दीवार C"),
        ("Donji", "निचला मॉड्यूल"), ("Gornji", "ऊपरी मॉड्यूल"),
        ("Iverica Korpus", "पार्टिकल बोर्ड कार्कस"), ("Iverica Front", "पार्टिकल बोर्ड फ्रंट"), ("Iverica Leđa", "HDF बैक"),
        ("Iverica", "पार्टिकल बोर्ड"), ("Lesonit Leđa", "HDF बैक"), ("Leđa", "बैक"),
        ("Radna ploča Korpus", "वर्कटॉप"), ("Radna ploča", "वर्कटॉप"),
        ("Leđna ploča", "बैक पैनल"), ("Parcijalna leđna ploča", "आंशिक बैक पैनल"),
        ("Back panel / opening", "बैक पैनल / ओपनिंग"), ("Leđa / prolaz", "बैक पैनल / ओपनिंग"),
        ("Leva strana", "बायां साइड"), ("Desna strana", "दायां साइड"),
        ("Srednja vertikala", "मध्य वर्टिकल"), ("Polica (podesiva)", "एडजस्टेबल शेल्फ"),
        ("Polica", "शेल्फ"), ("Plafon", "टॉप पैनल"), ("Dno", "बॉटम पैनल"),
        ("Dno sanduka fioke", "दराज़ का बॉटम"), ("Prednja strana sanduka fioke", "दराज़ का फ्रंट"),
        ("Zadnja strana sanduka fioke", "दराज़ का बैक"), ("Bočna stranica sanduka fioke", "दराज़ का साइड"),
        ("Front fioke", "दराज़ फ्रंट"), ("Vrata", "दरवाज़े"), ("Sokla", "सोकल"),
        ("Filer panel", "फिलर पैनल"), ("Završna bočna ploča", "एंड साइड पैनल"),
        ("Šarka", "हिंग"), ("Ručka / pull", "हैंडल"), ("Klizač za fioku", "दराज़ स्लाइड"),
        ("Konfirmat vijak", "कन्फर्मेट स्क्रू"), ("Drvena tipla", "लकड़ी डॉवेल"),
        ("Sanitarni silikon", "सैनिटरी सिलिकॉन"), ("Sudopera", "सिंक"), ("Slavina", "नल"),
        ("Ugradna sudopera", "बिल्ट-इन सिंक"), ("Wall plugs / anchors", "वॉल प्लग / एंकर"),
        ("Purchased separately", "अलग से खरीदें"), ("Tools needed on site", "साइट पर ज़रूरी औज़ार"),
        ("Wall installation", "दीवार इंस्टॉलेशन"), ("Po šablonu proizvođača", "निर्माता के टेम्पलेट के अनुसार"),
        ("Po meri iz projekta", "प्रोजेक्ट माप के अनुसार"), ("Po geometriji zida i šablonu", "दीवार माप और टेम्पलेट के अनुसार"),
        ("Priprema i finalni rez", "तैयारी और अंतिम कट"), ("Otvor za sudoperu", "सिंक कट-आउट"),
        ("Otvor za ploču", "हॉब कट-आउट"), ("Utor za leđa", "बैक पैनल ग्रूव"),
        ("Ventilacija / otvor", "वेंटिलेशन / ओपनिंग"), ("Posebna obrada", "स्पेशल प्रोसेसिंग"),
        ("Servis", "वर्कशॉप"), ("Kuća / lice mesta", "साइट पर"),
    ]
    note_replacements: list[tuple[str, str]] = [
        ("In the workshop, do cutting and edging strictly by CUT dimensions. Verify openings and special machining against the notes.", "वर्कशॉप में CUT माप के अनुसार ही कटिंग और एजिंग करें, और नोट के अनुसार ओपनिंग तथा स्पेशल प्रोसेसिंग जांचें।"),
        ("In the workshop, cut strictly by CUT dimensions and verify edging in the separate table", "वर्कशॉप में केवल CUT माप के अनुसार कटिंग करें और अलग तालिका से एजिंग जांचें"),
        ("Take only cutting, edging and machining to the workshop", "वर्कशॉप में केवल कटिंग, एजिंग और प्रोसेसिंग वाला भाग ले जाएँ"),
        ("The workshop should work only from the workshop packet tables", "वर्कशॉप केवल वर्कशॉप पैकेट की तालिकाओं से काम करे"),
        ("Purchase ready-made appliances, hardware and tools separately", "तैयार उपकरण, हार्डवेयर और औज़ार अलग से खरीदें"),
        ("These items are not part of cutting and must be sourced separately", "ये आइटम कटिंग का हिस्सा नहीं हैं और अलग से लेने होंगे"),
        ("Back panel groove", "बैक पैनल ग्रूव"), ("Sink cut-out", "सिंक कट-आउट"),
        ("Hob cut-out", "हॉब कट-आउट"), ("Ventilation / opening", "वेंटिलेशन / ओपनिंग"),
        ("Preparation and final cut", "तैयारी और अंतिम कट"), ("Workshop + on site", "वर्कशॉप + साइट"),
        ("Workshop", "वर्कशॉप"), ("On site", "साइट पर"),
        ("According to wall geometry and template", "दीवार की ज्योमेट्री और टेम्पलेट के अनुसार"),
        ("Wall requirement / finished dimension", "दीवार की आवश्यकता / फिनिश माप"),
        ("CUT basis / purchase segment", "CUT आधार / खरीद लंबाई"),
        ("Workshop works strictly by CUT dimensions.", "वर्कशॉप केवल CUT माप के अनुसार काम करती है।"),
        ("Final cut is done on site.", "अंतिम कट साइट पर किया जाता है।"),
    ]
    if col in {"Napomena", "Napomena za servis", "Instrukcija", "Stavka", "Sta radis", "Šta radiš", "Polje", "Vrednost", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena", "Grupa", "Tip / Šifra", "Note", "Instruction", "Item", "Field", "Value", "Processing type", "Operations", "Execution basis", "Processing / note", "Group", "Type / Code"}:
        txt = _apply_ordered_replacements(txt, note_replacements)
    return _apply_ordered_replacements(txt, common)


def _translate_export_text(value: Any, lang: str = "sr", column: str = "") -> Any:
    txt = str(value or "")
    col = str(column or "")
    _lang = str(lang or "sr").lower().strip()

    if _lang != "en":
        replacements = {
            "ploca za kuvanje": "ploča za kuvanje",
            "plocu za kuvanje": "ploču za kuvanje",
            "masina za sudove": "mašina za sudove",
            "samostojeca": "samostojeća",
            "frizider": "frižider",
            "zamrzivac": "zamrzivač",
            "spajz": "špajz",
            "Otvor za plocu": "Otvor za ploču",
            "sarke": "šarke",
            "sarki": "šarki",
            "busenje:": "bušenje:",
            "O35mm": "Ø35mm",
            "sablonu": "šablonu",
            "radnoj ploci": "radnoj ploči",
            "radnu plocu": "radnu ploču",
            "radna ploca": "radna ploča",
            "Radna ploca": "Radna ploča",
            "sina": "šina",
            "Sipka": "Šipka",
            "sipka": "šipka",
            "vesanje": "vešanje",
            "odece": "odeće",
            "nosaca": "nosača",
            "Nosaci": "Nosači",
            "nosaci": "nosači",
            "nosac": "nosač",
            "duzina": "dužina",
            "sirina": "širina",
        }
        for src, dst in replacements.items():
            txt = txt.replace(src, dst)
        txt = txt.replace("višina", "visina").replace("Višina", "Visina")
        if _lang == "es":
            return _translate_export_text_es(txt, col)
        if _lang == "ru":
            return _translate_export_text_ru(txt, col)
        if _lang == "pt-br":
            return _translate_export_text_ptbr(txt, col)
        if _lang == "zh-cn":
            return _translate_export_text_zhcn(txt, col)
        if _lang == "hi":
            return _translate_export_text_hi(txt, col)
        return txt

    if col in {"Zid", "Wall"}:
        return txt.replace("Zid ", "Wall ")

    if col in {"Modul", "Module"}:
        txt = txt.replace("Sokla - Zid ", "Plinth - Wall ")
        txt = txt.replace("Radna ploča - Zid ", "Worktop - Wall ")
        txt = txt.replace("Radna ploca - Zid ", "Worktop - Wall ")
        txt = txt.replace("Modul ", "Module ")
        txt = txt.replace("Zid ", "Wall ")
        txt = txt.replace("Donji (fioke)", "Base (drawers)")
        txt = txt.replace("Donji uski (flaše/ulja/začini)", "Base narrow (bottles/oils/spices)")
        txt = txt.replace("Donji uski (flase/ulja/zacini)", "Base narrow (bottles/oils/spices)")
        txt = txt.replace("Donji otvoreni (bez vrata)", "Base open (no doors)")
        txt = txt.replace("Donji (1 vrata)", "Base (1 door)")
        txt = txt.replace("Donji (2 vrata)", "Base (2 doors)")
        txt = txt.replace("Donji (vrata + fioka)", "Base (doors + drawer)")
        txt = txt.replace("Donji (sudopera)", "Base (sink)")
        return txt

    if col in {"Pozicija", "Position"}:
        mapping = {
            "LEVO": "LEFT",
            "DESNO": "RIGHT",
            "GORE": "TOP",
            "DOLE": "BOTTOM",
            "CENTAR": "CENTER",
            "SREDINA": "CENTER",
            "PREDNJA": "FRONT",
            "ZADNJA": "BACK",
            "NAPRED": "FRONT",
            "POZADI": "BACK",
        }
        return mapping.get(txt.upper(), txt)

    if col in {"Kategorija", "Category"}:
        mapping = {
            "okov": "hardware",
            "potrosni": "consumable",
            "warning": "warning",
        }
        return mapping.get(txt.lower(), txt)

    if col in {"Orijentacija", "Orientation"}:
        mapping = {
            "horizontalna": "horizontal",
            "horizontalno": "horizontal",
            "vertikalna": "vertical",
            "vertikalno": "vertical",
        }
        return mapping.get(txt.lower(), txt)

    if col in {"Naziv", "Name"} and txt.startswith("UPOZORENJE"):
        return txt.replace("UPOZORENJE", "WARNING")

    if col in {"Deo", "Naziv", "Materijal", "Element", "Part", "Name", "Material", "Type / Code"}:
        replacements = [
            ("Iverica", "Chipboard"),
            ("Šper ploča", "Plywood"),
            ("Sper ploca", "Plywood"),
            ("Puno drvo", "Solid wood"),
            ("Akril", "Acrylic"),
            ("Furnir", "Veneer"),
            ("Lakobel", "Lacobel"),
            ("Radna ploča", "Worktop"),
            ("Radna ploča", "Worktop"),
            ("Leva strana", "Left side"),
            ("Leva stranica korpusa", "Left carcass side"),
            ("Desna strana", "Right side"),
            ("Desna stranica korpusa", "Right carcass side"),
            ("Dno sanduka fioke", "Drawer box bottom"),
            ("Dno sanduka", "Drawer box bottom"),
            ("Donja ploča korpusa", "Bottom carcass panel"),
            ("Donja ploca korpusa", "Bottom carcass panel"),
            ("Dno", "Bottom panel"),
            ("Plafon", "Top panel"),
            ("Gornja ploča korpusa", "Top carcass panel"),
            ("Gornja ploca korpusa", "Top carcass panel"),
            ("Srednja vertikala", "Center upright"),
            ("Polica (podesiva)", "Adjustable shelf"),
            ("Polica", "Shelf"),
            ("Horizontalna pregrada", "Horizontal divider"),
            ("Parcijalna leđna ploča", "Partial back panel"),
            ("Parcijalna ledjna ploca", "Partial back panel"),
            ("Leđna ploča — Krak 1", "Back panel - Arm 1"),
            ("Leđna ploča — Krak 2", "Back panel - Arm 2"),
            ("Leđna ploča", "Back panel"),
            ("Ledjna ploca", "Back panel"),
            ("Front vrata", "Door front"),
            ("Vrata (ispod ploče za kuvanje)", "Doors below hob"),
            ("Vrata (ispod sudopere)", "Doors below sink"),
            ("Vrata rerne", "Oven front"),
            ("Front za rernu", "Oven front"),
            ("Vrata", "Doors"),
            ("Front fioke (kuhinjska jedinica)", "Drawer front below oven"),
            ("Front fioke (ispod rerne)", "Drawer front below oven"),
            ("Front fioke (unif.)", "Drawer front"),
            ("Front fioke", "Drawer front"),
            ("Front uskog izvlačnog modula", "Narrow pull-out front"),
            ("Ugaoni front", "Corner front"),
            ("Klizna vrata — američki plakar", "Sliding doors - American wardrobe"),
            ("Klizna vrata — ormar", "Sliding wardrobe doors"),
            ("Staklena vrata", "Glass doors"),
            ("Front sortirnika", "Waste sorter front"),
            ("Donji servisni front", "Lower service front"),
            ("Front — mašina za sudove", "Dishwasher front"),
            ("Front integrisanog frižidera", "Integrated fridge front"),
            ("Gornji front frižidera", "Upper fridge front"),
            ("Donji front zamrzivača", "Lower freezer front"),
            ("Završna bočna ploča", "End side panel"),
            ("Zavrsna bocna ploca", "End side panel"),
            ("Bočna ploča", "Drawer box side"),
            ("Bocna ploca", "Drawer box side"),
            ("Bočna stranica sanduka fioke", "Drawer box side"),
            ("Bocna stranica sanduka fioke", "Drawer box side"),
            ("Prednja strana sanduka fioke", "Drawer box front"),
            ("Prednja strana sanduka", "Drawer box front"),
            ("Zadnja strana sanduka fioke", "Drawer box back"),
            ("Zadnja strana sanduka", "Drawer box back"),
            ("Vezna letva — MZS (ugradna)", "Cross rail - built-in dishwasher"),
            ("Sokla (lajsna)", "Plinth (toe kick)"),
            ("Radna ploča", "Worktop"),
            ("Radna ploča", "Worktop"),
            ("Nosač radne ploče", "Worktop support"),
            ("Nosač radne ploce", "Worktop support"),
            ("Leđa / prolaz", "Back panel / opening"),
            ("Zona priključka", "Connection zone"),
            ("Ventilacija kolone", "Tall-unit ventilation"),
            ("Drvena tipla", "Wooden dowel"),
            ("Konfirmat vijak", "Confirmat screw"),
            ("Klizač za fioku", "Drawer slide"),
            ("Klizac za fioku", "Drawer slide"),
            ("Nosač police (klin)", "Shelf support pin"),
            ("Nosac police (klin)", "Shelf support pin"),
            ("Ručka / pull", "Handle / pull"),
            ("Stopica (nogica)", "Cabinet leg"),
            ("Uski izvlačni mehanizam", "Narrow pull-out mechanism"),
            ("Klipsa za soklu", "Plinth clip"),
            ("Spojnica susednih korpusa", "Connector for adjacent cabinets"),
            ("Vijak / ekser za leđa", "Screw / pin for back panel"),
            ("Vijak / ekser za leđa", "Screw / pin for back panel"),
            ("Vijak / ugaonik za radnu ploču", "Screw / bracket for worktop"),
            ("Vijak za klizač", "Screw for drawer slide"),
            ("Vijak za ručku", "Handle screw"),
            ("Zaptivna lajsna / silikon uz zid", "Sealing trim / silicone along wall"),
            ("Po izboru korisnika", "Selected by the customer"),
            ("Podesiva h=100 mm", "Adjustable h=100 mm"),
            ("Providna / sivi", "Transparent / grey"),
            ("Providan / sivi", "Transparent / grey"),
            ("Sortirnik / pull-out set", "Waste sorter / pull-out set"),
            ("2x kanta + nosac", "2x bins + holder"),
            ("2x kanta + nosač", "2x bins + holder"),
            ("Baterija za sudoperu", "Sink tap"),
            ("Ugradna sudopera", "Built-in sink"),
            ("Sifon + preliv + spojnice", "Trap + overflow + connectors"),
            ("Sifon i odvodni set", "Trap and drain set"),
            ("Slavina", "Tap"),
            ("Sudopera", "Sink"),
            ("Postolje / nosač — MZS", "Dishwasher platform / support"),
            ("Postolje / nosač - MZS", "Dishwasher platform / support"),
            ("Postolje / nosac — MZS", "Dishwasher platform / support"),
            ("Postolje / nosac - MZS", "Dishwasher platform / support"),
            ("Donja maska — MZS", "Dishwasher lower filler"),
            ("Donja maska - MZS", "Dishwasher lower filler"),
            ("Zidna lajsna ili sanitarni silikon", "Wall trim or sanitary silicone"),
            ("Spojnica korpusa + vijak", "Cabinet connector + screw"),
            ("Vijak / ugaonik za radnu ploču", "Screw / bracket for worktop"),
            ("M4 x 50mm vijak za ručku", "M4 x 50 mm handle screw"),
            ("Tipl + vijak 8x80 mm", "Wall plug + screw 8x80 mm"),
            ("Nosač + šina za viseće elemente", "Bracket + rail for wall units"),
            ("Blum CLIP-top 110° (71B...)", "Blum CLIP-top 110° (71B...)"),
            ("Kupuje se kao gotov kupovni izvlačni program", "Purchased as a ready-made pull-out system"),
            ("Osnovno pričvršćenje leđa", "Basic back-panel fixing"),
            ("Osnovni set", "Basic set"),
            ("Završno zaptivanje", "Final sealing"),
            ("Ledjna ploca", "Leđna ploča"),
            ("Nosač radne ploce", "Nosač radne ploče"),
            ("Vijak za ručku", "Vijak za ručku"),
            ("M4 x 50mm vijak za ručku", "M4 x 50mm vijak za ručku"),
            ("Zastita", "Zaštita"),
            ("iskljucivo", "isključivo"),
            ("pricvrscenje", "pričvršćenje"),
            ("nosace", "nosače"),
            ("nosaca", "nosača"),
            ("klizaca", "klizača"),
            ("busenje:", "bušenje:"),
            ("O35mm dubina", "Ø35mm dubina"),
            ("ukljuceni", "uključeni"),
            ("montazna ploca", "montažna ploča"),
            ("radne ploce", "radne ploče"),
            ("ploce na nosace", "ploče na nosače"),
            ("po ručki", "po ručki"),
            ("sinu ili", "šinu ili"),
            ("za pricvrscenje", "za pričvršćenje"),
            ("duzina", "dužina"),
        ]
        out = txt
        for src, dst in replacements:
            out = out.replace(src, dst)
        return out

    if col in {"Napomena", "Napomena za servis", "Instrukcija", "Stavka", "Sta radis", "Šta radiš", "Polje", "Note", "Instruction", "Item", "Field"}:
        replacements = [
            ("Element koristi nepoznat template_id '", "The unit uses an unknown template_id '"),
            ("Proveri module_templates.json ili zameni element važećim template-om pre eksporta.", "Check module_templates.json or replace the unit with a valid template before export."),
            ("Proveri module_templates.json ili zameni element važećim templateom pre eksporta.", "Check module_templates.json or replace the unit with a valid template before export."),
            ("Element referiše na nepostojeći zid '", "The unit references a non-existent wall '"),
            ("Postavi wall_key na jedan od važećih zidova: ", "Set wall_key to one of the valid walls: "),
            ("Element ima nevažeće dimenzije", "The unit has invalid dimensions"),
            ("Proveri širinu, visinu i dubinu elementa pre eksporta.", "Check the unit width, height, and depth before export."),
            ("Preporuka:", "Recommendation:"),
            ("Kupovni vodoinstalaterski set za sudoperu", "Purchased plumbing set for the sink"),
            ("Kupuje se posebno prema izboru korisnika", "Purchased separately according to customer choice"),
            ("Kupuje se kao gotov proizvod; otvor se reže u radnoj ploči prema šablonu", "Purchased as a ready-made product; the opening is cut into the worktop according to the template"),
            ("Kupuje se kao gotov proizvod; otvor se reze u radnoj ploci prema sablonu", "Purchased as a ready-made product; the opening is cut into the worktop according to the template"),
            ("Osnovno pricvrscenje", "Basic fixing"),
            ("Osnovno pričvršćenje", "Basic fixing"),
            ("Za pricvrscenje", "For fixing"),
            ("Kupuje se posebno", "Purchased separately"),
            ("Kupuje se kao gotov proizvod", "Purchased as a ready-made product"),
            ("po izboru korisnika", "according to customer choice"),
            ("Front fuga je ", "Front gap is "),
            ("(van preporuke 1.0–4.0 mm).", "(outside the recommended 1.0-4.0 mm range)."),
            ("Postavi front_gap na 2.0 mm za stabilnu montažu.", "Set front_gap to 2.0 mm for stable assembly."),
            ("Dubina visećeg elementa ", "Wall unit depth "),
            (" je velika za ergonomiju.", " is too deep for ergonomic use."),
            ("Preporuka je 300–380 mm dubine za gornje elemente.", "Recommended wall-unit depth is 300-380 mm."),
            ("Dubina donjeg elementa ", "Base unit depth "),
            (" je mala za standardne uređaje/sudoperu.", " is too small for standard appliances or a sink."),
            ("Preporuka je min 560 mm za kuhinjske donje module.", "Recommended minimum depth for kitchen base units is 560 mm."),
            ("Dubina visokog elementa ", "Tall unit depth "),
            (" je mala za stabilan kuhinjski korpus.", " is too small for a stable tall cabinet."),
            ("Preporuka je min 560 mm za visoke elemente.", "Recommended minimum depth for tall units is 560 mm."),
            ("Visoki element ", "Tall unit "),
            (" je blizu/iznad raspoložive visine.", " is close to or above the available height."),
            ("Smanji visinu ili proveri realnu visinu plafona i stopice.", "Reduce the height or verify the real ceiling and leg height."),
            ("Element drugog reda nema pun oslonac na gornjem elementu ispod sebe.", "The second-row wall unit does not have full support on the wall unit below."),
            ("Poravnaj ga tako da cela širina leži na elementu ispod.", "Align it so the full width sits on the unit below."),
            ("Popuna iznad visokog nema pun oslonac na visokom elementu ispod sebe.", "The infill above the tall unit does not have full support on the tall unit below."),
            ("Poravnaj je tako da cela širina leži na elementu ispod.", "Align it so the full width sits on the unit below."),
            ("Fioke na dubini ", "Drawer unit at depth "),
            (" mogu imati ograničen izbor klizača.", " may have a limited choice of drawer slides."),
            ("Koristi kraće klizače (npr. 350/400) ili povećaj dubinu korpusa.", "Use shorter slides (for example 350/400) or increase cabinet depth."),
            ("Lift-up front širine ", "Lift-up front with width "),
            (" može biti pretežak.", " may be too heavy."),
            ("Razdvoj front na 2 segmenta ili koristi jači/međupodupirač.", "Split the front into 2 segments or use stronger hardware/support."),
            ("Širina za mašinu za sudove ", "Dishwasher niche width "),
            (" je manja od standardnih 600 mm.", " is smaller than the standard 600 mm."),
            ("Povećaj širinu na najmanje 600 mm.", "Increase the width to at least 600 mm."),
            ("Širina za frižider ", "Fridge niche width "),
            ("Širina za rernu/ploču ", "Oven/hob niche width "),
            ("Samostalna ploča za kuvanje širine ", "Standalone hob width "),
            (" je manja od minimalnih 450 mm.", " is below the minimum 450 mm."),
            ("Standardne ploče su 45 cm, 60 cm ili 90 cm — povećaj širinu.", "Standard hobs are 45 cm, 60 cm or 90 cm - increase the width."),
            ("Širina visoke appliance kolone ", "Tall appliance column width "),
            ("Sudoperski element širine ", "Sink base unit width "),
            (" je suviše mali za stabilan laički workflow.", " is too small for a stable standard workflow."),
            ("Koristi najmanje 600 mm širine.", "Use at least 600 mm of width."),
            ("Ugaoni element širine ", "Corner unit width "),
            (" je rizičan za stabilan raspored i pristup uglu.", " is risky for a stable layout and corner access."),
            ("Povećaj širinu na najmanje 800 mm.", "Increase the width to at least 800 mm."),
            ("Ugaoni element nije naslonjen na unutrasnji ugao zida A.", "The corner unit is not aligned to the inside corner of Wall A."),
            ("Postavi ga kao poslednji element desno, uz unutrasnji ugao.", "Place it as the last unit on the right, against the inside corner."),
            ("Ugaoni element nije naslonjen na unutrasnji ugao zida ", "The corner unit is not aligned to the inside corner of Wall "),
            ("Postavi ga kao prvi element ", "Place it as the first unit "),
            (" uz unutrasnji ugao.", " against the inside corner."),
            ("Širina modula za napu/mikrotalasnu ", "Hood/microwave unit width "),
            ("Dubina modula za napu/mikrotalasnu ", "Hood/microwave unit depth "),
            (" je mala za uređaj i ventilaciju.", " is too small for the appliance and ventilation."),
            ("Povećaj dubinu na najmanje 300 mm.", "Increase the depth to at least 300 mm."),
            ("Dubina samostojećeg uređaja ", "Freestanding appliance depth "),
            (" je manja od praktičnog minimuma 580 mm.", " is below the practical minimum of 580 mm."),
            ("Povećaj dubinu na najmanje 580 mm.", "Increase the depth to at least 580 mm."),
            ("Dubina samostojećeg frižidera ", "Freestanding fridge depth "),
            (" je manja od 600 mm.", " is below 600 mm."),
            ("Povećaj dubinu na najmanje 600 mm.", "Increase the depth to at least 600 mm."),
            ("Dubina visoke appliance kolone ", "Tall appliance column depth "),
            (" je manja od 560 mm.", " is below 560 mm."),
            ("Povećaj dubinu na najmanje 560 mm.", "Increase the depth to at least 560 mm."),
            ("Jednokrilni front širine ", "Single-door front width "),
            (" može biti nestabilan i težak za podešavanje.", " may be unstable and hard to adjust."),
            ("Podeli element na 2 vrata ili smanji širinu.", "Split the unit into 2 doors or reduce the width."),
            ("Jednokrilna vrata su preblizu levom zidu na strani sarke.", "The single door is too close to the left wall on the hinge side."),
            ("Promeni stranu rucke ili dodaj filer 30-50 mm uz levi zid.", "Change the handle side or add a 30-50 mm filler next to the left wall."),
            ("Jednokrilna vrata su preblizu desnom zidu na strani sarke.", "The single door is too close to the right wall on the hinge side."),
            ("Promeni stranu rucke ili dodaj filer 30-50 mm uz desni zid.", "Change the handle side or add a 30-50 mm filler next to the right wall."),
            ("Bar jedna fioka je niža od 80 mm, što je rizično za front i rukovanje.", "At least one drawer front is lower than 80 mm, which is risky for the front and handling."),
            ("Povećaj najmanju fioku na barem 80 mm.", "Increase the smallest drawer to at least 80 mm."),
            ("Zbir visina fioka ", "Total drawer heights "),
            (" je vrlo blizu visine modula ", " are very close to the unit height "),
            ("Ostavi tehničku rezervu 10–20 mm za fuge i frontove.", "Leave a 10-20 mm technical allowance for gaps and fronts."),
            ("Vrata u kombinaciji vrata + fioka imaju samo ", "In the doors + drawer combination, the doors are only "),
            (" mm visine.", " mm high."),
            ("Povećaj vrata na najmanje 180 mm.", "Increase the doors to at least 180 mm."),
            ("Vrata u kombinaciji vrata + fioka uzimaju skoro celu visinu modula (", "In the doors + drawer combination, the doors take almost the full unit height ("),
            ("Ostavi barem 120 mm za fioku i tehničke fuge.", "Leave at least 120 mm for the drawer and technical gaps."),
            ("Proveri pristup instalaciji i ostavi servisni prostor.", "Check installation access and leave service space."),
            ("Raised dishwasher", "Raised dishwasher"),
            ("raised dishwasher", "raised dishwasher"),
            ("Donja maska iznad sokle za raised dishwasher; ", "Lower filler above plinth for raised dishwasher; "),
            ("Postolje/podizanje mašine za sudove ", "Dishwasher platform / lift "),
            ("Postolje/podizanje mašine za sudove", "Dishwasher platform / lift"),
            ("donja maska ", "lower filler "),
            ("uz kontinualnu soklu", "with continuous plinth"),
            ("Donji elementi na zidu ", "Base units on Wall "),
            (" nisu u istoj ravni za kontinualnu radnu ploču.", " are not aligned for a continuous worktop."),
            (" nisu u istoj ravni za kontinualnu radnu plocu.", " are not aligned for a continuous worktop."),
            ("Top-level raspon je ", "Top-level range is "),
            ("Poravnaj visine svih BASE elemenata pre radne ploče ili koristi poseban segment/raskid ploče.", "Align all BASE unit heights before the worktop or use a separate segment / worktop break."),
            ("Poravnaj visine svih BASE elemenata pre radne ploce ili koristi poseban segment/raskid ploce.", "Align all BASE unit heights before the worktop or use a separate segment / worktop break."),
            ("Potrebna dužina radne ploče za zid ", "Required worktop length for wall "),
            ("Potrebna duzina radne ploce za zid ", "Required worktop length for wall "),
            (" a najveća standardna nabavna dužina je ", " while the longest standard purchase length is "),
            (" a najveca standardna nabavna duzina je ", " while the longest standard purchase length is "),
            ("Dodaj spoj/raskid ploče, koristi duži komercijalni format ili naruči specijalnu radnu ploču.", "Add a worktop joint/break, use a longer commercial stock length, or order a custom worktop."),
            ("Dodaj spoj/raskid ploce, koristi duzi komercijalni format ili naruci specijalnu radnu plocu.", "Add a worktop joint/break, use a longer commercial stock length, or order a custom worktop."),
            ("Izrez za sudoperu izlazi van radne ploče/modula", "The sink cut-out extends outside the worktop/unit"),
            ("Izrez za sudoperu izlazi van radne ploce/modula", "The sink cut-out extends outside the worktop/unit"),
            ("Izrez za ploču za kuvanje izlazi van radne ploče/modula", "The hob cut-out extends outside the worktop/unit"),
            ("Izrez za plocu za kuvanje izlazi van radne ploce/modula", "The hob cut-out extends outside the worktop/unit"),
            ("Smanji izrez ili ga pomeri tako da ostane unutar širine i dubine modula.", "Reduce the cut-out or move it so it stays within the unit width and depth."),
            ("Smanji izrez ili ga pomeri tako da ostane unutar sirine i dubine modula.", "Reduce the cut-out or move it so it stays within the unit width and depth."),
            ("Filer panel širine ", "Filler panel width "),
            ("Filer panel sirine ", "Filler panel width "),
            (" je širok za standardnu popunu prostora.", " is wide for a standard filler application."),
            (" je sirok za standardnu popunu prostora.", " is wide for a standard filler application."),
            ("Razmotri pravi završni panel ili poseban uski modul umesto širokog filera.", "Consider a proper end panel or a dedicated narrow unit instead of a wide filler."),
            ("Razmotri pravi zavrsni panel ili poseban uski modul umesto sirokog filera.", "Consider a proper end panel or a dedicated narrow unit instead of a wide filler."),
            ("Element izlazi van zida ", "The unit extends beyond Wall "),
            (" > duzina zida ", " > wall length "),
            ("Pomeri element ulevo ili smanjite mu sirinu da stane u raspored zida.", "Move the unit left or reduce its width so it fits the wall layout."),
            ("Elementi se preklapaju u zoni ", "Units overlap in zone "),
            (" na zidu ", " on Wall "),
            ("Pomeri elemente tako da im se X-rasponi ne preklapaju.", "Move the units so their X ranges do not overlap."),
            ("Ugaoni element i susedni modul nalezu bez servisnog razmaka: ", "The corner unit and adjacent module touch without a service gap: "),
            ("Proveri frontove i rucke u otvaranju; po potrebi dodaj filer ili tehnicki razmak.", "Check door fronts and handles when opening; add a filler or technical gap if needed."),
            ("Jednokrilni sused '", "Single-door adjacent unit '"),
            ("' otvara vrata ka uglu pored ugaonog modula.", "' opens toward the corner next to the corner unit."),
            ("Promeni stranu rucke ili koristi dvokrilni/fiokar pored ugla.", "Change the handle side or use a double-door or drawer unit next to the corner."),
            ("' je preširok uz ugaoni modul (", "' is too wide next to the corner unit ("),
            ("Koristi max ", "Use a maximum of "),
            ("dvokrilni element, fiokar ili ostavi filer uz ugao.", "a double-door unit, a drawer unit, or leave a filler near the corner."),
            ("Nema podataka.", "No data."),
            ("Nema podataka po modulima.", "No unit data."),
            ("Segment zida ", "Wall segment "),
            ("raspon ", "range "),
            ("1 segment na zidu ", "1 segment on wall "),
            ("ukupna dužina base zone ", "total base-zone length "),
            ("ukupna duzina base zone ", "total base-zone length "),
            ("radna ploča počinje na x=", "worktop starts at x="),
            ("radna ploca pocinje na x=", "worktop starts at x="),
            ("radna ploča dužine ", "worktop length "),
            ("radna ploca duzine ", "worktop length "),
            ("Ugaono spajanje s Zidom A", "Corner joint with Wall A"),
            ("rezati/stepenasto obradi levu stranu", "cut / stepped-machine the left side"),
            ("rezati/stepenasto obradi desnu stranu", "cut / stepped-machine the right side"),
            ("dubine Zida A", "of Wall A depth"),
            ("Dužina segmenta: ", "Segment length: "),
            ("Duzina segmenta: ", "Segment length: "),
            ("parcijalna leđa samo u donjoj servisnoj zoni, gornja zona iza rerne ostaje otvorena radi ventilacije i priključaka", "partial back panel only in the lower service zone; the upper zone behind the oven remains open for ventilation and connections"),
            ("ostaviti otvor za dovod/odvod vode prema poziciji instalacija", "leave opening for water supply/drain according to installation position"),
            ("ostaviti otvor/prolaz za odvod nape prema modelu uređaja", "leave opening/duct path for hood according to appliance model"),
            ("ostaviti otvor za kabl i ventilaciju mikrotalasne", "leave opening for cable and microwave ventilation"),
            ("utor 8mm", "8 mm back-panel groove"),
            ("Gornji vez", "Top rail"),
            ("pričvršćuje radnu ploču", "secures the worktop"),
            ("pricvrscuje radnu plocu", "secures the worktop"),
            ("prednja ivica kantovana", "front edge edged"),
            ("radna ploča", "worktop"),
            ("radnoj ploci", "worktop"),
            ("radna ploca", "worktop"),
            ("zidu ", "wall "),
            ("ploču za kuvanje", "hob"),
            ("ploča za kuvanje", "hob"),
            ("sudoperu", "sink"),
            ("sudopere", "sink"),
            ("slavina", "tap"),
            ("sifon", "trap"),
            ("ugradnu mikrotalasnu", "built-in microwave"),
            ("ugradnu rernu", "built-in oven"),
            ("ugradne MZS", "built-in dishwasher"),
            ("ugradnog frižidera", "integrated fridge"),
            ("frontovi", "fronts"),
            ("korpuse", "carcasses"),
            ("okove", "hardware"),
            ("uredjaje", "appliances"),
            ("servis", "workshop"),
            ("secenje", "cutting"),
            ("kantovanje", "edging"),
            ("obradu", "machining"),
            ("obrade", "machining"),
            ("montazu", "assembly"),
            ("kuce", "site"),
            ("lice mesta", "on site"),
            ("zidne elemente", "wall units"),
            ("visoke", "tall units"),
            ("proveriti", "verify"),
            ("proveri", "check"),
            ("izabrati", "choose"),
            ("dodaj BASE elemente", "add BASE units"),
        ]
        out = txt
        for src, dst in replacements:
            out = out.replace(src, dst)
        out = out.replace("Zid ", "Wall ")
        out = out.replace("Modul ", "Module ")
        out = out.replace("lajsna", "trim")
        return out

    return txt


def _normalize_sr_export_text(value: Any) -> str:
    return str(_translate_export_text(value, "sr") or "")


def _format_material_role(material: Any, thickness: Any, role: str, lang: str = "sr") -> str:
    _lang = str(lang or "sr").lower().strip()
    mat = str(material or "").strip()
    thk = str(thickness or "").strip()
    role_map_sr = {"carcass": "Korpus", "front": "Front", "back": "Leđa"}
    role_map_en = {"carcass": "Carcass", "front": "Front", "back": "Back"}
    role_map_es = {"carcass": "Cuerpo", "front": "Frente", "back": "Fondo"}
    role_map_ru = {"carcass": "Корпус", "front": "Фасад", "back": "Задник"}
    role_map_pt = {"carcass": "Corpo", "front": "Frente", "back": "Fundo"}
    role_map_zh = {"carcass": "柜体", "front": "门板", "back": "背板"}
    role_map_hi = {"carcass": "कार्कस", "front": "फ्रंट", "back": "बैक"}
    if _lang == "es":
        mat = mat.replace("Iverica", "Aglomerado").replace("Lesonit", "HDF").replace("Radna ploča", "Encimera")
    if _lang == "ru":
        mat = mat.replace("Iverica", "ЛДСП").replace("Lesonit", "ХДФ").replace("Radna ploča", "Столешница")
    if _lang == "pt-br":
        mat = mat.replace("Iverica", "Aglomerado").replace("Radna ploča", "Bancada")
    if _lang == "zh-cn":
        mat = mat.replace("Iverica", "刨花板").replace("Lesonit", "HDF").replace("Radna ploča", "台面")
    if _lang == "hi":
        mat = mat.replace("Iverica", "पार्टिकल बोर्ड").replace("Lesonit", "HDF").replace("Radna ploča", "वर्कटॉप")
    if _lang == "en":
        role_label = role_map_en.get(role, role)
    elif _lang == "es":
        role_label = role_map_es.get(role, role)
    elif _lang == "ru":
        role_label = role_map_ru.get(role, role)
    elif _lang == "pt-br":
        role_label = role_map_pt.get(role, role)
    elif _lang == "zh-cn":
        role_label = role_map_zh.get(role, role)
    elif _lang == "hi":
        role_label = role_map_hi.get(role, role)
    else:
        role_label = role_map_sr.get(role, role)
    if not mat:
        return f"{role_label} / {thk} mm" if thk else role_label
    if role == "back":
        if mat.upper() == "HDF":
            mat = "HDF" if _lang == "en" else "HDF"
        elif mat.lower() == "lesonit":
            mat = "Lesonit"
    mat_label = f"{mat} {role_label}"
    return f"{mat_label} / {thk} mm" if thk else mat_label


def _material_role_from_part_name(part_name: Any) -> str:
    part = str(part_name or "").lower()
    if any(k in part for k in ("leđ", "ledj", "ledn", "back", "fundo", "traseir")):
        return "back"
    if any(k in part for k in ("front", "vrata", "frente", "porta")):
        return "front"
    return "carcass"


def _translate_export_df(df: pd.DataFrame | None, lang: str = "sr") -> pd.DataFrame | None:
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in (
        "Zid", "Wall",
        "Modul", "Module",
        "Deo", "Part",
        "Pozicija", "Position",
        "Kategorija", "Category",
        "Naziv", "Name",
        "Materijal", "Material",
        "Orijentacija", "Orientation",
        "Element",
        "Type / Code",
        "Napomena", "Napomena za servis", "Note",
        "Instrukcija", "Instruction",
        "Stavka", "Item",
        "Sta radis",
        "Šta radiš",
        "Polje", "Field",
        "Vrednost", "Value",
        "Tip obrade", "Processing type",
        "Izvodi", "Operations",
        "Osnov izvođenja", "Execution basis",
        "Obrada / napomena", "Processing / note",
        "Grupa", "Group",
        "Tip / Šifra", "Type / Code",
    ):
        if col in out.columns:
            out[col] = out[col].map(lambda v, _c=col: _translate_export_text(v, lang, _c))
    return out


def _sanitize_export_value(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if value is None:
        return ""
    txt = str(value).strip()
    if txt.lower() in {"nan", "none", "null"}:
        return ""
    return value


def _sanitize_export_df(
    df: pd.DataFrame | None,
    *,
    require_positive_dims: bool = False,
    part_col: str = "Deo",
    material_col: str = "Materijal",
    width_candidates: tuple[str, ...] = ("Dužina [mm]", "CUT_W [mm]", "width_mm"),
    height_candidates: tuple[str, ...] = ("Širina [mm]", "CUT_H [mm]", "height_mm"),
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    out = df.copy()
    out = out.replace([pd.NA], "")
    out = out.where(pd.notna(out), "")
    for col in out.columns:
        out[col] = out[col].map(_sanitize_export_value)
    if require_positive_dims:
        _part_ok = out[part_col].astype(str).str.strip().ne("") if part_col in out.columns else pd.Series(True, index=out.index)
        _mat_ok = out[material_col].astype(str).str.strip().ne("") if material_col in out.columns else pd.Series(True, index=out.index)
        _w_ok = pd.Series(True, index=out.index)
        _h_ok = pd.Series(True, index=out.index)
        for _wc in width_candidates:
            if _wc in out.columns:
                _w_ok = pd.to_numeric(out[_wc], errors="coerce").fillna(0).gt(0)
                break
        for _hc in height_candidates:
            if _hc in out.columns:
                _h_ok = pd.to_numeric(out[_hc], errors="coerce").fillna(0).gt(0)
                break
        out = out[_part_ok & _mat_ok & _w_ok & _h_ok].copy()
    return out.reset_index(drop=True)


def get_final_cutlist_dataset(kitchen: Dict[str, Any], lang: str = "sr") -> Dict[str, Any]:
    _lang = str(lang or "sr").lower().strip()
    raw_sections = generate_cutlist(kitchen)
    sections: Dict[str, pd.DataFrame] = {}
    for key, df in (raw_sections or {}).items():
        _needs_dims = key in {"carcass", "backs", "fronts", "drawer_boxes", "worktop", "plinth"}
        sections[key] = _translate_export_df(
            _sanitize_export_df(df, require_positive_dims=_needs_dims),
            _lang,
        )
    service_packet_raw = build_service_packet(kitchen, sections, lang=_lang)
    service_packet: Dict[str, pd.DataFrame] = {}
    for key, df in (service_packet_raw or {}).items():
        _needs_dims = key in {"service_cuts", "service_edge"}
        service_packet[key] = _translate_export_df(
            _sanitize_export_df(
                df,
                require_positive_dims=_needs_dims,
                part_col="Deo",
                material_col="Materijal",
                width_candidates=("CUT_W [mm]", "Dužina [mm]", "width_mm"),
                height_candidates=("CUT_H [mm]", "Širina [mm]", "height_mm"),
            ),
            _lang,
        )
    summary = generate_cutlist_summary(sections, lang=_lang)
    for key, df in list(summary.items()):
        summary[key] = _translate_export_df(
            _sanitize_export_df(
                df,
                require_positive_dims=key in {"summary_all", "summary_detaljna", "summary_carcass", "summary_fronts", "summary_backs"},
                part_col="Deo",
                material_col="Materijal",
                width_candidates=("Dužina [mm]",),
                height_candidates=("Širina [mm]",),
            ),
            _lang,
        )
    return {
        "sections": sections,
        "service_packet": service_packet,
        "summary": summary,
    }


def _format_pdf_table_cell(value: Any, column_name: str | None = None) -> str:
    try:
        if pd.isna(value):
            value = ""
    except Exception:
        pass
    txt = str(value or "")
    if txt.strip().lower() in {"nan", "none", "null"}:
        txt = ""
    col = str(column_name or "").strip().lower()
    if col in {"zid", "wall", "cut_w [mm]", "cut_h [mm]", "kant", "edge"} and not txt.strip():
        return "-"
    note_cols = {
        "napomena",
        "napomena za servis",
        "obrada / napomena",
        "instrukcija",
        "objašnjenje",
        "objasnjenje",
        "note",
        "workshop note",
        "instruction",
        "explanation",
    }
    if col in {"sta radis", "šta radiš"} and txt:
        txt = txt.replace("Sta radis", "Šta radiš")
    if col in note_cols:
        txt = txt.replace(" | ", "<br/>")
        txt = txt.replace(" Preporuka: ", "<br/>Preporuka: ")
        txt = txt.replace(" Recommendation: ", "<br/>Recommendation: ")
        txt = txt.replace(". Finalni rez", ".<br/>Finalni rez")
        txt = txt.replace(". Final cut", ".<br/>Final cut")
    return txt


def build_service_packet(
    kitchen: Dict[str, Any],
    sections: Dict[str, pd.DataFrame] | None = None,
    lang: str = "sr",
) -> Dict[str, pd.DataFrame]:
    """Workshop-oriented packet: cuts, edging, processing, shopping, checklist."""
    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)
    sections = sections or generate_cutlist(kitchen)
    packet: Dict[str, pd.DataFrame] = {}
    _profile_key = (kitchen.get("manufacturing", {}) or {}).get("profile", "EU_SRB")
    _profile = MANUFACTURING_PROFILES.get(_profile_key, MANUFACTURING_PROFILES["EU_SRB"])
    front_gap = float(_profile.get("front_gap_mm", 2.0))

    header_df = build_project_header(kitchen, lang=_lang)
    packet["project_header"] = header_df

    board_frames = []
    for _key in ("carcass", "backs", "fronts", "drawer_boxes", "worktop", "plinth"):
        _df = sections.get(_key)
        if _df is not None and not _df.empty:
            board_frames.append(_df.copy())
    board_df = pd.concat(board_frames, ignore_index=True) if board_frames else pd.DataFrame()

    if not board_df.empty:
        for _col in ("Materijal", "Deb.", "CUT_W [mm]", "CUT_H [mm]", "Kol.", "Kant", "Napomena"):
            if _col not in board_df.columns:
                board_df[_col] = ""
        if "Zid" not in board_df.columns:
            board_df["Zid"] = ""

        service_cuts = (
            board_df.groupby(
                ["Zid", "Materijal", "Deb.", "CUT_W [mm]", "CUT_H [mm]", "Kant"],
                as_index=False,
            )
            .agg({"Kol.": "sum"})
            .sort_values(["Zid", "Materijal", "CUT_W [mm]", "CUT_H [mm]"], ascending=[True, True, False, False])
            .reset_index(drop=True)
        )
        service_cuts.insert(0, "RB", range(1, len(service_cuts) + 1))
        service_cuts["Napomena za servis"] = _t(
            "U servisu seci po CUT merama i proveri kantovanje u posebnoj tabeli",
            "In the workshop, cut strictly by CUT dimensions and verify edging in the separate table",
        )
        packet["service_cuts"] = service_cuts

        edge_df = board_df[
            board_df["Kant"].astype(str).str.strip().replace("", "-") != "-"
        ].copy()
        if not edge_df.empty:
            edge_df = edge_df[[
                c for c in [
                    "PartCode", "Zid", "Modul", "Deo", "Kol.", "Materijal", "Deb.",
                    "CUT_W [mm]", "CUT_H [mm]", "Kant", "L1", "L2", "K1", "K2", "Napomena",
                ] if c in edge_df.columns
            ]]
            packet["service_edge"] = edge_df.reset_index(drop=True)

        def _proc_row(
            part: str,
            zid: str,
            modul: str,
            deo: str,
            cut_w: Any,
            cut_h: Any,
            kol: Any,
            tip: str,
            izvodi: str,
            osnov: str,
            nap: str,
        ) -> Dict[str, Any]:
            return {
                "PartCode": part,
                "Zid": zid,
                "Modul": modul,
                "Deo": deo,
                "CUT_W [mm]": cut_w,
                "CUT_H [mm]": cut_h,
                "Kol.": kol,
                "Tip obrade": tip,
                "Izvodi": izvodi,
                "Osnov izvođenja": osnov,
                "Obrada / napomena": nap,
            }

        proc_rows: list[dict[str, Any]] = []
        proc_mask = pd.Series(False, index=board_df.index)
        for _kw in ("otvor", "utor", "ventil", "kabl", "odvod", "dovod", "instal", "prolaz", "sudoper", "napa"):
            proc_mask = proc_mask | board_df["Napomena"].astype(str).str.lower().str.contains(_kw, na=False)
        proc_df = board_df[proc_mask].copy()
        if not proc_df.empty:
            for _r in proc_df.to_dict("records"):
                _part_code = str(_r.get("PartCode", "") or "")
                _part_name = str(_r.get("Deo", "") or "").strip().lower()
                if _part_code.startswith("M0-W") or _part_name in {"radna ploča", "radna ploca", "worktop"}:
                    continue
                _nap = str(_r.get("Napomena", "") or "")
                _nap_l = _nap.lower()
                _tip = _t("Posebna obrada", "Special machining")
                _izv = _t("Servis", "Workshop")
                _osnov = _t("Po meri iz projekta", "According to project dimensions")
                if "utor" in _nap_l:
                    _tip = _t("Utor za leđa", "Back panel groove")
                elif "sudoper" in _nap_l:
                    _tip = _t("Otvor za sudoperu", "Sink cut-out")
                    _osnov = _t("Po šablonu proizvođača", "According to the manufacturer's template")
                elif "napa" in _nap_l or "ventil" in _nap_l:
                    _tip = _t("Ventilacija / otvor", "Ventilation / opening")
                    _osnov = _t("Po šablonu proizvođača", "According to the manufacturer's template")
                elif "kabl" in _nap_l:
                    _tip = _t("Prolaz za kabl", "Cable pass-through")
                    _izv = _t("Kuća / lice mesta", "On site")
                elif "odvod" in _nap_l or "dovod" in _nap_l or "instal" in _nap_l:
                    _tip = _t("Instalacioni prolaz", "Service pass-through")
                    _izv = _t("Kuća / lice mesta", "On site")
                proc_rows.append(_proc_row(
                    str(_r.get("PartCode", "")),
                    str(_r.get("Zid", "")),
                    str(_r.get("Modul", "")),
                    str(_r.get("Deo", "")),
                    _r.get("CUT_W [mm]", ""),
                    _r.get("CUT_H [mm]", ""),
                    _r.get("Kol.", ""),
                    _tip,
                    _izv,
                    _osnov,
                    _nap,
                ))

        for _m in (kitchen.get("modules", []) or []):
            _tid = str((_m or {}).get("template_id", "")).upper()
            _mid = int((_m or {}).get("id", 0))
            _lbl = str((_m or {}).get("label", _tid) or _tid)
            _zid = f"Zid {str((_m or {}).get('wall_key', 'A') or 'A').upper()}"
            if _tid == "SINK_BASE":
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Radna ploča", "Worktop"), "-", "-", 1,
                    _t("Otvor za sudoperu", "Sink cut-out"), _t("Servis", "Workshop"), _t("Po šablonu proizvođača", "According to the manufacturer's template"),
                    _t("Izrezati otvor za sudoperu u radnoj ploči prema šablonu proizvođača sudopere.", "Cut the sink opening in the worktop according to the sink manufacturer's template."),
                ))
            elif _tid in {"BASE_COOKING_UNIT", "BASE_HOB"}:
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Radna ploča", "Worktop"), "-", "-", 1,
                    _t("Otvor za ploču", "Hob cut-out"), _t("Servis", "Workshop"), _t("Po šablonu proizvođača", "According to the manufacturer's template"),
                    _t("Izrezati otvor za ploču za kuvanje u radnoj ploči prema šablonu proizvođača.", "Cut the hob opening in the worktop according to the manufacturer's template."),
                ))
            elif _tid == "WALL_HOOD":
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Leđa / prolaz", "Back panel / opening"), "-", "-", 1,
                    _t("Ventilacija / otvor", "Ventilation / opening"), _t("Servis", "Workshop"), _t("Po šablonu proizvođača", "According to the manufacturer's template"),
                    _t("Obezbediti otvor i prolaz za odvod nape prema modelu i osi instalacije.", "Provide the opening and duct path for the hood according to the model and the installation axis."),
                ))
            elif _tid == "FILLER_PANEL":
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Filer panel", "Filler panel"), "-", "-", 1,
                    _t("Vidljiva dekorativna ploča", "Visible decorative panel"), _t("Servis", "Workshop"), _t("Po meri iz projekta", "According to project dimensions"),
                    _t("Obraditi kao usku završnu popunu; proveriti vidljive ivice i montažni zazor prema zidu/frontu.", "Treat as a narrow finishing filler; verify visible edges and installation clearance toward the wall/front."),
                ))
            elif _tid == "END_PANEL":
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Završna bočna ploča", "End side panel"), "-", "-", 1,
                    _t("Vidljiva dekorativna ploča", "Visible decorative panel"), _t("Servis", "Workshop"), _t("Po meri iz projekta", "According to project dimensions"),
                    _t("Obraditi kao završnu bočnu stranu; proveriti smer goda, vidljive ivice i eventualni prepust.", "Treat as the finished end side; verify grain direction, visible edges, and any reveal."),
                ))
            elif _tid == "WALL_MICRO":
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Leđa / prolaz", "Back panel / opening"), "-", "-", 1,
                    _t("Prolaz za kabl", "Cable pass-through"), _t("Kuća / lice mesta", "On site"), _t("Po šablonu proizvođača", "According to the manufacturer's template"),
                    _t("Obezbediti prolaz za kabl i ventilacioni razmak za ugradnu mikrotalasnu.", "Provide a cable pass-through and the required ventilation gap for the built-in microwave."),
                ))
            elif _tid == "BASE_DISHWASHER":
                _dish = dishwasher_installation_metrics(kitchen, _m)
                _raised = bool(_dish.get("dishwasher_raised_mode", False))
                _platform_h = int(_dish.get("dishwasher_platform_height", 0))
                _lower_fill_h = int(_dish.get("dishwasher_lower_filler_height", 0))
                _lower_fill_cut_h = int(max(0.0, float(_lower_fill_h) - 2 * front_gap))
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Zona priključka", "Connection zone"), "-", "-", 1,
                    _t("Instalacioni prolaz", "Service pass-through"), _t("Kuća / lice mesta", "On site"), _t("Po meri iz projekta", "According to project dimensions"),
                    _t(
                        "Proveriti prolaz creva i kabla za MZS bez konflikta sa korpusom i susednim elementima.",
                        "Check the hose and cable routing for the built-in dishwasher so it does not conflict with the cabinet or adjacent units.",
                    ) + (
                        "" if not _raised else
                        _t(
                            f" Raised mode: podici mašinu na postolje {_platform_h} mm i za servis koristiti CUT visinu donje maske {_lower_fill_cut_h} mm iznad kontinuirane sokle.",
                            f" Raised mode: lift the appliance on a {_platform_h} mm platform and use the lower filler CUT height {_lower_fill_cut_h} mm above the continuous plinth.",
                        )
                    ),
                ))
            elif _tid in {"TALL_FRIDGE", "TALL_FRIDGE_FREEZER"}:
                proc_rows.append(_proc_row(
                    f"M{_mid:02d}", _zid, _lbl, _t("Ventilacija kolone", "Tall-unit ventilation"), "-", "-", 1,
                    _t("Ventilacija / otvor", "Ventilation / opening"), _t("Servis", "Workshop"), _t("Po šablonu proizvođača", "According to the manufacturer's template"),
                    _t("Obezbediti ulaz/izlaz vazduha za integrisani rashladni uređaj prema preporuci proizvođača.", "Provide the intake and exhaust ventilation for the integrated cooling appliance according to the manufacturer's recommendations."),
                ))

        _wt_df = sections.get("worktop", pd.DataFrame())
        if _wt_df is not None and not _wt_df.empty:
            for _idx, _r in enumerate(_wt_df.to_dict("records"), start=1):
                if str(_r.get("Deo", "") or "").strip().lower() not in {"radna ploča", "radna ploca", "worktop"}:
                    continue
                _zid = str(_r.get("Zid", "") or "")
                _mod = str(_r.get("Modul", "") or _zid or f"Worktop {_idx}")
                _req_raw = _r.get("Required length [mm]", 0)
                _buy_raw = _r.get("Purchase length [mm]", 0)
                _req = int(float(_req_raw)) if pd.notna(_req_raw) and str(_req_raw).strip() else 0
                _buy = int(float(_buy_raw)) if pd.notna(_buy_raw) and str(_buy_raw).strip() else 0
                _joint = str(_r.get("Joint type", "") or "").strip()
                _cutouts = str(_r.get("Cutouts", "") or "").strip()
                _field_cut = str(_r.get("Field cut", "") or "").strip().upper() == "TRUE"
                _nap = _t(
                    f"Zidna mera / finished dimension: {_req} mm. CUT osnova / purchase segment: {_buy} mm. Servis radi isključivo po CUT merama. Finalni rez na licu mesta.",
                    f"Wall requirement / finished dimension: {_req} mm. CUT basis / purchase segment: {_buy} mm. Workshop works strictly by CUT dimensions. Final cut is done on site.",
                )
                if _joint:
                    _nap += " " + _t(f"Spoj u uglu: {_joint}.", f"Corner joint: {_joint}.")
                if _cutouts:
                    _nap += " " + _t(f"Izrezi: {_cutouts}.", f"Cut-outs: {_cutouts}.")
                proc_rows.append(_proc_row(
                    f"W{_idx:02d}",
                    _zid,
                    _mod,
                    _t("Radna ploča", "Worktop"),
                    _r.get("Dužina [mm]", ""),
                    _r.get("Širina [mm]", ""),
                    _r.get("Kol.", 1),
                    _t("Priprema i finalni rez", "Preparation and final cut"),
                    _t("Servis + lice mesta", "Workshop + on site") if _field_cut else _t("Servis", "Workshop"),
                    _t("Po geometriji zida i šablonu", "According to wall geometry and template"),
                    _nap,
                ))

        if proc_rows:
            proc_df = pd.DataFrame(proc_rows).drop_duplicates().reset_index(drop=True)
            packet["service_processing"] = proc_df

    hw_df = sections.get("hardware", pd.DataFrame())
    hw_df = hw_df.copy() if hw_df is not None else pd.DataFrame()
    if not hw_df.empty:
        shop_df = hw_df[hw_df["Kategorija"].astype(str).str.lower() != "warning"].copy()
        _ready_mask = pd.Series(False, index=shop_df.index)
        for _kw in (
            "frižider", "frizider", "rerna", "mikrotalas", "mašina za sudove", "masina za sudove",
            "sudopera", "slavina", "sifon", "napa", "aspirator", "ploča za kuvanje", "ploca za kuvanje",
        ):
            _ready_mask = _ready_mask | shop_df["Naziv"].astype(str).str.lower().str.contains(_kw, na=False)
        def _shopping_group(_row: pd.Series) -> str:
            _naziv = str(_row.get("Naziv", "")).lower()
            _kat = str(_row.get("Kategorija", "")).lower()
            _mod = str(_row.get("Modul", ""))
            if any(_kw in _naziv for _kw in ("zidni nosac", "anker za zid", "anti-tip", "zidni ")):
                return _t("Montaža na zid", "Wall installation")
            if _mod == "Projekat":
                return _t("Projektni potrosni materijal", "Project consumables")
            if _kat == "okov":
                return _t("Okovi i mehanizmi po elementu", "Hardware and mechanisms by unit")
            if _kat == "potrosni":
                return _t("Sitni materijal po elementu", "Consumables by unit")
            return _t("Kupuje se posebno", "Purchased separately")

        shop_df["Grupa"] = shop_df.apply(_shopping_group, axis=1)
        shop_df.loc[_ready_mask, "Grupa"] = _t("Kupuje se gotovo", "Purchased ready-made")
        shop_df = (
            shop_df.groupby(["Grupa", "Naziv", "Tip / Šifra"], as_index=False)
            .agg({"Kol.": "sum", "Napomena": "first"})
            .sort_values(["Grupa", "Naziv"])
            .reset_index(drop=True)
        )

        extra_rows = [
            {"Grupa": _t("Alat koji treba kod kuće", "Tools needed on site"), "Naziv": _t("Aku šrafilica / odvijač", "Cordless drill / driver"), "Tip / Šifra": "-", "Kol.": 1,
             "Napomena": _t("Treba za sklapanje korpusa i montažu okova", "Needed for cabinet assembly and hardware installation")},
            {"Grupa": _t("Alat koji treba kod kuće", "Tools needed on site"), "Naziv": _t("Metar + libela", "Tape measure + spirit level"), "Tip / Šifra": "-", "Kol.": 1,
             "Napomena": _t("Treba za proveru mera, ravni i nivelacije", "Needed to verify dimensions, plumb and levelling")},
            {"Grupa": _t("Alat koji treba kod kuće", "Tools needed on site"), "Naziv": _t("Stege", "Clamps"), "Tip / Šifra": "-", "Kol.": 2,
             "Napomena": _t("Pomažu da delovi ostanu poravnati tokom sklapanja", "They help keep parts aligned during assembly")},
        ]
        if any(str((m or {}).get("zone", "")).lower() in ("wall", "wall_upper", "tall") for m in (kitchen.get("modules", []) or [])):
            extra_rows.append({
                "Grupa": _t("Za montažu na zid", "For wall installation"),
                "Naziv": _t("Tipli / ankeri za zid", "Wall plugs / anchors"),
                "Tip / Šifra": "-",
                "Kol.": 1,
                "Napomena": _t("Izabrati prema tipu zida na licu mesta", "Choose according to the actual wall type on site"),
            })
        shop_df = pd.concat([shop_df, pd.DataFrame(extra_rows)], ignore_index=True)
        packet["shopping_list"] = shop_df
        ready_df = shop_df[shop_df["Grupa"].astype(str) == _t("Kupuje se gotovo", "Purchased ready-made")].copy()
        if not ready_df.empty:
            packet["ready_made_items"] = ready_df.reset_index(drop=True)

    packet["user_guide"] = pd.DataFrame([
        {"Korak": 1, "Šta radiš": _t("Proveri projekat pre naručivanja", "Review the project before ordering"), "Napomena": _t("Potvrdi zid, mere, verziju i da li je ovo poslednja izmena", "Confirm the wall, dimensions, version and that this is the latest revision")},
        {"Korak": 2, "Šta radiš": _t("U servis nosiš samo sečenje, kantovanje i obradu", "Take only cutting, edging and machining to the workshop"), "Napomena": _t("Servis koristi samo tabele iz servis paketa", "The workshop should work only from the workshop packet tables")},
        {"Korak": 3, "Šta radiš": _t("Posebno kupuješ gotove uređaje, okove i alat", "Purchase ready-made appliances, hardware and tools separately"), "Napomena": _t("To nije deo sečenja i mora da se nabavi posebno", "These items are not part of cutting and must be sourced separately")},
        {"Korak": 4, "Šta radiš": _t("Kod kuće razvrstaš delove po modulu", "Sort parts by unit on site"), "Napomena": _t("Prvo odvoji korpuse, zatim frontove, pa okove", "Separate carcass parts first, then fronts, then hardware")},
        {"Korak": 5, "Šta radiš": _t("Prvo sklapaš korpuse, pa vrata, fioke i uređaje", "Assemble the carcasses first, then doors, drawers and appliances"), "Napomena": _t("Visoke i zidne elemente obavezno pričvrsti za zid", "Always secure tall and wall units to the wall")},
    ])

    packet["workshop_checklist"] = pd.DataFrame([
        {"RB": 1, "Stavka": _t("Proveri da li su tačni materijal i debljine za korpus, front i leđa", "Check that the material and thickness values are correct for carcass, fronts and backs"), "Status": ""},
        {"RB": 2, "Stavka": _t("Proveri da li su upisane sve CUT mere i broj komada", "Check that all CUT dimensions and quantities are listed"), "Status": ""},
        {"RB": 3, "Stavka": _t("Proveri kantovanje po svakoj ivici pre predaje u servis", "Check edge banding on every edge before sending to the workshop"), "Status": ""},
        {"RB": 4, "Stavka": _t("Proveri da li su ubačeni frontovi, leđa, sokla i posebni paneli", "Check that fronts, backs, plinths and special panels are included"), "Status": ""},
        {"RB": 5, "Stavka": _t("Proveri da li je sve što se kupuje gotovo izdvojeno van sečenja", "Check that all ready-made purchased items are separated from cut parts"), "Status": ""},
        {"RB": 6, "Stavka": _t("Proveri sve otvore i posebne obrade za instalacije i ventilaciju", "Check all openings and special machining for services and ventilation"), "Status": ""},
        {"RB": 7, "Stavka": _t("Proveri da lista kupovine sadrži okove, sitni materijal i alat", "Check that the shopping list includes hardware, consumables and tools"), "Status": ""},
    ])
    packet["home_checklist"] = pd.DataFrame([
        {"RB": 1, "Stavka": _t("Prebroj sve isečene ploče i uporedi ih sa listom", "Count all cut panels and compare them against the list"), "Status": ""},
        {"RB": 2, "Stavka": _t("Razvrstaj delove po elementu pre početka sklapanja", "Sort parts by unit before starting assembly"), "Status": ""},
        {"RB": 3, "Stavka": _t("Proveri da li su kupljeni svi okovi, uređaji i alat", "Check that all hardware, appliances and tools have been purchased"), "Status": ""},
        {"RB": 4, "Stavka": _t("Prvo sastavi korpuse, zatim vrata i fioke, pa tek onda uređaje", "Assemble the carcasses first, then doors and drawers, and only then the appliances"), "Status": ""},
        {"RB": 5, "Stavka": _t("Visoke i zidne elemente obavezno pričvrsti za zid", "Always secure tall and wall units to the wall"), "Status": ""},
    ])

    packet["service_instructions"] = pd.DataFrame([
        {"RB": 1, "Stavka": _t("Sečenje", "Cutting"), "Instrukcija": _t("Seci isključivo po CUT merama iz servis paketa.", "Cut strictly by the CUT dimensions from the workshop packet.")},
        {"RB": 2, "Stavka": _t("Kantovanje", "Edging"), "Instrukcija": _t("Kantuj samo ivice označene u tabeli kantovanja i proveri tip ABS-a.", "Apply edging only to the edges marked in the edging table and verify the ABS type.")},
        {"RB": 3, "Stavka": _t("Obrade", "Machining"), "Instrukcija": _t("Posebne otvore, utore i ventilaciju radi samo tamo gde su eksplicitno navedeni.", "Make special openings, grooves and ventilation cuts only where they are explicitly specified.")},
        {"RB": 4, "Stavka": _t("Radna ploča", "Worktop"), "Instrukcija": _t("Otvor za sudoperu ili ploču radi po šablonu proizvođača, ne po slobodnoj proceni.", "Cut the sink or hob opening according to the manufacturer's template, not by estimation.")},
        {"RB": 5, "Stavka": _t("Ko izvodi", "Who performs it"), "Instrukcija": _t("Kolona 'Izvodi' jasno razlikuje da li posao radi servis ili se obrada potvrđuje kod kuće / na licu mesta.", "The 'Operations' column clearly shows whether the workshop performs the job or whether it must be confirmed on site.")},
        {"RB": 6, "Stavka": _t("Po kom pravilu", "Execution basis"), "Instrukcija": _t("Kolona 'Osnov izvođenja' pokazuje da li se radi po meri iz projekta ili po šablonu proizvođača uređaja.", "The 'Execution basis' column shows whether the work follows project dimensions or the appliance manufacturer's template.")},
        {"RB": 7, "Stavka": _t("Instalacije", "Services"), "Instrukcija": _t("Ako pozicija instalacija odstupa od projekta, obrade potvrditi pre sečenja ili na licu mesta.", "If the service positions differ from the project, confirm all machining before cutting or on site.")},
        {"RB": 8, "Stavka": _t("Kontrola", "Final check"), "Instrukcija": _t("Pre isporuke proveriti broj komada, oznake delova i da li su svi paneli/frontovi uključeni.", "Before delivery, verify the piece count, part labels and that all panels/fronts are included.")},
    ])

    return packet


def build_cutlist_pdf_bytes(
    kitchen: Dict[str, Any],
    cutlist_sections: Dict[str, pd.DataFrame],
    project_title: str = "Krojna lista PRO - M1 (jedan zid)",
    lang: str = "sr",
) -> bytes:
    _register_fonts()
    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)

    def _pdf_clean_text(value: Any) -> str:
        txt = str(value or "")
        replacements = {
            "â€“": "-",
            "â€”": "-",
            "â€œ": "\"",
            "â€": "\"",
            "â€˜": "'",
            "â€™": "'",
            "Ä": "č",
            "Ä‡": "ć",
            "Å¾": "ž",
            "Å¡": "š",
            "Ä‘": "đ",
            "ÄŒ": "Č",
            "Ä†": "Ć",
            "Å½": "Ž",
            "Å ": "Š",
            "Ä": "Đ",
            "LeÄ‘a": "Leđa",
            "LeÄ‘na ploÄa": "Leđna ploča",
            "BoÄna ploÄa": "Bočna ploča",
            "ploÄa": "ploča",
            "seÄenje": "sečenje",
            "seÄi": "seci",
            "ureÄ‘aji": "uređaji",
            "sliÄno": "slično",
            "potroÅ¡ni": "potrošni",
            "MontaÅ¾na": "Montažna",
            "montaÅ¾u": "montažu",
            "vodiÄ": "vodič",
            "kuÄ‡nog": "kućnog",
            "potvrÄ‘uje": "potvrđuje",
            "izvoÄ‘enja": "izvođenja",
            "Å¡ablonu": "šablonu",
            "proizvoÄ‘aÄa": "proizvođača",
            "ureÄ‘aja": "uređaja",
            "plocasti": "pločasti",
            "ploca": "ploča",
            "radna ploca": "radna ploča",
            "Radna ploca": "Radna ploča",
            "uredjaji": "uređaji",
            "uredjaja": "uređaja",
            "kucnog": "kućnog",
            "kuce": "kuće",
            "vodic": "vodič",
            "kriticna": "kritična",
            "Kriticna": "Kritična",
            "secenje": "sečenje",
            "Secenje": "Sečenje",
            "srafilica": "šrafilica",
            "sifra": "šifra",
            "sablonu": "šablonu",
            "proizvodjaca": "proizvođača",
            "montaza": "montaža",
            "Montaza": "Montaža",
            "ledja": "leđa",
            "Ledja": "Leđa",
            "ubaceni": "ubačeni",
            "precvrsti": "pričvrsti",
            "pricvrsti": "pričvrsti",
        }
        for src, dst in replacements.items():
            txt = txt.replace(src, dst)
        return txt

    service_packet = build_service_packet(kitchen, cutlist_sections, lang=lang)
    display_sections = {k: _translate_export_df(v, _lang) for k, v in (cutlist_sections or {}).items()}
    display_service_packet = {k: _translate_export_df(v, _lang) for k, v in (service_packet or {}).items()}

    # Pick a font that supports Cyrillic
    try:
        pdfmetrics.getFont("DejaVuSans")
        FONT_REGULAR = "DejaVuSans"
        FONT_BOLD = "DejaVuSans-Bold"
    except Exception as ex:
        _LOG.debug("Falling back to Helvetica fonts for PDF build: %s", ex)
        FONT_REGULAR = "Helvetica"
        FONT_BOLD = "Helvetica-Bold"

    profile_key = (kitchen.get("manufacturing", {}) or {}).get("profile", "EU_SRB")
    profile = MANUFACTURING_PROFILES.get(profile_key, MANUFACTURING_PROFILES["EU_SRB"])
    mats = kitchen.get("materials", {}) or {}

    materials_line = " | ".join([
        _format_material_role(mats.get('carcass_material', ''), mats.get('carcass_thk', 18), 'carcass', _lang),
        _format_material_role(mats.get('front_material', ''), mats.get('front_thk', 19), 'front', _lang),
        _format_material_role(mats.get('back_material', ''), mats.get('back_thk', 3), 'back', _lang),
    ])

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=project_title,
    )

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle(
        "TitleCyr",
        parent=styles["Title"],
        fontName=FONT_BOLD,
        fontSize=16,
        leading=18,
        spaceAfter=8,
    )
    s_norm = ParagraphStyle(
        "NormCyr",
        parent=styles["Normal"],
        fontName=FONT_REGULAR,
        fontSize=10,
        leading=12,
    )
    s_h2 = ParagraphStyle(
        "H2Cyr",
        parent=styles["Heading2"],
        fontName=FONT_BOLD,
        fontSize=12,
        leading=14,
        spaceBefore=12,
        spaceAfter=6,
    )

    # ---- Helper: napravi ReportLab tabelu iz DataFrame-a ----
    # Landscape A4: 297mm - 2*10mm margina = 277mm upotrebljive sirine
    _PAGE_W_MM = 277.0
    # Podrazumevane sirine kolona po imenu (u mm)
    _COL_W: Dict[str, float] = {
        "RB": 10, "PartCode": 22, "Zid": 14, "Modul": 44, "Deo": 44,
        "Pozicija": 16, "SklopKorak": 14, "Kom": 11, "Kol.": 11,
        "Dužina [mm]": 18, "Širina [mm]": 18, "Deb.": 11, "Deb. [mm]": 11,
        "Visina [mm]": 16, "CUT_W [mm]": 18, "CUT_H [mm]": 18,
        "Materijal": 36, "Smer goda": 16, "Orijentacija": 20,
        "L1": 9, "L2": 9, "K1": 9, "K2": 9, "Kant": 16,
        "Napomena": 32, "Napomena za servis": 32,
        "Tip obrade": 18, "Izvodi": 14, "Osnov izvođenja": 20,
        "Obrada / napomena": 28, "Stavka": 28, "Instrukcija": 50,
        "Polje": 40, "Vrednost": 60,
        "ID": 10, "TYPE": 12, "Element": 30, "Naziv": 36,
        "Tip / Šifra": 28, "Kategorija": 20,
    }

    def _df_to_table(df: pd.DataFrame, cols: List[str]) -> Table:
        _table_df = df.copy()
        for _fin_col, _cut_col in (("Dužina [mm]", "CUT_W [mm]"), ("Širina [mm]", "CUT_H [mm]")):
            if _fin_col in _table_df.columns and _cut_col in _table_df.columns:
                _table_df[_fin_col] = _table_df[_fin_col].where(
                    _table_df[_fin_col].notna() & (_table_df[_fin_col].astype(str).str.strip().str.lower() != "nan") & (_table_df[_fin_col].astype(str).str.strip() != ""),
                    _table_df[_cut_col],
                )
        _available_cols = [c for c in cols if c in _table_df.columns]
        _is_summary_table = _available_cols == [
            "RB", "Deo", "Kom", "Dužina [mm]", "Širina [mm]", "Deb.",
            "Materijal", "Orijentacija", "L1", "L2", "K1", "K2",
        ]
        # Izracunaj colWidths — skaliraj na ukupnu sirinu stranice
        if _is_summary_table:
            _summary_w = {
                "RB": 7, "Deo": 28, "Kom": 9, "Dužina [mm]": 15, "Širina [mm]": 15,
                "Deb.": 7, "Materijal": 21, "Orijentacija": 9, "L1": 6, "L2": 6, "K1": 6, "K2": 6,
            }
            _raw_w = [_summary_w.get(c, _COL_W.get(c, 20)) for c in _available_cols]
        else:
            _raw_w = [_COL_W.get(c, 20) for c in _available_cols]
        _total_raw = sum(_raw_w)
        _scale = _PAGE_W_MM / _total_raw if _total_raw > 0 else 1.0
        _cw = [w * _scale * mm for w in _raw_w]
        # Stilovi za Paragraph unutar celija (tekst se prelama umesto da izlazi van)
        _font_size = 6.0 if _is_summary_table else 7
        _leading = 7.0 if _is_summary_table else 8.5
        _pad_lr = 0.8 if _is_summary_table else 3
        _th = ParagraphStyle("th_tbl", fontName=FONT_BOLD, fontSize=_font_size, leading=_leading,
                             wordWrap="CJK")
        _td = ParagraphStyle("td_tbl", fontName=FONT_REGULAR, fontSize=_font_size, leading=_leading,
                             wordWrap="CJK")
        _header_map_pt = {
            "RB": "RB",
            "PartCode": "PartCode",
            "Korak": "Passo",
            "Šta radiš": "O que fazer",
            "Sta radis": "O que fazer",
            "Pojam": "Termo",
            "Objašnjenje": "Explicação",
            "Zid": "Parede",
            "Modul": "Módulo",
            "Deo": "Peça",
            "Pozicija": "Posição",
            "SklopKorak": "Passo",
            "Kom": "Qtd.",
            "Kol.": "Qtd.",
            "Dužina [mm]": "Comprimento [mm]",
            "Širina [mm]": "Largura [mm]",
            "Visina [mm]": "Altura [mm]",
            "Deb.": "Esp.",
            "Deb. [mm]": "Esp. [mm]",
            "Materijal": "Material",
            "Smer goda": "Veio",
            "Orijentacija": "Orientação",
            "Kant": "Borda",
            "Napomena": "Nota",
            "Napomena za servis": "Nota para a marcenaria",
            "Tip obrade": "Tipo de usinagem",
            "Izvodi": "Executa",
            "Osnov izvođenja": "Base de execução",
            "Obrada / napomena": "Usinagem / nota",
            "Stavka": "Item",
            "Instrukcija": "Instrução",
            "Polje": "Campo",
            "Vrednost": "Valor",
            "Naziv": "Nome",
            "Tip / Šifra": "Tipo / código",
            "Kategorija": "Categoria",
        }
        _header_map_en = {
            "RB": "No.",
            "PartCode": "Code",
            "Korak": "Step",
            "Šta radiš": "What you do",
            "Sta radis": "What you do",
            "Pojam": "Term",
            "Objašnjenje": "Explanation",
            "Zid": "Wall",
            "Modul": "Module",
            "Deo": "Part",
            "Pozicija": "Position",
            "SklopKorak": "Step",
            "Kom": "Qty",
            "Kol.": "Qty",
            "Dužina [mm]": "Length [mm]",
            "Širina [mm]": "Width [mm]",
            "Visina [mm]": "Height [mm]",
            "Deb.": "Thk.",
            "Deb. [mm]": "Thk. [mm]",
            "Materijal": "Material",
            "Smer goda": "Grain",
            "Orijentacija": "Orientation",
            "Kant": "Edge",
            "Napomena": "Note",
            "Napomena za servis": "Workshop note",
            "Tip obrade": "Processing type",
            "Izvodi": "Operations",
            "Osnov izvođenja": "Execution basis",
            "Obrada / napomena": "Processing / note",
            "Stavka": "Item",
            "Instrukcija": "Instruction",
            "Polje": "Field",
            "Vrednost": "Value",
            "Naziv": "Name",
            "Tip / Šifra": "Type / Code",
            "Kategorija": "Category",
        }
        _header_map_es = {
            "RB": "RB",
            "PartCode": "PartCode",
            "Korak": "Paso",
            "Šta radiš": "Qué hacer",
            "Sta radis": "Qué hacer",
            "Pojam": "Término",
            "Objašnjenje": "Explicación",
            "Zid": "Pared",
            "Modul": "Módulo",
            "Deo": "Pieza",
            "Pozicija": "Posición",
            "SklopKorak": "Paso",
            "Kom": "Cant.",
            "Kol.": "Cant.",
            "Dužina [mm]": "Longitud [mm]",
            "Širina [mm]": "Ancho [mm]",
            "Visina [mm]": "Altura [mm]",
            "Deb.": "Esp.",
            "Deb. [mm]": "Esp. [mm]",
            "Materijal": "Material",
            "Smer goda": "Veta",
            "Orijentacija": "Orientación",
            "Kant": "Canto",
            "Napomena": "Nota",
            "Napomena za servis": "Nota para el taller",
            "Tip obrade": "Tipo de mecanizado",
            "Izvodi": "Ejecuta",
            "Osnov izvođenja": "Base de ejecución",
            "Obrada / napomena": "Mecanizado / nota",
            "Stavka": "Ítem",
            "Instrukcija": "Instrucción",
            "Polje": "Campo",
            "Vrednost": "Valor",
            "Naziv": "Nombre",
            "Tip / Šifra": "Tipo / código",
            "Kategorija": "Categoría",
        }
        _header_map_ru = {
            "RB": "№",
            "PartCode": "Код",
            "Korak": "Шаг",
            "Šta radiš": "Что делать",
            "Sta radis": "Что делать",
            "Pojam": "Термин",
            "Objašnjenje": "Пояснение",
            "Zid": "Стена",
            "Modul": "Модуль",
            "Deo": "Деталь",
            "Pozicija": "Позиция",
            "SklopKorak": "Шаг",
            "Kom": "Кол.",
            "Kol.": "Кол.",
            "Dužina [mm]": "Длина [мм]",
            "Širina [mm]": "Ширина [мм]",
            "Visina [mm]": "Высота [мм]",
            "Deb.": "Толщ.",
            "Deb. [mm]": "Толщ. [мм]",
            "Materijal": "Материал",
            "Smer goda": "Волокно",
            "Orijentacija": "Ориентация",
            "Kant": "Кромка",
            "Napomena": "Примечание",
            "Napomena za servis": "Примечание для мастерской",
            "Tip obrade": "Тип обработки",
            "Izvodi": "Выполняет",
            "Osnov izvođenja": "Основание",
            "Obrada / napomena": "Обработка / примечание",
            "Stavka": "Позиция",
            "Instrukcija": "Инструкция",
            "Polje": "Поле",
            "Vrednost": "Значение",
            "Naziv": "Название",
            "Tip / Šifra": "Тип / код",
            "Kategorija": "Категория",
        }
        _header_map_zh = {
            "RB": "序号", "PartCode": "代码", "Korak": "步骤", "Šta radiš": "操作内容", "Sta radis": "操作内容",
            "Pojam": "术语", "Objašnjenje": "说明", "Zid": "墙", "Modul": "模块", "Deo": "部件",
            "Pozicija": "位置", "SklopKorak": "步骤", "Kom": "数量", "Kol.": "数量", "Dužina [mm]": "长度 [mm]",
            "Širina [mm]": "宽度 [mm]", "Visina [mm]": "高度 [mm]", "Deb.": "厚度", "Deb. [mm]": "厚度 [mm]",
            "Materijal": "材料", "Smer goda": "纹理", "Orijentacija": "方向", "Kant": "封边", "Napomena": "备注",
            "Napomena za servis": "车间备注", "Tip obrade": "加工类型", "Izvodi": "执行", "Osnov izvođenja": "执行依据",
            "Obrada / napomena": "加工 / 备注", "Stavka": "项目", "Instrukcija": "说明", "Polje": "字段", "Vrednost": "值",
            "Naziv": "名称", "Tip / Šifra": "类型 / 编码", "Kategorija": "类别",
        }
        _header_map_hi = {
            "RB": "क्रम", "PartCode": "कोड", "Korak": "चरण", "Šta radiš": "क्या करना है", "Sta radis": "क्या करना है",
            "Pojam": "टर्म", "Objašnjenje": "समझ", "Zid": "दीवार", "Modul": "मॉड्यूल", "Deo": "भाग",
            "Pozicija": "स्थिति", "SklopKorak": "चरण", "Kom": "मात्रा", "Kol.": "मात्रा", "Dužina [mm]": "लंबाई [mm]",
            "Širina [mm]": "चौड़ाई [mm]", "Visina [mm]": "ऊंचाई [mm]", "Deb.": "मोटाई", "Deb. [mm]": "मोटाई [mm]",
            "Materijal": "सामग्री", "Smer goda": "ग्रेन", "Orijentacija": "ओरिएंटेशन", "Kant": "एज", "Napomena": "नोट",
            "Napomena za servis": "वर्कशॉप नोट", "Tip obrade": "प्रोसेसिंग प्रकार", "Izvodi": "ऑपरेशन", "Osnov izvođenja": "आधार",
            "Obrada / napomena": "प्रोसेसिंग / नोट", "Stavka": "आइटम", "Instrukcija": "निर्देश", "Polje": "फ़ील्ड", "Vrednost": "मान",
            "Naziv": "नाम", "Tip / Šifra": "प्रकार / कोड", "Kategorija": "श्रेणी",
        }
        header = [
            Paragraph(_pdf_clean_text(
                _header_map_pt.get(c, c) if _lang == "pt-br"
                else (_header_map_en.get(c, c) if _lang == "en"
                      else (_header_map_es.get(c, c) if _lang == "es"
                      else (_header_map_ru.get(c, c) if _lang == "ru"
                            else (_header_map_zh.get(c, c) if _lang == "zh-cn"
                                  else (_header_map_hi.get(c, c) if _lang == "hi" else c)))))
            ), _th)
            for c in _available_cols
        ]
        rows = []
        for row_vals in _table_df[_available_cols].fillna("").astype(object).values.tolist():
            rows.append([
                Paragraph(_pdf_clean_text(_format_pdf_table_cell(v, c)), _td)
                for c, v in zip(_available_cols, row_vals)
            ])
        data = [header] + rows
        tbl = Table(data, repeatRows=1, colWidths=_cw)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), _pad_lr),
            ("RIGHTPADDING", (0, 0), (-1, -1), _pad_lr),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            # Alternirajuce boje redova
            *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8F8F8"))
              for i in range(2, len(data), 2)],
        ]))
        return tbl

    def _rl_image_from_uri(uri: str, width_mm: float) -> RLImage | None:
        txt = str(uri or "").strip()
        if not txt.startswith("data:image/"):
            return None
        try:
            header, payload = txt.split(",", 1)
            raw = base64.b64decode(payload)
            return RLImage(BytesIO(raw), width=width_mm * mm)
        except Exception as ex:
            _LOG.debug("Failed to decode preview image URI: %s", ex)
            return None

    story: List[Any] = []
    if _lang in {"es", "pt-br", "ru", "zh-cn", "hi"}:
        project_title = _translate_export_text(project_title, _lang, "Polje")

    # ---- Zaglavlje dokumenta ----
    story.append(Paragraph(_pdf_clean_text(project_title), s_title))
    story.append(Paragraph(
        _pdf_clean_text(
            f"{_t('Profil', 'Profile')}: {profile.get('label', profile_key)} | {_t('Standard', 'Standard')}: {kitchen.get('standard', 'SRB')} | "
            f"{_t('Generisano', 'Generated')}: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ),
        s_norm
    ))
    story.append(Paragraph(_pdf_clean_text(f"{_t('Materijali', 'Materials')}: {materials_line}"), s_norm))
    story.append(Spacer(1, 4 * mm))

    # ---- Kratko uputstvo za laika ----
    story.append(Paragraph(_pdf_clean_text(_t("Kako koristiš ovaj dokument", "How to use this document")), s_h2))
    _howto_df = pd.DataFrame([
        {"Korak": 1, "Šta radiš": _t("Prvo proveri osnovne mere i materijale", "First check the main dimensions and materials"), "Napomena": _t("Ako ovde vidiš grešku, nemoj naručivati sečenje.", "If you see an error here, do not order cutting yet.")},
        {"Korak": 2, "Šta radiš": _t("U servis nosi samo deo 'Za servis'", "Take only the 'For workshop' section to the workshop"), "Napomena": _t("Servis radi po CUT merama i napomenama iz tog dela.", "The workshop works by CUT dimensions and notes from that section.")},
        {"Korak": 3, "Šta radiš": _t("Posebno kupi gotove uređaje, okove i alat", "Purchase ready-made appliances, hardware and tools separately"), "Napomena": _t("To nije deo sečenja i ne izrađuje se iz pločastog materijala.", "These items are not cut from board material.")},
        {"Korak": 4, "Šta radiš": _t("Za sklapanje prati deo 'Za montažu'", "Use the 'For assembly' section during assembly"), "Napomena": _t("Radi redom: korpusi, frontovi, fioke, uređaji, završna provera.", "Follow the order: carcasses, fronts, drawers, appliances, final check.")},
    ])
    story.append(_df_to_table(_howto_df, ["Korak", "Šta radiš", "Napomena"]))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph(_pdf_clean_text(_t("Legenda za početnika", "Beginner legend")), s_h2))
    _legend_df = pd.DataFrame([
        {"Pojam": _t("Deo", "Part"), "Objašnjenje": _t("Naziv ploče ili dela elementa.", "The name of the panel or unit part.")},
        {"Pojam": _t("Kom", "Qty"), "Objašnjenje": _t("Koliko komada tog dela treba.", "How many pieces of that part are needed.")},
        {"Pojam": _t("Deb.", "Thk."), "Objašnjenje": _t("Debljina ploče u milimetrima.", "Panel thickness in millimeters.")},
        {"Pojam": _t("Dužina / Širina", "Length / Width"), "Objašnjenje": _t("Gotove mere dela.", "Finished part dimensions.")},
        {"Pojam": _t("Materijal", "Material"), "Objašnjenje": _t("Od čega se deo izrađuje.", "What the part is made of.")},
        {"Pojam": _t("Orijentacija", "Orientation"), "Objašnjenje": _t("Smer ploče ili dezena.", "Board or grain direction.")},
        {"Pojam": _t("Kant", "Edge"), "Objašnjenje": _t("Koje ivice se kantuju.", "Which edges get edge banding.")},
        {"Pojam": "L1 / L2 / K1 / K2", "Objašnjenje": _t("L1/L2 su duže ivice, K1/K2 su kraće ivice.", "L1/L2 are long edges, K1/K2 are short edges.")},
        {"Pojam": "1 / 0", "Objašnjenje": _t("1 = kantuje se, 0 = ne kantuje se.", "1 = edged, 0 = not edged.")},
        {"Pojam": "CUT", "Objašnjenje": _t("Mera za servis pre završne obrade, gde je primenljivo.", "Workshop size before final processing, where applicable.")},
    ])
    story.append(_df_to_table(_legend_df, ["Pojam", "Objašnjenje"]))
    story.append(Spacer(1, 5 * mm))

    # ---- Radni nalog / project header ----
    story.append(Paragraph(_pdf_clean_text(_t("Radni nalog", "Work order")), s_h2))
    _hdr_df = display_service_packet.get("project_header", pd.DataFrame())
    if _hdr_df is not None and not _hdr_df.empty:
        story.append(_df_to_table(_hdr_df, ["Polje", "Vrednost"]))
    story.append(Spacer(1, 4 * mm))

    # ---- Katalog prikaz ----
    try:
        from visualization import _render as _viz_render, _wall_len_h  # type: ignore
        import struct as _struct

        # Izracunaj stvarne proporcije zida (iste granice kao _setup_view katalog mod)
        _wl_v, _wh_v = _wall_len_h(kitchen)
        _x_total = 170 + _wl_v + 620   # left_pad + wall_len + right_pad
        _y_total = 420 + _wh_v + 150   # bottom_pad + wall_h + top_pad
        _aspect  = _x_total / max(_y_total, 1)   # data width / data height
        _fig_h_in = 5.0
        _fig_w_in = max(_fig_h_in * _aspect, 4.0)

        _fig = plt.figure(figsize=(_fig_w_in, _fig_h_in))
        _ax = _fig.add_subplot(111)
        _viz_render(
            ax=_ax,
            kitchen=kitchen,
            view_mode="katalog",
            show_grid=False,
            grid_mm=50,
            show_bounds=False,
            kickboard=True,
            ceiling_filler=False,
        )
        _img_buf = BytesIO()
        _fig.savefig(_img_buf, format="png", dpi=150, bbox_inches="tight",
                     facecolor="white", edgecolor="none")
        plt.close(_fig)
        _img_buf.seek(0)
        # Citaj stvarne pixel dimenzije iz PNG headera — bbox_inches='tight' menja
        # stvarnu velicinu pa se virtuelni ratio ne sme koristiti.
        _raw_png = _img_buf.read()
        _pw_px = _struct.unpack('>I', _raw_png[16:20])[0]
        _ph_px = _struct.unpack('>I', _raw_png[20:24])[0]
        _img_w_mm = 270.0
        _img_h_mm = _img_w_mm * _ph_px / max(_pw_px, 1)
        _img_buf.seek(0)
        _img = RLImage(_img_buf, width=_img_w_mm * mm, height=_img_h_mm * mm)
        story.append(_img)
    except Exception as _viz_err:
        story.append(Paragraph(_pdf_clean_text(f"[{_t('Katalog slika nije dostupna', 'Catalog image is not available')}: {_viz_err}]"), s_norm))

    story.append(Spacer(1, 5 * mm))

    # ---- Sekcija: Koordinate projekta ----
    wall = kitchen.get("wall", {}) or {}
    foot_mm = kitchen.get("foot_height_mm", 100)
    base_h = kitchen.get("base_korpus_h_mm", 720)
    wt = kitchen.get("worktop", {}) or {}
    wt_mm = int(round(float(wt.get("thickness", 0.0)) * 10.0))
    radna_visina = foot_mm + base_h + wt_mm

    story.append(Paragraph(_t("Koordinate projekta", "Project coordinates"), s_h2))
    _coord_data = [
        [_t("Zid (duž. x vis.)", "Wall (L x H)"), f"{wall.get('length_mm', 0):.0f} x {wall.get('height_mm', 0):.0f} mm"],
        [_t("Stopice", "Legs"), f"{foot_mm} mm"],
        [_t("Korpus donjih el.", "Base carcass height"), f"{base_h} mm"],
        [_t("Debljina radne ploče", "Worktop thickness"), f"{wt_mm} mm"],
        [_t("Radna visina (ukupno)", "Working height (total)"), f"{radna_visina} mm"],
    ]
    _ct = Table(_coord_data)
    _ct.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONT", (0, 0), (0, -1), FONT_BOLD),
        ("FONT", (1, 0), (1, -1), FONT_REGULAR),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(_ct)
    story.append(Spacer(1, 5 * mm))

    # ---- Sekcijski marker: za korisnika ----
    story.append(Paragraph(_t("ZA KORISNIKA", "FOR CUSTOMER"), s_h2))
    story.append(Paragraph(
        _t(
            "Ovde proveravaš osnovne mere, važne panele i da li je projekat spreman za dalje korake.",
            "Use this section to review the main dimensions, important panels and overall project readiness.",
        ),
        s_norm,
    ))
    story.append(Spacer(1, 2 * mm))

    # ---- Sumarna krojna lista (identično PDF strana 4) ----
    _summary_pdf = generate_cutlist_summary(cutlist_sections, lang=_lang)
    df_sum_all = _summary_pdf.get("summary_all", pd.DataFrame())

    story.append(Paragraph(_t("Sumarna krojna lista — ploče", "Summary cut list - panels"), s_h2))
    story.append(Paragraph(_t("Dimenzije su gotove mere, posle kantovanja.", "Dimensions are finished sizes, after edging."), s_norm))
    if df_sum_all is not None and not df_sum_all.empty:
        df_sum_all = df_sum_all.copy()
        if "Materijal" in df_sum_all.columns and "Deo" in df_sum_all.columns and "Deb." in df_sum_all.columns:
            df_sum_all["Materijal"] = df_sum_all.apply(
                lambda _r: _format_material_role(
                    _r.get("Materijal", ""),
                    _r.get("Deb.", ""),
                    _material_role_from_part_name(_r.get("Deo", "")),
                    _lang,
                ),
                axis=1,
            )
        _cols_sum = ["RB", "Deo", "Kom", "Dužina [mm]", "Širina [mm]", "Deb.",
                     "Materijal", "Orijentacija", "L1", "L2", "K1", "K2"]
        story.append(_df_to_table(df_sum_all, _cols_sum))
        story.append(Spacer(1, 8 * mm))

    # ---- Sokla ----
    df_pl = display_sections.get("plinth", pd.DataFrame())
    if df_pl is not None and not df_pl.empty:
        story.append(Paragraph(_t("Sokla (lajsna)", "Plinth / toe kick"), s_h2))
        _cols_pl = ["ID", "Zid", "Deo", "Materijal", "Deb. [mm]", "Dužina [mm]", "Visina [mm]", "Kol.", "Napomena"]
        story.append(_df_to_table(df_pl, _cols_pl))
        story.append(Spacer(1, 8 * mm))

    # ---- Detaljna krojna lista po modulima (identično PDF strane 5-7) ----
    df_det = _summary_pdf.get("summary_detaljna", pd.DataFrame())
    if df_det is not None and not df_det.empty:
        story.append(Paragraph(_t("Detaljna krojna lista — po modulima", "Detailed cut list - by units"), s_h2))
        story.append(Paragraph(_t("Svaki deo je prikazan po modulu. Dimenzije su gotove mere, posle kantovanja.", "Each part is shown by unit. Dimensions are finished sizes, after edging."), s_norm))
        df_det = df_det.copy()
        if "Materijal" in df_det.columns and "Deo" in df_det.columns and "Deb." in df_det.columns:
            df_det["Materijal"] = df_det.apply(
                lambda _r: _format_material_role(
                    _r.get("Materijal", ""),
                    _r.get("Deb.", ""),
                    _material_role_from_part_name(_r.get("Deo", "")),
                    _lang,
                ),
                axis=1,
            )
        _cols_det = ["PartCode", "Zid", "Modul", "Deo", "Pozicija", "SklopKorak", "Kom", "Dužina [mm]", "Širina [mm]", "Deb.",
                     "Materijal", "Orijentacija", "L1", "L2", "K1", "K2"]
        story.append(_df_to_table(df_det, _cols_det))
        story.append(Spacer(1, 8 * mm))

    # ---- Servis paket ----
    story.append(Paragraph(_t("ZA SERVIS", "FOR WORKSHOP"), s_h2))
    story.append(Paragraph(
        _t(
            "Ovaj deo nosiš u servis. Servis radi po CUT merama, kantovanju i napomenama iz ovih tabela.",
            "Take this section to the workshop. The workshop works by CUT dimensions, edging notes and machining notes from these tables.",
        ),
        s_norm,
    ))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(_t("Servis paket za pločasti materijal", "Workshop packet for board material"), s_h2))
    story.append(Paragraph(_t("Ovo se nosi u servis: sečenje po CUT merama, kantovanje po tabeli i obrade po napomenama.", "This goes to the workshop: cutting by CUT dimensions, edging by table, and machining by notes."), s_norm))
    _svc_legend_df = pd.DataFrame([
        {"Pojam": "CUT", "Objašnjenje": _t("Mera po kojoj servis seče ploču pre završne obrade.", "Size used by the workshop before final processing.")},
        {"Pojam": _t("Kant", "Edge"), "Objašnjenje": _t("Ivica koja dobija kant traku.", "Edge that receives edge banding.")},
        {"Pojam": _t("Kol.", "Qty"), "Objašnjenje": _t("Koliko komada servis treba da pripremi.", "How many pieces the workshop should prepare.")},
        {"Pojam": _t("Napomena", "Note"), "Objašnjenje": _t("Dodatno objašnjenje za sečenje, kantovanje ili obradu.", "Extra instruction for cutting, edging or machining.")},
    ])
    story.append(_df_to_table(_svc_legend_df, ["Pojam", "Objašnjenje"]))
    story.append(Spacer(1, 3 * mm))
    _svc_cuts = display_service_packet.get("service_cuts", pd.DataFrame())
    if _svc_cuts is not None and not _svc_cuts.empty:
        story.append(Paragraph(_t("Tabela za sečenje", "Cutting table"), s_norm))
        story.append(_df_to_table(_svc_cuts, ["RB", "Zid", "Materijal", "Deb.", "CUT_W [mm]", "CUT_H [mm]", "Kant", "Kol.", "Napomena za servis"]))
        story.append(Spacer(1, 3 * mm))
    _svc_edge = display_service_packet.get("service_edge", pd.DataFrame())
    if _svc_edge is not None and not _svc_edge.empty:
        story.append(Paragraph(_t("Tabela za kantovanje", "Edging table"), s_norm))
        story.append(_df_to_table(_svc_edge, ["PartCode", "Zid", "Modul", "Deo", "Kol.", "CUT_W [mm]", "CUT_H [mm]", "Kant", "Napomena"]))
        story.append(Spacer(1, 3 * mm))
    _svc_proc = display_service_packet.get("service_processing", pd.DataFrame())
    if _svc_proc is not None and not _svc_proc.empty:
        _wt_spec = _svc_proc[
            _svc_proc["Deo"].astype(str).str.lower().isin(["radna ploča", "radna ploca", "worktop"])
        ].copy() if "Deo" in _svc_proc.columns else pd.DataFrame()
        if _wt_spec is not None and not _wt_spec.empty:
            story.append(Paragraph(_t("Specifikacija radne ploče", "Worktop specification"), s_h2))
            story.append(Paragraph(
                _t(
                    "Servis radi isključivo po CUT merama. Finalni rez i prilagođavanje rade se na licu mesta.",
                    "The workshop works strictly by CUT dimensions. Final trimming and fitting are done on site.",
                ),
                s_norm,
            ))
            _wt_cols = ["Zid", "Modul", "Kol.", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena"]
            story.append(_df_to_table(_wt_spec, _wt_cols))
            story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(_t("Tabela za obradu", "Machining table"), s_norm))
        story.append(_df_to_table(_svc_proc, ["PartCode", "Zid", "Modul", "Deo", "CUT_W [mm]", "CUT_H [mm]", "Tip obrade", "Izvodi", "Osnov izvođenja", "Kol.", "Obrada / napomena"]))
        story.append(Spacer(1, 3 * mm))
    _svc_instr = display_service_packet.get("service_instructions", pd.DataFrame())
    if _svc_instr is not None and not _svc_instr.empty:
        story.append(Paragraph(_t("Instrukcije za servis", "Workshop instructions"), s_norm))
        story.append(_df_to_table(_svc_instr, ["RB", "Stavka", "Instrukcija"]))
        story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(_t("Napomena: uređaji, sudopera, slavina, sifon i slično kupuju se kao gotovi proizvodi i ne ulaze u sečenje.", "Note: appliances, sink, tap, trap, and similar items are purchased as ready-made products and are not part of the cutting list."), s_norm))
    story.append(Spacer(1, 8 * mm))

    # ---- Montazna mapa po modulu ----
    story.append(Paragraph(_t("ZA MONTAŽU", "FOR ASSEMBLY"), s_h2))
    story.append(Paragraph(
        _t(
            "Ovde pratiš redosled sklapanja, raspored delova i završne kontrolne liste.",
            "Use this section for assembly order, part orientation and final checklists.",
        ),
        s_norm,
    ))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(_t("Montažna mapa po modulu (orijentacija delova)", "Assembly map by unit (part orientation)"), s_h2))
    _all_for_map = []
    for _k in ("carcass", "backs", "fronts", "drawer_boxes", "worktop", "plinth"):
        _dfk = display_sections.get(_k, pd.DataFrame())
        if _dfk is not None and not _dfk.empty:
            _all_for_map.append(_dfk.copy())
    if _all_for_map:
        _mm = pd.concat(_all_for_map, ignore_index=True)
        if "ID" in _mm.columns:
            _mm["ID"] = pd.to_numeric(_mm["ID"], errors="coerce")
            _mm = _mm[_mm["ID"].notna()].copy()
        if not _mm.empty:
            _mm["ID"] = _mm["ID"].astype(int)
            _mm["_korak"] = pd.to_numeric(_mm.get("SklopKorak", "-"), errors="coerce").fillna(99).astype(int)
            for _mid in sorted(_mm["ID"].unique().tolist()):
                _mrows = _mm[_mm["ID"] == _mid].sort_values(["_korak", "PartCode"]).copy()
                _modul = str(_mrows.iloc[0].get("Modul", _t(f"Modul {_mid}", f"Module {_mid}")))
                story.append(Paragraph(f"#{_mid} - {_modul}", s_norm))

                _map_data = [
                    ["", _t("GORE", "TOP"), ""],
                    [_t("LEVO", "LEFT"), _t("CENTAR", "CENTER"), _t("DESNO", "RIGHT")],
                    ["", _t("DOLE", "BOTTOM"), ""],
                ]
                _map_tbl = Table(_map_data, colWidths=[18 * mm, 22 * mm, 18 * mm], rowHeights=[8 * mm, 10 * mm, 8 * mm])
                _map_tbl.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONT", (0, 0), (-1, -1), FONT_BOLD),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#F3F3F3")),
                ]))

                _plist = _mrows.copy()
                if "Kol." in _plist.columns:
                    _plist = _plist.rename(columns={"Kol.": "Kom"})
                _cols_mp = ["PartCode", "Deo", "Pozicija", "SklopKorak", "Kom"]
                _avail = [c for c in _cols_mp if c in _plist.columns]
                _pt = _df_to_table(_plist, _avail)

                _row = Table([[_map_tbl, _pt]], colWidths=[62 * mm, 190 * mm])
                _row.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]))
                story.append(_row)
                story.append(Spacer(1, 2 * mm))

    story.append(Spacer(1, 8 * mm))

    # ---- Po elementima: detalji i sklapanje ----
    _mods = kitchen.get("modules", []) or []
    if _all_for_map and _mods:
        story.append(PageBreak())
        story.append(Paragraph(_t("Po elementima - detalji i sklapanje", "By unit - details and assembly"), s_h2))
        story.append(Paragraph(
            _t(
                "Za svaki element su povezani slika, oznake delova, tabela i redosled sklapanja.",
                "Each unit links the image, part labels, table and assembly order.",
            ),
            s_norm,
        ))
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(_t("Legenda oznaka", "Label legend"), s_norm))
        story.append(Paragraph(_t("C = korpus, B = leđa, F = front, D = fioka", "C = carcass, B = back, F = front, D = drawer"), s_norm))
        story.append(Paragraph(_kant_legend(_lang), s_norm))
        story.append(Spacer(1, 4 * mm))

        _parts_all = pd.concat(_all_for_map, ignore_index=True)
        if "ID" in _parts_all.columns:
            _parts_all["ID"] = pd.to_numeric(_parts_all["ID"], errors="coerce")
            _parts_all = _parts_all[_parts_all["ID"].notna()].copy()
            _parts_all["ID"] = _parts_all["ID"].astype(int)

        for _m in _mods:
            _mid = int(_m.get("id", 0) or 0)
            _mparts = _parts_all[_parts_all["ID"] == _mid].copy() if not _parts_all.empty else pd.DataFrame()
            if _mparts.empty:
                continue

            _mlbl = _normalize_sr_export_text(_m.get("label", "") or "")
            if _lang in {"es", "pt-br", "ru"}:
                _mlbl = _translate_export_text(_mlbl, _lang, "Modul")
            _mtid = str(_m.get("template_id", "") or "")
            _mz = str(_m.get("zone", "") or "").lower()
            _mw = int(_m.get("w_mm", 0) or 0)
            _mh = int(_m.get("h_mm", 0) or 0)
            _md = int(_m.get("d_mm", 0) or 0)
            _wall = str(_m.get("wall_key", "") or "")

            story.append(Paragraph(_pdf_clean_text(f"#{_mid} - {_mlbl}"), s_h2))
            if _lang == "pt-br":
                _module_meta = f"Tipo: {_mtid}  |  Dimensões: {_mw} x {_mh} x {_md} mm" + (f"  |  Parede: {_wall}" if _wall else "")
            elif _lang == "es":
                _module_meta = f"Tipo: {_mtid}  |  Medidas: {_mw} x {_mh} x {_md} mm" + (f"  |  Pared: {_wall}" if _wall else "")
            elif _lang == "ru":
                _module_meta = f"Тип: {_mtid}  |  Размеры: {_mw} x {_mh} x {_md} мм" + (f"  |  Стена: {_wall}" if _wall else "")
            elif _lang == "zh-cn":
                _module_meta = f"类型: {_mtid}  |  尺寸: {_mw} x {_mh} x {_md} mm" + (f"  |  墙面: {_wall}" if _wall else "")
            elif _lang == "hi":
                _module_meta = f"प्रकार: {_mtid}  |  माप: {_mw} x {_mh} x {_md} mm" + (f"  |  दीवार: {_wall}" if _wall else "")
            else:
                _module_meta = _t(
                    f"Tip: {_mtid}  |  Dimenzije: {_mw} x {_mh} x {_md} mm" + (f"  |  Zid: {_wall}" if _wall else ""),
                    f"Type: {_mtid}  |  Dimensions: {_mw} x {_mh} x {_md} mm" + (f"  |  Wall: {_wall}" if _wall else ""),
                )
            story.append(Paragraph(_pdf_clean_text(_module_meta), s_norm))
            story.append(Spacer(1, 1 * mm))

            try:
                from visualization import render_element_preview  # local import to avoid circular import
                _uri_2d, _uri_3d = render_element_preview(
                    _m,
                    kitchen,
                    label_mode="part_codes",
                    part_rows=_mparts.to_dict("records"),
                )
                _img2d = _rl_image_from_uri(_uri_2d, 72)
                _img3d = _rl_image_from_uri(_uri_3d, 72)
                _img_rows = []
                if _img2d:
                    _img_rows.append([Paragraph("2D", s_norm)])
                    _img_rows.append([_img2d])
                if _img3d:
                    _img_rows.append([Paragraph("3D", s_norm)])
                    _img_rows.append([_img3d])
                if _img_rows:
                    _img_tbl = Table(_img_rows, colWidths=[75 * mm])
                    _img_tbl.setStyle(TableStyle([
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                    ]))
                    story.append(_img_tbl)
                    story.append(Paragraph(_t("Oznake na slici prate istu oznaku kao u tabeli i koracima.", "Image labels use the same labels as the table and steps."), s_norm))
                    story.append(Spacer(1, 2 * mm))
            except Exception as ex:
                _LOG.debug("By-unit preview render failed for module id=%s: %s", _mid, ex)

            _map_parts = _mparts[["PartCode", "Deo", "Pozicija", "SklopKorak", "Kol."]].copy()
            _map_parts["Oznaka"] = _map_parts["PartCode"].map(_short_part_code)
            _map_parts["Deo"] = _map_parts["Deo"].map(lambda v: _friendly_part_name(v, _lang))
            _map_parts["Pozicija"] = _map_parts["Pozicija"].map(lambda v: _friendly_position_name(v, _lang))
            _map_parts = _map_parts.sort_values(["SklopKorak", "PartCode"]).reset_index(drop=True)
            story.append(Paragraph(_t("Mapa delova za sklapanje", "Assembly parts map"), s_norm))
            story.append(_df_to_table(_map_parts.rename(columns={"Oznaka": _t("Oznaka", "Label"), "Deo": _t("Deo", "Part"), "Pozicija": _t("Gde ide", "Where it goes"), "SklopKorak": _t("Korak", "Step"), "Kol.": _t("Kom.", "Qty")}), [_t("Oznaka", "Label"), _t("Deo", "Part"), _t("Gde ide", "Where it goes"), _t("Korak", "Step"), _t("Kom.", "Qty")]))
            story.append(Spacer(1, 2 * mm))

            _cuts = _mparts.copy()
            _cuts["Oznaka"] = _cuts["PartCode"].map(_short_part_code)
            _cuts["Deo"] = _cuts["Deo"].map(lambda v: _friendly_part_name(v, _lang))
            for _fin_col, _cut_col in (("Dužina [mm]", "CUT_W [mm]"), ("Širina [mm]", "CUT_H [mm]")):
                if _cut_col in _cuts.columns:
                    if _fin_col not in _cuts.columns:
                        _cuts[_fin_col] = _cuts[_cut_col]
                    else:
                        _cuts[_fin_col] = _cuts[_fin_col].where(
                            _cuts[_fin_col].notna() & (_cuts[_fin_col].astype(str).str.strip() != ""),
                            _cuts[_cut_col],
                        )
            story.append(Paragraph(_t("Rezovi", "Cut parts"), s_norm))
            story.append(_df_to_table(_cuts.rename(columns={"Oznaka": _t("Oznaka", "Label"), "Deo": _t("Deo", "Part"), "Kol.": _t("Kom.", "Qty"), "Dužina [mm]": _t("Dužina [mm]", "Length [mm]"), "Širina [mm]": _t("Širina [mm]", "Width [mm]"), "Deb.": _t("Deb.", "Thk."), "Kant": _t("Kant", "Edge")}), [_t("Oznaka", "Label"), _t("Deo", "Part"), "Pozicija", "SklopKorak", _t("Dužina [mm]", "Length [mm]"), _t("Širina [mm]", "Width [mm]"), _t("Deb.", "Thk."), _t("Kom.", "Qty"), _t("Kant", "Edge")]))
            story.append(Spacer(1, 2 * mm))

            _role_notes: list[str] = []
            for _, _pr in _mparts.iterrows():
                _note = _part_role_note(
                    _friendly_part_name(_pr.get("Deo", ""), _lang),
                    _pr.get("Materijal", ""),
                    _pr.get("Deb.", ""),
                    _lang,
                )
                if _note and _note not in _role_notes:
                    _role_notes.append(_note)
            if _role_notes:
                story.append(Paragraph(_t("Važne napomene o delovima", "Important part notes"), s_norm))
                for _note in _role_notes:
                    story.append(Paragraph(_pdf_clean_text(f"• {_note}"), s_norm))
                story.append(Spacer(1, 1 * mm))

            story.append(Paragraph(_t("Potreban alat i okov", "Required tools and hardware"), s_norm))
            for _line in _module_tool_hardware_lines(_mtid, _mz, _lang):
                story.append(Paragraph(_pdf_clean_text(f"• {_line}"), s_norm))
            story.append(Spacer(1, 1 * mm))

            story.append(Paragraph(_t("Proveri pre sklapanja", "Check before assembly"), s_norm))
            for _line in _module_preassembly_lines(_mtid, _mz, _lang):
                story.append(Paragraph(_pdf_clean_text(f"• {_line}"), s_norm))
            story.append(Spacer(1, 1 * mm))

            story.append(Paragraph(_t("Uputstvo za montažu", "Assembly instructions"), s_norm))
            for _line in assembly_instructions(_mtid, _mz, m=_m, kitchen=kitchen, lang=_lang):
                _txt = _pdf_clean_text(_line)
                if not str(_txt).strip():
                    story.append(Spacer(1, 1 * mm))
                elif str(_txt).strip().startswith("--"):
                    story.append(Paragraph(_txt.replace("--", "").strip(), s_h2))
                else:
                    story.append(Paragraph(_txt, s_norm))

            story.append(Spacer(1, 4 * mm))
            story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey))
            story.append(Spacer(1, 4 * mm))

    # ---- Okovi ----
    story.append(Paragraph(_t("Okovi i potrošni materijal", "Hardware and consumables"), s_h2))
    df_hw = display_sections.get("hardware", pd.DataFrame())
    if df_hw is not None and not df_hw.empty:
        if "Kategorija" in df_hw.columns:
            _wdf = df_hw[df_hw["Kategorija"].astype(str).str.lower() == "warning"]
            if _wdf is not None and not _wdf.empty:
                story.append(Paragraph(_t("Kritična upozorenja pre proizvodnje", "Critical warnings before production"), s_h2))
                _cols_w = ["Modul", "Naziv", "Napomena"]
                story.append(_df_to_table(_wdf, _cols_w))
                story.append(Spacer(1, 4 * mm))
        _cols_hw = ["PartCode", "Modul", "Kategorija", "Naziv", "Tip / Šifra", "Kol.", "SklopKorak", "Napomena"]
        story.append(_df_to_table(df_hw, _cols_hw))

    # ---- Shopping list ----
    _shop_df = display_service_packet.get("shopping_list", pd.DataFrame())
    if _shop_df is not None and not _shop_df.empty:
        story.append(Spacer(1, 8 * mm))
        story.append(Paragraph(_t("Lista kupovine", "Shopping list"), s_h2))
        story.append(Paragraph(_t("Ovde su stvari koje kupuješ posebno i ne idu na sečenje.", "These are items you purchase separately and they are not part of cutting."), s_norm))
        story.append(_df_to_table(_shop_df, ["Grupa", "Naziv", "Tip / Šifra", "Kol.", "Napomena"]))

    _ready_df = display_service_packet.get("ready_made_items", pd.DataFrame())
    if _ready_df is not None and not _ready_df.empty:
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(_t("Kupuje se kao gotov proizvod — ne ulazi u sečenje", "Purchased ready-made - not included in cutting"), s_h2))
        story.append(_df_to_table(_ready_df, ["Grupa", "Naziv", "Tip / Šifra", "Kol.", "Napomena"]))

    _guide_df = display_service_packet.get("user_guide", pd.DataFrame())
    if _guide_df is not None and not _guide_df.empty:
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(_t("Kratak vodič za korisnika", "Quick user guide"), s_h2))
        story.append(_df_to_table(_guide_df, ["Korak", "Šta radiš", "Napomena"]))

    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(_t("Kontrolna lista pre servisa i sklapanja", "Checklist before workshop and assembly"), s_h2))
    _wcl = display_service_packet.get("workshop_checklist", pd.DataFrame())
    if _wcl is not None and not _wcl.empty:
        story.append(Paragraph(_t("Pre odlaska u servis", "Before going to the workshop"), s_norm))
        story.append(_df_to_table(_wcl, ["RB", "Stavka", "Status"]))
        story.append(Spacer(1, 3 * mm))
    _hcl = display_service_packet.get("home_checklist", pd.DataFrame())
    if _hcl is not None and not _hcl.empty:
        story.append(Paragraph(_t("Pre kućnog sklapanja", "Before home assembly"), s_norm))
        story.append(_df_to_table(_hcl, ["RB", "Stavka", "Status"]))

    doc.build(story)
    return buf.getvalue()


def generate_cutlist_pdf(
    kitchen: Dict[str, Any],
    title: str = "Krojna lista PRO - M1 (jedan zid)",
    lang: str = "sr",
) -> bytes:
    """Returns PDF bytes."""
    if str(lang or "sr").lower().strip() == "en" and title in {"Krojna lista PRO", "Krojna lista PRO - M1 (jedan zid)", "Krojna lista PRO – M1 (jedan zid)"}:
        title = "CabinetCut PRO" if title == "Krojna lista PRO" else "CabinetCut PRO - M1 (single wall)"
    if str(lang or "sr").lower().strip() == "es" and title in {"Krojna lista PRO", "Krojna lista PRO - M1 (jedan zid)", "Krojna lista PRO – M1 (jedan zid)"}:
        title = "CabinetCut PRO" if title == "Krojna lista PRO" else "CabinetCut PRO - M1 (una pared)"
    if str(lang or "sr").lower().strip() == "ru" and title in {"Krojna lista PRO", "Krojna lista PRO - M1 (jedan zid)", "Krojna lista PRO – M1 (jedan zid)"}:
        title = "CabinetCut PRO" if title == "Krojna lista PRO" else "CabinetCut PRO - M1 (одна стена)"
    if str(lang or "sr").lower().strip() == "pt-br" and title in {"Krojna lista PRO", "Krojna lista PRO - M1 (jedan zid)", "Krojna lista PRO – M1 (jedan zid)"}:
        title = "CabinetCut PRO" if title == "Krojna lista PRO" else "CabinetCut PRO - M1 (uma parede)"
    if str(lang or "sr").lower().strip() == "zh-cn" and title in {"Krojna lista PRO", "Krojna lista PRO - M1 (jedan zid)", "Krojna lista PRO – M1 (jedan zid)"}:
        title = "CabinetCut PRO" if title == "Krojna lista PRO" else "CabinetCut PRO - M1 (单面墙)"
    if str(lang or "sr").lower().strip() == "hi" and title in {"Krojna lista PRO", "Krojna lista PRO - M1 (jedan zid)", "Krojna lista PRO – M1 (jedan zid)"}:
        title = "CabinetCut PRO" if title == "Krojna lista PRO" else "CabinetCut PRO - M1 (एक दीवार)"
    final_ds = get_final_cutlist_dataset(kitchen, lang=lang)
    sections = final_ds["sections"]
    return build_cutlist_pdf_bytes(kitchen, sections, project_title=title, lang=lang)


def generate_cutlist_excel(
    kitchen: Dict[str, Any],
    title: str = "Krojna lista PRO",
    lang: str = "sr",
) -> bytes:
    """
    Generiše Excel fajl (.xlsx) sa kompletnom krojnom listom.

    Struktura sheetova:
      Sheet 1  "Pregled"   — sve ploče agregirane po materijalu + deb + dimenzijama
      Sheet 2  "Korpus"    — carcass ploče (stranice, dno, plafon)
      Sheet 3  "Ledne"     — leđne ploče
      Sheet 4  "Frontovi"  — frontovi
      Sheet 5  "Fioke"     — sanduci fioka
      Sheet 6  "Radna"     — radna ploča i nosači
      Sheet 7  "Sokla"     — sokla / lajsna

    Vraća bytes kompletnog .xlsx fajla.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)
    if _lang != "sr" and title == "Krojna lista PRO":
        title = "CabinetCut PRO"

    final_ds = get_final_cutlist_dataset(kitchen, lang=_lang)
    sections = final_ds["sections"]
    service_packet = final_ds["service_packet"]
    mats = kitchen.get("materials", {}) or {}
    wall = kitchen.get("wall", {}) or {}

    # ── Boje ────────────────────────────────────────────────────────────────
    CLR_HDR_BG = "1F3864"   # tamno plava — header
    CLR_HDR_FG = "FFFFFF"
    CLR_TTL_BG = "2E75B6"   # srednje plava — naslov lista
    CLR_TTL_FG = "FFFFFF"
    CLR_ODD    = "EBF3FB"   # svijetlo plava — neparne vrste
    CLR_EVEN   = "FFFFFF"
    CLR_WARN   = "FFF2CC"   # žuta — vrste sa greškom
    CLR_BORDER = "BFBFBF"

    def _hdr_fill():  return PatternFill("solid", fgColor=CLR_HDR_BG)
    def _ttl_fill():  return PatternFill("solid", fgColor=CLR_TTL_BG)
    def _odd_fill():  return PatternFill("solid", fgColor=CLR_ODD)
    def _warn_fill(): return PatternFill("solid", fgColor=CLR_WARN)
    def _border():
        s = Side(style="thin", color=CLR_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr_font():  return Font(name="Calibri", bold=True,  color=CLR_HDR_FG, size=10)
    def _ttl_font():  return Font(name="Calibri", bold=True,  color=CLR_TTL_FG, size=11)
    def _norm_font(): return Font(name="Calibri", bold=False, size=10)
    def _info_font(): return Font(name="Calibri", italic=True, size=9, color="555555")

    _now  = datetime.now().strftime("%d.%m.%Y  %H:%M")
    _wl   = wall.get("length_mm", 0)
    _wh   = wall.get("height_mm", 0)
    _info = (
        f"{_t('Zid', 'Wall')}: {_wl}×{_wh} mm   |   "
        f"{_format_material_role(mats.get('carcass_material','?'), mats.get('carcass_thk','?'), 'carcass', _lang)}   |   "
        f"{_format_material_role(mats.get('front_material','?'), mats.get('front_thk','?'), 'front', _lang)}   |   "
        f"{_format_material_role((mats.get('back_material') or ''), (mats.get('back_thk') or ''), 'back', _lang)}   |   "
        f"{_t('Generisano', 'Generated')}: {_now}"
    )

    # ── Specifikacija kolona ─────────────────────────────────────────────────
    # (df_field, excel_header, col_width)
    SEC_COLS = [
        ("PartCode",    "PartCode",      12),
        ("Zid",         _t("Zid", "Wall"),           10),
        ("Modul",       _t("Modul", "Module"),        22),
        ("Deo",         _t("Deo", "Part"),          24),
        ("Pozicija",    _t("Pozicija", "Position"),      9),
        ("SklopKorak",  _t("Korak", "Step"),         7),
        ("Materijal",   _t("Materijal", "Material"),    12),
        ("Deb.",        "Deb.",          6),
        ("Kol.",        _t("Kol.", "Qty"),          5),
        ("CUT_W [mm]",  _t("CUT Duž.", "CUT Length"),      9),
        ("CUT_H [mm]",  _t("CUT Šir.", "CUT Width"),      9),
        ("Dužina [mm]", _t("FIN Duž.", "FIN Length"),      9),
        ("Širina [mm]", _t("FIN Šir.", "FIN Width"),      9),
        ("Smer goda",   _t("Smer goda", "Grain"),     8),
        ("Kant",        _t("Kant", "Edge"),         24),
        ("L1",          "L1",            4),
        ("L2",          "L2",            4),
        ("K1",          "K1",            4),
        ("K2",          "K2",            4),
        ("Napomena",    _t("Napomena", "Note"),     30),
    ]
    HARDWARE_COLS = [
        ("PartCode",    "PartCode",      12),
        ("Zid",         _t("Zid", "Wall"),           10),
        ("Modul",       _t("Modul", "Module"),         22),
        ("Kategorija",  _t("Kategorija", "Category"),    12),
        ("Naziv",       _t("Naziv", "Name"),         20),
        ("Tip / Šifra", _t("Tip / Sifra", "Type / Code"),   20),
        ("Kol.",        _t("Kol.", "Qty"),           6),
        ("SklopKorak",  _t("Korak", "Step"),          7),
        ("Napomena",    _t("Napomena", "Note"),      32),
    ]
    PROJECT_COLS = [
        ("Polje", _t("Polje", "Field"), 24),
        ("Vrednost", _t("Vrednost", "Value"), 60),
    ]
    SERVICE_CUTS_COLS = [
        ("RB", "RB", 6),
        ("Zid", _t("Zid", "Wall"), 10),
        ("Materijal", _t("Materijal", "Material"), 18),
        ("Deb.", "Deb.", 8),
        ("CUT_W [mm]", _t("CUT Dužina", "CUT Length"), 12),
        ("CUT_H [mm]", _t("CUT Širina", "CUT Width"), 12),
        ("Kant", _t("Kant", "Edge"), 24),
        ("Kol.", _t("Kol.", "Qty"), 8),
        ("Napomena za servis", _t("Napomena za servis", "Workshop note"), 42),
    ]
    SERVICE_EDGE_COLS = [
        ("PartCode", "PartCode", 12),
        ("Zid", _t("Zid", "Wall"), 10),
        ("Modul", _t("Modul", "Module"), 22),
        ("Deo", _t("Deo", "Part"), 24),
        ("Kol.", _t("Kol.", "Qty"), 6),
        ("CUT_W [mm]", _t("CUT Dužina", "CUT Length"), 10),
        ("CUT_H [mm]", _t("CUT Širina", "CUT Width"), 10),
        ("Kant", _t("Kant", "Edge"), 24),
        ("Napomena", _t("Napomena", "Note"), 36),
    ]
    SERVICE_PROC_COLS = [
        ("PartCode", "PartCode", 12),
        ("Zid", _t("Zid", "Wall"), 10),
        ("Modul", _t("Modul", "Module"), 22),
        ("Deo", _t("Deo", "Part"), 24),
        ("CUT_W [mm]", _t("CUT Dužina", "CUT Length"), 10),
        ("CUT_H [mm]", _t("CUT Širina", "CUT Width"), 10),
        ("Tip obrade", _t("Tip obrade", "Processing type"), 18),
        ("Izvodi", _t("Izvodi", "Operations"), 12),
        ("Osnov izvođenja", _t("Osnov izvođenja", "Execution basis"), 20),
        ("Kol.", _t("Kol.", "Qty"), 6),
        ("Obrada / napomena", _t("Obrada / napomena", "Processing / note"), 30),
    ]
    SERVICE_INSTR_COLS = [
        ("RB", "RB", 6),
        ("Stavka", _t("Stavka", "Item"), 20),
        ("Instrukcija", _t("Instrukcija", "Instruction"), 70),
    ]
    SHOP_COLS = [
        ("Grupa", _t("Grupa", "Group"), 20),
        ("Naziv", _t("Naziv", "Name"), 24),
        ("Tip / Šifra", _t("Tip / Sifra", "Type / Code"), 22),
        ("Kol.", _t("Kol.", "Qty"), 8),
        ("Napomena", _t("Napomena", "Note"), 40),
    ]
    GUIDE_COLS = [
        ("Korak", _t("Korak", "Step"), 8),
        ("Šta radiš", _t("Šta radiš", "What you do"), 36),
        ("Napomena", _t("Napomena", "Note"), 48),
    ]
    CHECK_COLS = [
        ("RB", "RB", 6),
        ("Stavka", _t("Stavka", "Item"), 70),
        ("Status", _t("Status", "Status"), 14),
    ]
    SUM_COLS = [
        ("Materijal",   _t("Materijal", "Material"),    14),
        ("Deb.",        "Deb.",          6),
        ("CUT_W [mm]",  _t("CUT Dužina", "CUT Length"),    9),
        ("CUT_H [mm]",  _t("CUT Širina", "CUT Width"),    9),
        ("Dužina [mm]", _t("FIN Dužina", "FIN Length"),    9),
        ("Širina [mm]", _t("FIN Širina", "FIN Width"),    9),
        ("Kol.",        _t("Kol.", "Qty"),          6),
        ("Kant",        _t("Kant", "Edge"),         24),
    ]
    NUM_FIELDS = {
        "Deb.", "Kol.", "CUT_W [mm]", "CUT_H [mm]",
        "Dužina [mm]", "Širina [mm]", "L1", "L2", "K1", "K2",
    }

    # ── Pomoćna funkcija za pisanje jednog lista ─────────────────────────────
    def _write_sheet(
        ws,
        df,
        cols: list,
        sheet_title: str,
    ) -> None:
        n = len(cols)

        # Red 1: Naslov lista
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
        c = ws.cell(1, 1, sheet_title)
        c.font      = _ttl_font()
        c.fill      = _ttl_fill()
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 22

        # Red 2: Info o projektu
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        c = ws.cell(2, 1, _info)
        c.font      = _info_font()
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 15

        # Red 3: Zaglavlje kolona
        for ci, (field, lbl, w) in enumerate(cols, 1):
            c = ws.cell(3, ci, lbl)
            c.font      = _hdr_font()
            c.fill      = _hdr_fill()
            c.border    = _border()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(n)}3"

        # Prazna sekcija
        if df is None or df.empty:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n)
            c = ws.cell(4, 1, _t("— nema podataka —", "— no data —"))
            c.font      = _info_font()
            c.alignment = Alignment(horizontal="center")
            return

        # Redovi s podacima
        for ri, (_, row) in enumerate(df.iterrows(), 4):
            nap = str(row.get("Napomena", "") or "")
            _nap_u = nap.upper()
            if ("GREŠKA" in _nap_u or "GRJEŠKA" in _nap_u or "UPOZORENJE" in _nap_u
                    or "WARNING" in _nap_u or str(row.get("Kategorija", "")).lower() == "warning"):
                row_fill = _warn_fill()
            elif ri % 2 == 0:
                row_fill = _odd_fill()
            else:
                row_fill = PatternFill("solid", fgColor=CLR_EVEN)

            for ci, (field, _lbl, _w) in enumerate(cols, 1):
                raw = row.get(field, "")
                if field in NUM_FIELDS:
                    try:
                        val = round(float(raw), 1) if raw != "" and raw is not None else ""
                    except (TypeError, ValueError):
                        val = raw
                else:
                    val = raw if raw is not None else ""

                c = ws.cell(ri, ci, val)
                c.font   = _norm_font()
                c.fill   = row_fill
                c.border = _border()
                if field in NUM_FIELDS:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    def _extract_worktop_spec_df(proc_df: pd.DataFrame) -> pd.DataFrame:
        cols = ["Zid", "Modul", "Kol.", "Tip obrade", "Izvodi", "Osnov izvođenja", "Obrada / napomena"]
        if proc_df is None or proc_df.empty or "Deo" not in proc_df.columns:
            return pd.DataFrame(columns=cols)
        mask = proc_df["Deo"].astype(str).str.lower().isin(["radna ploča", "radna ploca", "worktop", "bancada"])
        out = proc_df.loc[mask, [c for c in cols if c in proc_df.columns]].copy()
        for col in cols:
            if col not in out.columns:
                out[col] = ""
        return out[cols]

    # ── Kreira workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # ukloni default prazan sheet

    # ── Uvodni sheet: kako koristiti dokument ───────────────────────────────
    ws_intro = wb.create_sheet(_t("Kako koristiti", "How to use"), 0)
    _intro_df = pd.DataFrame([
        {"Korak": 1, "Šta radiš": _t("Prvo proveri osnovne mere i materijale", "First check the main dimensions and materials"), "Napomena": _t("Ako ovde vidiš grešku, nemoj slati dokument u servis.", "If you see an error here, do not send the document to the workshop.")},
        {"Korak": 2, "Šta radiš": _t("Za servis koristi samo sheet-ove za rezanje, kantovanje i obradu", "Use only the cutting, edging and machining sheets for the workshop"), "Napomena": _t("Servis radi po CUT merama i napomenama iz tih tabela.", "The workshop works by CUT dimensions and notes from those sheets.")},
        {"Korak": 3, "Šta radiš": _t("Posebno kupi gotove uređaje, okove i alat", "Purchase ready-made appliances, hardware and tools separately"), "Napomena": _t("To nije deo sečenja i ne izrađuje se iz pločastog materijala.", "These items are not cut from board material.")},
        {"Korak": 4, "Šta radiš": _t("Za sklapanje koristi vodič i checkliste", "Use the guide and checklists during assembly"), "Napomena": _t("Radi redom: korpusi, frontovi, fioke, uređaji, završna provera.", "Follow the order: carcasses, fronts, drawers, appliances, final check.")},
    ])
    _write_sheet(ws_intro, _intro_df, GUIDE_COLS, f"{title}  |  {_t('Kako koristiti dokument', 'How to use this document')}")

    # ── Sheet 1: Pregled ─────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(_t("Pregled", "Overview"))
    all_dfs = [df for df in sections.values() if df is not None and not df.empty]
    if all_dfs:
        combined = pd.concat(all_dfs, ignore_index=True)
        # Osiguraj da kolone postoje
        for col in ["Dužina [mm]", "Širina [mm]", "Kant"]:
            if col not in combined.columns:
                combined[col] = ""
        grp = [c for c in
               ["Materijal", "Deb.", "CUT_W [mm]", "CUT_H [mm]", "Dužina [mm]", "Širina [mm]", "Kant"]
               if c in combined.columns]
        summary = (
            combined
            .groupby(grp, as_index=False)
            .agg({"Kol.": "sum"})
            .sort_values(
                ["Materijal", "CUT_W [mm]", "CUT_H [mm]"],
                ascending=[True, False, False],
            )
            .reset_index(drop=True)
        )
    else:
        summary = pd.DataFrame(columns=[f for f, _, _ in SUM_COLS])
    _write_sheet(ws_sum, summary, SUM_COLS, f"{title}  |  {_t('Pregled svih ploča', 'Overview of all panels')}")

    # ── Sheet "Rezanje" — sumarna krojna lista po PDF formatu (strana 4) ──────
    REZANJE_COLS = [
        ("RB",           "RB",            4),
        ("Deo",          _t("Deo", "Part"),          24),
        ("Kom",          _t("Kom", "Qty"),           5),
        ("Dužina [mm]",  _t("Dužina [mm]", "Length [mm]"),  10),
        ("Širina [mm]",  _t("Širina [mm]", "Width [mm]"),  10),
        ("Deb.",         "Deb. [mm]",     8),
        ("Materijal",    _t("Materijal", "Material"),    16),
        ("Smer goda",    _t("Orijent.", "Grain"),     10),
        ("L1",           ("Borda L1" if _lang == "pt-br" else ("Borde L1" if _lang == "es" else ("Кромка L1" if _lang == "ru" else ("封边 L1" if _lang == "zh-cn" else ("एज L1" if _lang == "hi" else _t("Kant L1", "Edge L1")))))),       7),
        ("L2",           ("Borda L2" if _lang == "pt-br" else ("Borde L2" if _lang == "es" else ("Кромка L2" if _lang == "ru" else ("封边 L2" if _lang == "zh-cn" else ("एज L2" if _lang == "hi" else _t("Kant L2", "Edge L2")))))),       7),
        ("K1",           ("Borda K1" if _lang == "pt-br" else ("Borde K1" if _lang == "es" else ("Кромка K1" if _lang == "ru" else ("封边 K1" if _lang == "zh-cn" else ("एज K1" if _lang == "hi" else _t("Kant K1", "Edge K1")))))),       7),
        ("K2",           ("Borda K2" if _lang == "pt-br" else ("Borde K2" if _lang == "es" else ("Кромка K2" if _lang == "ru" else ("封边 K2" if _lang == "zh-cn" else ("एज K2" if _lang == "hi" else _t("Kant K2", "Edge K2")))))),       7),
    ]
    DET_COLS = [
        ("PartCode",     "PartCode",     12),
        ("Modul",        _t("Modul", "Module"),        22),
        ("Deo",          _t("Deo", "Part"),          24),
        ("Kom",          _t("Kom", "Qty"),           5),
        ("Dužina [mm]",  _t("Dužina [mm]", "Length [mm]"),  10),
        ("Širina [mm]",  _t("Širina [mm]", "Width [mm]"),  10),
        ("Deb.",         "Deb. [mm]",     8),
        ("Materijal",    _t("Materijal", "Material"),    16),
        ("Smer goda",    _t("Orijent.", "Grain"),     10),
        ("L1",           ("Borda L1" if _lang == "pt-br" else ("Borde L1" if _lang == "es" else ("Кромка L1" if _lang == "ru" else ("封边 L1" if _lang == "zh-cn" else ("एज L1" if _lang == "hi" else _t("Kant L1", "Edge L1")))))),       7),
        ("L2",           ("Borda L2" if _lang == "pt-br" else ("Borde L2" if _lang == "es" else ("Кромка L2" if _lang == "ru" else ("封边 L2" if _lang == "zh-cn" else ("एज L2" if _lang == "hi" else _t("Kant L2", "Edge L2")))))),       7),
        ("K1",           ("Borda K1" if _lang == "pt-br" else ("Borde K1" if _lang == "es" else ("Кромка K1" if _lang == "ru" else ("封边 K1" if _lang == "zh-cn" else ("एज K1" if _lang == "hi" else _t("Kant K1", "Edge K1")))))),       7),
        ("K2",           ("Borda K2" if _lang == "pt-br" else ("Borde K2" if _lang == "es" else ("Кромка K2" if _lang == "ru" else ("封边 K2" if _lang == "zh-cn" else ("एज K2" if _lang == "hi" else _t("Kant K2", "Edge K2")))))),       7),
    ]
    try:
        _sum2 = final_ds["summary"]
        ws_rez = wb.create_sheet(_t("Rezanje", "Cutting"))
        _write_sheet(ws_rez, _sum2.get("summary_all"), REZANJE_COLS,
                     f"{title}  |  {_t('Sumarna krojna lista ploča', 'Summary cut list of panels')}")
        ws_det = wb.create_sheet(_t("Po modulima", "By units"))
        _write_sheet(ws_det, _sum2.get("summary_detaljna"), DET_COLS,
                     f"{title}  |  {_t('Detaljna krojna lista po modulima', 'Detailed cut list by units')}")
    except Exception as _sex:
        _LOG.warning("generate_cutlist_summary failed, preskacam Rezanje sheet: %s", _sex)

    # ── Sheetovi po sekcijama (detalji) ──────────────────────────────────────
    SHEET_CFG = [
        ("carcass",      _t("Korpus", "Carcass"),    _t("Korpus — stranice, dno, plafon", "Carcass — sides, bottom, top")),
        ("backs",        _t("Leđa", "Backs"),        _t("Leđne ploče", "Back panels")),
        ("fronts",       _t("Frontovi", "Fronts"),   _t("Frontovi", "Fronts")),
        ("drawer_boxes", _t("Fioke", "Drawers"),     _t("Sanduk fioke", "Drawer box")),
        ("worktop",      _t("Radna", "Worktop"),     _t("Radna ploča i nosači", "Worktop and supports")),
        ("plinth",       _t("Sokla", "Plinth"),      _t("Sokla i lajsne", "Plinths and trims")),
    ]
    for key, sheet_name, long_name in SHEET_CFG:
        ws = wb.create_sheet(sheet_name)
        _write_sheet(ws, _translate_export_df(sections.get(key), _lang), SEC_COLS, f"{title}  |  {long_name}")

    ws_hw = wb.create_sheet(_t("Okovi", "Hardware"))
    _write_sheet(ws_hw, _translate_export_df(sections.get("hardware"), _lang), HARDWARE_COLS, f"{title}  |  {_t('Okovi i potrošni materijal', 'Hardware and consumables')}")

    ws_proj = wb.create_sheet(_t("Radni nalog", "Work order"))
    _write_sheet(ws_proj, _translate_export_df(service_packet.get("project_header"), _lang), PROJECT_COLS, f"{title}  |  {_t('Radni nalog', 'Work order')}")

    ws_sc = wb.create_sheet(_t("Servis sečenje", "Workshop cutting"))
    _write_sheet(ws_sc, _translate_export_df(service_packet.get("service_cuts"), _lang), SERVICE_CUTS_COLS, f"{title}  |  {_t('Servis paket - sečenje', 'Workshop packet - cutting')}")

    ws_se = wb.create_sheet(_t("Servis kant", "Workshop edging"))
    _write_sheet(ws_se, _translate_export_df(service_packet.get("service_edge"), _lang), SERVICE_EDGE_COLS, f"{title}  |  {_t('Servis paket - kantovanje', 'Workshop packet - edging')}")

    _svc_proc_df = _translate_export_df(service_packet.get("service_processing"), _lang)

    ws_sp = wb.create_sheet(_t("Servis obrada", "Workshop machining"))
    _write_sheet(ws_sp, _svc_proc_df, SERVICE_PROC_COLS, f"{title}  |  {_t('Servis paket - obrada', 'Workshop packet - machining')}")

    ws_wt = wb.create_sheet(_t("Spec ploce", "Worktop spec"))
    _write_sheet(
        ws_wt,
        _extract_worktop_spec_df(_svc_proc_df),
        [
            ("Zid", _t("Zid", "Wall"), 10),
            ("Modul", _t("Modul", "Module"), 22),
            ("Kol.", _t("Kol.", "Qty"), 8),
            ("Tip obrade", _t("Tip obrade", "Processing type"), 18),
            ("Izvodi", _t("Izvodi", "Operations"), 18),
            ("Osnov izvođenja", _t("Osnov izvođenja", "Execution basis"), 18),
            ("Obrada / napomena", _t("Obrada / napomena", "Processing / note"), 48),
        ],
        f"{title}  |  {_t('Specifikacija radne ploče', 'WORKTOP SPECIFICATION')}",
    )

    ws_si = wb.create_sheet(_t("Servis uputstvo", "Workshop instructions"))
    _write_sheet(ws_si, _translate_export_df(service_packet.get("service_instructions"), _lang), SERVICE_INSTR_COLS, f"{title}  |  {_t('Instrukcije za servis', 'Workshop instructions')}")

    ws_shop = wb.create_sheet(_t("Lista kupovine", "Shopping"))
    _write_sheet(ws_shop, _translate_export_df(service_packet.get("shopping_list"), _lang), SHOP_COLS, f"{title}  |  {_t('Lista kupovine', 'Shopping list')}")

    ws_ready = wb.create_sheet(_t("Kupuje se", "Ready-made"))
    _write_sheet(ws_ready, _translate_export_df(service_packet.get("ready_made_items"), _lang), SHOP_COLS, f"{title}  |  {_t('Ovo se kupuje gotovo', 'Purchased ready-made')}")

    ws_guide = wb.create_sheet(_t("Vodič", "Guide"))
    _write_sheet(ws_guide, _translate_export_df(service_packet.get("user_guide"), _lang), GUIDE_COLS, f"{title}  |  {_t('Kratak vodič za korisnika', 'Quick user guide')}")

    ws_wchk = wb.create_sheet(_t("Checklist servis", "Workshop checklist"))
    _write_sheet(ws_wchk, _translate_export_df(service_packet.get("workshop_checklist"), _lang), CHECK_COLS, f"{title}  |  {_t('Kontrolna lista pre servisa', 'Checklist before workshop')}")

    ws_hchk = wb.create_sheet(_t("Checklist dom", "Home checklist"))
    _write_sheet(ws_hchk, _translate_export_df(service_packet.get("home_checklist"), _lang), CHECK_COLS, f"{title}  |  {_t('Kontrolna lista pre sklapanja', 'Checklist before assembly')}")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_cutlist_csv(kitchen: Dict[str, Any], lang: str = "sr") -> bytes:
    """Generise glavni CSV eksport iz istog finalnog dataseta kao PDF i XLSX."""
    import csv
    from io import StringIO

    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)
    final_ds = get_final_cutlist_dataset(kitchen, lang=_lang)
    summary = final_ds.get("summary", {}) or {}
    df = summary.get("summary_detaljna", pd.DataFrame())
    if df is None:
        df = pd.DataFrame()

    edge_thk = float(((kitchen.get("materials", {}) or {}).get("edge_abs_thk", 2.0)) or 2.0)

    out = StringIO()
    wr = csv.writer(out, delimiter=",", lineterminator="\n")
    wr.writerow([
        "module",
        "part_name",
        "material",
        "length_mm",
        "width_mm",
        "qty",
        "edge_front_mm",
        "edge_back_mm",
        "edge_left_mm",
        "edge_right_mm",
        "notes",
    ])

    if not df.empty:
        sort_cols = [c for c in ("Modul", "Deo", "PartCode") if c in df.columns]
        if sort_cols:
            df = df.sort_values(sort_cols).reset_index(drop=True)
        for _, row in df.iterrows():
            wr.writerow([
                _sanitize_export_value(row.get("Modul", "")),
                _sanitize_export_value(row.get("Deo", "")),
                _sanitize_export_value(row.get("Materijal", "")),
                _sanitize_export_value(row.get("Dužina [mm]", "")),
                _sanitize_export_value(row.get("Širina [mm]", "")),
                _sanitize_export_value(row.get("Kom", row.get("Kol.", ""))),
                edge_thk if bool(row.get("L1", False)) else "",
                edge_thk if bool(row.get("L2", False)) else "",
                edge_thk if bool(row.get("K1", False)) else "",
                edge_thk if bool(row.get("K2", False)) else "",
                _sanitize_export_value(row.get("Napomena", "")),
            ])

    return out.getvalue().encode("utf-8-sig")


# =========================================================
# Sumarna krojna lista — agregirane ploce po dimenziji
# =========================================================

# Redosled tipova ploča u objedinjenoj sumarnoj listi (identično PDF formatu)
_TIP_ORDER: Dict[str, int] = {
    "Korpus":      0,
    "Leđna ploča": 1,
    "Front":       2,
    "Radna ploča": 3,
}

def generate_cutlist_summary(sections: Dict[str, pd.DataFrame], lang: str = "sr") -> Dict[str, pd.DataFrame]:
    """
    Vraca sumarnu krojnu listu po PDF formatu:
      - summary_all: JEDNA objedinjena tabela svih ploča (identično strana 4 PDF-a)
        Kolone: Deo | Kom | Dužina [mm] | Širina [mm] | Deb. [mm] | Materijal | Orijent. | Kant L1 | Kant L2 | Kant K1 | Kant K2
        Dimenzije su FIN (gotove mere posle kanta), sortirano po Deo

      - summary_detaljna: detaljna lista sa Modul kolonom (identično strane 5-7 PDF-a)
        Kolone: Modul | Deo | Kom | Dužina [mm] | Širina [mm] | Deb. [mm] | Materijal | Orijent. | ...

      - summary_carcass / summary_fronts / summary_backs: backwards compat
    """
    result: Dict[str, pd.DataFrame] = {
        "summary_all":     pd.DataFrame(),
        "summary_detaljna": pd.DataFrame(),
        "summary_carcass": pd.DataFrame(),
        "summary_fronts":  pd.DataFrame(),
        "summary_backs":   pd.DataFrame(),
    }

    # -------------------------------------------------------
    # summary_all — OBJEDINJENA tabela svih ploča (PDF strana 4)
    # Grupiše iste ploče (isti Deo, iste FIN dimenzije, isti materijal)
    # -------------------------------------------------------
    all_frames = []
    _sec_to_tip = [
        ("carcass", "Korpus"),
        ("backs",   "Leđna ploča"),
        ("fronts",  "Front"),
        ("drawer_boxes", "Sanduk fioke"),
        ("worktop", "Radna ploča"),
    ]
    _needed_all = {"Deo", "Materijal", "Deb.", "Dužina [mm]", "Širina [mm]", "Kol."}

    for sec_key, tip_label in _sec_to_tip:
        df = sections.get(sec_key)
        if df is None or df.empty:
            continue
        if not _needed_all.issubset(set(df.columns)):
            continue
        df2 = df.copy()
        df2["_Tip"] = tip_label
        all_frames.append(df2)

    if all_frames:
        combined = pd.concat(all_frames, ignore_index=True)
        # Grupisanje: Deo + dimenzije + materijal + kant flags (FIN dimenzije)
        grp_cols = ["_Tip", "Deo", "Materijal", "Deb.", "Dužina [mm]", "Širina [mm]",
                    "Orijentacija", "Smer goda", "L1", "L2", "K1", "K2"]
        grp_cols = [c for c in grp_cols if c in combined.columns]
        # VAŽNO: popuniti NaN u grp_cols pre groupby!
        # Sekcije poput "backs" i "worktop" nemaju kolone "Smer goda" —
        # pd.concat ih popunjava NaN-om, a groupby() po defaultu izbacuje
        # redove sa NaN ključem, što uzrokuje nestajanje tih redova iz sumarnog prikaza.
        for _gc in grp_cols:
            if combined[_gc].dtype == object or combined[_gc].isna().any():
                combined[_gc] = combined[_gc].fillna("").astype(str)
        try:
            agg_d = {"Kol.": "sum"}
            if "Kant" in combined.columns:
                agg_d["Kant"] = "first"
            agg = (
                combined.groupby(grp_cols, as_index=False)
                .agg(agg_d)
            )
            # Sortiranje: Tip order → Materijal → Duzina desc
            agg["_sort"] = agg["_Tip"].map(_TIP_ORDER).fillna(99)
            agg = agg.sort_values(["_sort", "Materijal", "Dužina [mm]"],
                                  ascending=[True, True, False])
            agg = agg.drop(columns=["_sort", "_Tip"]).reset_index(drop=True)
            # Preimenovati Kol. → Kom (kao u PDF-u)
            agg = agg.rename(columns={"Kol.": "Kom"})
            # Dodati redni broj
            agg.insert(0, "RB", range(1, len(agg) + 1))
            # Uредiti kolone po PDF redosledu
            _cols_order = ["RB", "Deo", "Kom", "Dužina [mm]", "Širina [mm]", "Deb.",
                           "Materijal", "Smer goda", "Orijentacija", "L1", "L2", "K1", "K2", "Kant"]
            _cols_show = [c for c in _cols_order if c in agg.columns]
            result["summary_all"] = _translate_export_df(agg[_cols_show], lang)
        except Exception as ex:
            _LOG.debug("Failed summary_all aggregation, using combined fallback: %s", ex)
            result["summary_all"] = _translate_export_df(combined.copy(), lang)

    # -------------------------------------------------------
    # summary_detaljna — detaljna lista sa Modul kolonom (PDF strane 5-7)
    # -------------------------------------------------------
    det_frames = []
    for sec_key, _ in _sec_to_tip:
        df = sections.get(sec_key)
        if df is None or df.empty:
            continue
        det_frames.append(df.copy())

    if det_frames:
        det = pd.concat(det_frames, ignore_index=True)
        _det_needed = {"ID", "Modul", "Deo", "Kol.", "Dužina [mm]", "Širina [mm]", "Deb.", "Materijal"}
        if _det_needed.issubset(set(det.columns)):
            det = det.rename(columns={"Kol.": "Kom"})
            # Compatibility: Cut List tab and PDF "by unit" views still expect ID and Kol.
            det["Kol."] = det["Kom"]
            _sort_cols = [c for c in ("ID", "Modul", "Deo") if c in det.columns]
            det = det.sort_values(_sort_cols).reset_index(drop=True)
            _det_cols = ["ID", "PartCode", "Modul", "Deo", "Pozicija", "SklopKorak", "Kom", "Kol.", "Dužina [mm]", "Širina [mm]", "Deb.",
                         "Materijal", "Smer goda", "Orijentacija", "L1", "L2", "K1", "K2", "Kant", "Napomena"]
            _det_cols = [c for c in _det_cols if c in det.columns]
            result["summary_detaljna"] = _translate_export_df(det[_det_cols], lang)
        else:
            result["summary_detaljna"] = _translate_export_df(det.copy(), lang)

    # -------------------------------------------------------
    # Backwards compat — odvojene sekcije
    # -------------------------------------------------------
    for sec_key, out_key in (
        ("carcass", "summary_carcass"),
        ("fronts",  "summary_fronts"),
        ("backs",   "summary_backs"),
    ):
        df = sections.get(sec_key)
        if df is None or df.empty:
            continue
        needed = {"Materijal", "Deb.", "Dužina [mm]", "Širina [mm]", "Kol."}
        if not needed.issubset(set(df.columns)):
            result[out_key] = _translate_export_df(df.copy(), lang)
            continue
        try:
            agg2 = (
                df.groupby(["Materijal", "Deb.", "Dužina [mm]", "Širina [mm]"], as_index=False)
                .agg({"Kol.": "sum"})
                .sort_values(["Materijal", "Dužina [mm]"])
                .reset_index(drop=True)
            )
            agg2.insert(0, "RB", range(1, len(agg2) + 1))
            result[out_key] = _translate_export_df(agg2, lang)
        except Exception as ex:
            _LOG.debug("Failed %s aggregation, using section fallback: %s", out_key, ex)
            result[out_key] = _translate_export_df(df.copy(), lang)

    return result


# =========================================================
# Lista ugradjenih uredjaja (frizider, rerna, ploca, napa...)
# =========================================================
_APPLIANCE_KEYWORDS: Dict[str, str] = {
    "FRIDGE_FREEZE": "Frizider sa zamrzivačem",
    "FRIDGE":        "Frizider",
    "OVEN_HOB":      "Rerna + ploča za kuvanje (kombo)",
    "OVEN_MICRO":    "Rerna + mikrotalasna (kombo)",
    "HOB":           "Ploča za kuvanje",
    "OVEN":          "Ugradna rerna",
    "MICRO":         "Mikrotalasna pećnica",
    "DISHWASHER":    "Mašina za sudove",
    "HOOD":          "Aspirator / napa",
    "SINK":          "Sudopera",
}


def generate_appliance_list(kitchen: Dict[str, Any], lang: str = "sr") -> pd.DataFrame:
    """
    Vraca DataFrame sa listom ugradnih uredjaja koji se NABAVLJAJU
    (ne seku se iz ploce — frizider, rerna, ploca za kuvanje, mašina itd.)
    """
    modules = kitchen.get("modules", []) or []
    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)
    appliance_map = {
        "Frizider sa zamrzivačem": "Fridge with freezer",
        "Frizider": "Fridge",
        "Rerna + ploča za kuvanje (kombo)": "Oven + hob (combo)",
        "Rerna + mikrotalasna (kombo)": "Oven + microwave (combo)",
        "Ploča za kuvanje": "Hob",
        "Ugradna rerna": "Built-in oven",
        "Mikrotalasna pećnica": "Microwave oven",
        "Mašina za sudove": "Dishwasher",
        "Aspirator / napa": "Cooker hood",
        "Sudopera": "Sink",
    }
    rows: List[Dict[str, Any]] = []

    for m in modules:
        tid = str(m.get("template_id", "")).upper()
        for kw, naziv in _APPLIANCE_KEYWORDS.items():
            if kw in tid:
                rows.append({
                    "ID":           m.get("id", "?"),
                    "Uređaj":       appliance_map.get(naziv, naziv) if _lang == "en" else naziv,
                    "Oznaka":       m.get("label", tid),
                    "Zona":         _t(str(m.get("zone", "")).capitalize(), str(m.get("zone", "")).capitalize()),
                    "Š [mm]":       m.get("w_mm", ""),
                    "V [mm]":       m.get("h_mm", ""),
                    "D [mm]":       m.get("d_mm", ""),
                    "Kol.":         1,
                    "Napomena":     "",
                })
                break  # svaki modul samo jednom

    return pd.DataFrame(rows)



def generate_wardrobe_sections_csv(kitchen: Dict[str, Any], lang: str = "sr") -> bytes:
    """Generise CSV za americke sekcije ormara (leva/centar/desna/spremnik)."""
    import csv
    from io import StringIO
    _lang = str(lang or "sr").lower().strip()
    def _t(sr: str, en: str) -> str:
        return _pdf_t(sr, en, _lang)

    mats = kitchen.get("materials", {}) or {}
    carcass_mat = str(mats.get("carcass_material", _t("Iverica", "Chipboard")))
    front_mat = str(mats.get("front_material", carcass_mat))
    dflt_thk = float(mats.get("carcass_thk", 18) or 18)

    header = [
        "module_id", "module_label", "section", "component", "qty",
        "width_mm", "height_mm", "depth_mm", "material", "thickness_mm", "status", "note",
    ]
    out = StringIO()
    wr = csv.writer(out)
    wr.writerow(header)

    def _safe_int(v: Any, default: int = 0) -> int:
        try:
            return int(float(v))
        except Exception:
            return int(default)

    for m in (kitchen.get("modules", []) or []):
        tid = str(m.get("template_id", "")).upper()
        p = m.get("params", {}) or {}
        is_wardrobe = bool(p.get("wardrobe", False)) or ("WARDROBE" in tid)
        if not is_wardrobe:
            continue

        mid = _safe_int(m.get("id", 0), 0)
        mlbl = str(m.get("label", tid))
        mw = _safe_int(m.get("w_mm", 0), 0)
        mh = _safe_int(m.get("h_mm", 0), 0)
        md = _safe_int(m.get("d_mm", 0), 0)
        if mw <= 0 or mh <= 0 or md <= 0:
            wr.writerow([mid, mlbl, "-", "validation", 0, mw, mh, md, "-", dflt_thk, "ERROR", _t("Nevalidne dimenzije modula", "Invalid module dimensions")])
            continue

        sec = p.get("american_sections") if isinstance(p.get("american_sections"), dict) else {}
        if sec:
            lp = max(1, _safe_int(sec.get("left_pct", 33), 33))
            cp = max(1, _safe_int(sec.get("center_pct", 34), 34))
            rp = max(1, _safe_int(sec.get("right_pct", 33), 33))
            s = lp + cp + rp
            top_h = max(80, min(_safe_int(sec.get("top_h_mm", 420), 420), int(mh * 0.45)))
            base_h = max(80, mh - top_h)
            lw = int(mw * lp / s)
            cw = int(mw * cp / s)
            rw = max(40, mw - lw - cw)

            parts = [
                ("left", lw, sec.get("left", {}) or {}),
                ("center", cw, sec.get("center", {}) or {}),
                ("right", rw, sec.get("right", {}) or {}),
                ("top", mw, sec.get("top", {}) or {}),
            ]
            for sname, sw, sp in parts:
                sh = top_h if sname == "top" else base_h
                ns = max(0, _safe_int(sp.get("shelves", 0), 0))
                nd = max(0, _safe_int(sp.get("drawers", 0), 0))
                nh = max(0, _safe_int(sp.get("hangers", 0), 0))

                if ns > 0:
                    wr.writerow([mid, mlbl, sname, "shelf", ns, max(40, sw - 16), 18, max(120, md - 20), carcass_mat, dflt_thk, "OK", _t("Polica", "Shelf")])
                if nd > 0:
                    dh = max(80, int((sh * 0.30) / max(1, nd)))
                    wr.writerow([mid, mlbl, sname, "drawer_front", nd, max(40, sw - 16), dh, 18, front_mat, dflt_thk, "OK", _t("Fioka front", "Drawer front")])
                    wr.writerow([mid, mlbl, sname, "drawer_box", nd, max(40, sw - 24), max(60, dh - 12), max(120, md - 40), carcass_mat, dflt_thk, "OK", _t("Fioka sanduk", "Drawer box")])
                if nh > 0:
                    wr.writerow([mid, mlbl, sname, "hanger_rod", nh, max(60, sw - 20), 20, 20, _t("Metal", "Metal"), 20, "OK", _t("Sipka za vesanje", "Hanging rail")])

                used_h = (18 * max(0, ns)) + (max(80, int((sh * 0.30) / max(1, nd))) * max(0, nd))
                if used_h > sh:
                    wr.writerow([mid, mlbl, sname, "validation", 0, sw, sh, md, "-", dflt_thk, "WARN", _t("Police/Fioke prekoracile visinu sekcije", "Shelves/drawers exceeded the section height")])
        else:
            ns = max(0, _safe_int(p.get("n_shelves", 4), 4))
            nd = max(0, _safe_int(p.get("n_drawers", 2), 2))
            nh = max(0, _safe_int(p.get("hanger_sections", 1), 1))
            if ns > 0:
                wr.writerow([mid, mlbl, "generic", "shelf", ns, max(40, mw - 16), 18, max(120, md - 20), carcass_mat, dflt_thk, "OK", _t("Polica", "Shelf")])
            if nd > 0:
                dh = max(80, int((mh * 0.22) / max(1, nd)))
                wr.writerow([mid, mlbl, "generic", "drawer_front", nd, max(40, mw - 16), dh, 18, front_mat, dflt_thk, "OK", _t("Fioka front", "Drawer front")])
            if nh > 0:
                wr.writerow([mid, mlbl, "generic", "hanger_rod", nh, max(60, mw - 20), 20, 20, _t("Metal", "Metal"), 20, "OK", _t("Sipka za vesanje", "Hanging rail")])

    return out.getvalue().encode("utf-8-sig")
