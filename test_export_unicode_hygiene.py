# -*- coding: utf-8 -*-
from __future__ import annotations

from io import BytesIO

import openpyxl
from pypdf import PdfReader

from cutlist import generate_cutlist_excel, generate_cutlist_pdf
from test_export_consistency import build_sample_kitchen


def run_export_unicode_hygiene_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    pdf_text = "\n".join(
        (page.extract_text() or "")
        for page in PdfReader(BytesIO(generate_cutlist_pdf(kitchen, lang="sr"))).pages[:4]
    )

    wb = openpyxl.load_workbook(BytesIO(generate_cutlist_excel(kitchen, lang="sr")), data_only=True)
    ws = wb["Kako koristiti"]
    xlsx_text_parts: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        xlsx_text_parts.extend("" if value is None else str(value) for value in row)
    xlsx_text = " | ".join(xlsx_text_parts)

    bad_markers = ["Ã", "â", "Ä", "Å"]
    for marker in bad_markers:
        if marker in pdf_text:
            return False, f"pdf_unicode_marker_present:{marker}"
        if marker in xlsx_text:
            return False, f"xlsx_unicode_marker_present:{marker}"
    return True, "ok"
