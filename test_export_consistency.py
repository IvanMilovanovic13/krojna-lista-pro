# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
from io import BytesIO, StringIO

import openpyxl
import pandas as pd
from pypdf import PdfReader

from cutlist import (
    _format_pdf_table_cell,
    _manufacturing_warnings,
    _sanitize_export_df,
    _translate_export_text,
    build_cutlist_pdf_bytes,
    generate_cutlist,
    generate_cutlist_csv,
    generate_cutlist_excel,
    generate_cutlist_pdf,
    get_final_cutlist_dataset,
)
from state_logic import _default_kitchen
from ui_cutlist_tab import _friendly_part_name as cutlist_friendly_part_name
from ui_pdf_export import _friendly_part_name as pdf_friendly_part_name


def build_sample_kitchen() -> dict:
    k = _default_kitchen()
    k["wall"]["length_mm"] = 3000
    k["wall"]["height_mm"] = 2600
    k["foot_height_mm"] = 150
    k["worktop"] = {
        "enabled": True,
        "material": "Iverica",
        "thickness": 38,
        "depth_mm": 600,
    }
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {
            "id": 1,
            "template_id": "BASE_2DOOR",
            "zone": "base",
            "x_mm": 0,
            "w_mm": 800,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Donji 2V",
            "params": {},
            "wall_key": "A",
        },
        {
            "id": 2,
            "template_id": "SINK_BASE",
            "zone": "base",
            "x_mm": 800,
            "w_mm": 800,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Sudopera",
            "params": {
                "sink_cutout_x_mm": 120,
                "sink_cutout_width_mm": 500,
                "sink_cutout_depth_mm": 480,
            },
            "wall_key": "A",
        },
        {
            "id": 3,
            "template_id": "BASE_COOKING_UNIT",
            "zone": "base",
            "x_mm": 1600,
            "w_mm": 600,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Rerna+ploca",
            "params": {
                "hob_cutout_x_mm": 40,
                "hob_cutout_width_mm": 510,
                "hob_cutout_depth_mm": 470,
            },
            "wall_key": "A",
        },
        {
            "id": 4,
            "template_id": "WALL_2DOOR",
            "zone": "wall",
            "x_mm": 0,
            "w_mm": 800,
            "h_mm": 720,
            "d_mm": 320,
            "label": "Gornji 2V",
            "params": {},
            "wall_key": "A",
        },
    ]
    return k


def build_reference_linear_kitchen() -> dict:
    k = _default_kitchen()
    k["wall"]["length_mm"] = 3600
    k["wall"]["height_mm"] = 2600
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_1DOOR", "zone": "base", "x_mm": 0, "w_mm": 400, "h_mm": 720, "d_mm": 560, "label": "B400", "params": {}, "wall_key": "A"},
        {"id": 2, "template_id": "BASE_DISHWASHER", "zone": "base", "x_mm": 400, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "MZS", "params": {}, "wall_key": "A"},
        {"id": 3, "template_id": "SINK_BASE", "zone": "base", "x_mm": 1000, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Sudopera", "params": {"sink_cutout_x_mm": 120, "sink_cutout_width_mm": 500, "sink_cutout_depth_mm": 480}, "wall_key": "A"},
        {"id": 4, "template_id": "BASE_COOKING_UNIT", "zone": "base", "x_mm": 1800, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "Rerna", "params": {"hob_cutout_x_mm": 40, "hob_cutout_width_mm": 510, "hob_cutout_depth_mm": 470}, "wall_key": "A"},
        {"id": 5, "template_id": "BASE_DRAWERS_3", "zone": "base", "x_mm": 2400, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Fioke", "params": {"n_drawers": 3}, "wall_key": "A"},
        {"id": 6, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 0, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G800", "params": {}, "wall_key": "A"},
        {"id": 7, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 800, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G801", "params": {}, "wall_key": "A"},
        {"id": 8, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 1800, "w_mm": 600, "h_mm": 720, "d_mm": 320, "label": "G602", "params": {}, "wall_key": "A"},
        {"id": 9, "template_id": "WALL_1DOOR", "zone": "wall", "x_mm": 2400, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G803", "params": {}, "wall_key": "A"},
    ]
    return k


def build_reference_tall_block_kitchen() -> dict:
    k = _default_kitchen()
    k["wall"]["length_mm"] = 3000
    k["wall"]["height_mm"] = 2600
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "TALL_FRIDGE_FREEZER", "zone": "tall", "x_mm": 0, "w_mm": 600, "h_mm": 2100, "d_mm": 560, "label": "Frižider", "params": {}, "wall_key": "A"},
        {"id": 2, "template_id": "TALL_OVEN_MICRO", "zone": "tall", "x_mm": 600, "w_mm": 600, "h_mm": 2100, "d_mm": 560, "label": "Rerna mikro", "params": {}, "wall_key": "A"},
        {"id": 3, "template_id": "BASE_DRAWERS_3", "zone": "base", "x_mm": 1200, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Fioke", "params": {"n_drawers": 3}, "wall_key": "A"},
        {"id": 4, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 2000, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B800", "params": {"n_shelves": 1}, "wall_key": "A"},
        {"id": 5, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 1200, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G800", "params": {}, "wall_key": "A"},
        {"id": 6, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 2000, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G801", "params": {}, "wall_key": "A"},
    ]
    return k


def build_reference_utility_kitchen() -> dict:
    k = _default_kitchen()
    k["wall"]["length_mm"] = 3200
    k["wall"]["height_mm"] = 2600
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_TRASH", "zone": "base", "x_mm": 0, "w_mm": 400, "h_mm": 720, "d_mm": 560, "label": "Sortirnik", "params": {}, "wall_key": "A"},
        {"id": 2, "template_id": "BASE_CORNER", "zone": "base", "x_mm": 400, "w_mm": 900, "h_mm": 720, "d_mm": 900, "label": "Ugao", "params": {}, "wall_key": "A"},
        {"id": 3, "template_id": "BASE_HOB", "zone": "base", "x_mm": 1300, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "Ploča", "params": {"hob_cutout_x_mm": 45, "hob_cutout_width_mm": 500, "hob_cutout_depth_mm": 470}, "wall_key": "A"},
        {"id": 4, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 1900, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B801", "params": {"n_shelves": 1}, "wall_key": "A"},
        {"id": 5, "template_id": "FILLER_PANEL", "zone": "base", "x_mm": 2700, "w_mm": 80, "h_mm": 720, "d_mm": 560, "label": "Filer", "params": {}, "wall_key": "A"},
        {"id": 6, "template_id": "END_PANEL", "zone": "base", "x_mm": 2780, "w_mm": 18, "h_mm": 720, "d_mm": 560, "label": "Bočna", "params": {}, "wall_key": "A"},
        {"id": 7, "template_id": "WALL_1DOOR", "zone": "wall", "x_mm": 1300, "w_mm": 600, "h_mm": 720, "d_mm": 320, "label": "G600", "params": {}, "wall_key": "A"},
        {"id": 8, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 1900, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G801", "params": {}, "wall_key": "A"},
    ]
    return k


def build_reference_raised_dishwasher_kitchen() -> dict:
    k = _default_kitchen()
    k["wall"]["length_mm"] = 3200
    k["wall"]["height_mm"] = 2700
    k["foot_height_mm"] = 150
    k["base_korpus_h_mm"] = 812
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 0, "w_mm": 800, "h_mm": 812, "d_mm": 560, "label": "B800", "params": {"n_shelves": 1}, "wall_key": "A"},
        {"id": 2, "template_id": "BASE_DISHWASHER", "zone": "base", "x_mm": 800, "w_mm": 600, "h_mm": 812, "d_mm": 560, "label": "MZS", "params": {}, "wall_key": "A"},
        {"id": 3, "template_id": "SINK_BASE", "zone": "base", "x_mm": 1400, "w_mm": 800, "h_mm": 812, "d_mm": 560, "label": "Sudopera", "params": {"sink_cutout_x_mm": 120, "sink_cutout_width_mm": 500, "sink_cutout_depth_mm": 480}, "wall_key": "A"},
        {"id": 4, "template_id": "BASE_DRAWERS_3", "zone": "base", "x_mm": 2200, "w_mm": 800, "h_mm": 812, "d_mm": 560, "label": "Fioke", "params": {"n_drawers": 3}, "wall_key": "A"},
        {"id": 5, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 0, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G800", "params": {}, "wall_key": "A"},
        {"id": 6, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 1400, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G801", "params": {}, "wall_key": "A"},
    ]
    return k


def build_reference_l_kitchen() -> dict:
    k = _default_kitchen()
    k["layout"] = "l_oblik"
    k["wall"]["length_mm"] = 3000
    k["wall"]["height_mm"] = 2600
    k["wall_lengths_mm"] = {"A": 3000, "B": 2600}
    k["wall_heights_mm"] = {"A": 2600, "B": 2600}
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_CORNER", "zone": "base", "x_mm": 2100, "w_mm": 900, "h_mm": 720, "d_mm": 900, "label": "Ugao A", "params": {}, "wall_key": "A"},
        {"id": 2, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 1300, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B800 A", "params": {"n_shelves": 1}, "wall_key": "A"},
        {"id": 3, "template_id": "BASE_1DOOR", "zone": "base", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "B600 B", "params": {}, "wall_key": "B"},
        {"id": 4, "template_id": "SINK_BASE", "zone": "base", "x_mm": 600, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Sudopera B", "params": {"sink_cutout_x_mm": 120, "sink_cutout_width_mm": 500, "sink_cutout_depth_mm": 480}, "wall_key": "B"},
        {"id": 5, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 1300, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G800 A", "params": {}, "wall_key": "A"},
        {"id": 6, "template_id": "WALL_1DOOR", "zone": "wall", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 320, "label": "G600 B", "params": {}, "wall_key": "B"},
    ]
    return k


def build_reference_galley_kitchen() -> dict:
    k = _default_kitchen()
    k["layout"] = "galija"
    k["wall"]["length_mm"] = 3400
    k["wall"]["height_mm"] = 2600
    k["wall_lengths_mm"] = {"A": 3400, "B": 3400}
    k["wall_heights_mm"] = {"A": 2600, "B": 2600}
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 0, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B800 A", "params": {"n_shelves": 1}, "wall_key": "A"},
        {"id": 2, "template_id": "BASE_DISHWASHER", "zone": "base", "x_mm": 800, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "MZS A", "params": {}, "wall_key": "A"},
        {"id": 3, "template_id": "BASE_DRAWERS_3", "zone": "base", "x_mm": 1400, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Fioke A", "params": {"n_drawers": 3}, "wall_key": "A"},
        {"id": 4, "template_id": "SINK_BASE", "zone": "base", "x_mm": 0, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Sudopera B", "params": {"sink_cutout_x_mm": 120, "sink_cutout_width_mm": 500, "sink_cutout_depth_mm": 480}, "wall_key": "B"},
        {"id": 5, "template_id": "BASE_COOKING_UNIT", "zone": "base", "x_mm": 800, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "Rerna B", "params": {"hob_cutout_x_mm": 40, "hob_cutout_width_mm": 510, "hob_cutout_depth_mm": 470}, "wall_key": "B"},
        {"id": 6, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 1400, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B800 B", "params": {"n_shelves": 1}, "wall_key": "B"},
    ]
    return k


def build_reference_u_kitchen() -> dict:
    k = _default_kitchen()
    k["layout"] = "u_oblik"
    k["wall"]["length_mm"] = 3200
    k["wall"]["height_mm"] = 2600
    k["wall_lengths_mm"] = {"A": 3200, "B": 2400, "C": 2400}
    k["wall_heights_mm"] = {"A": 2600, "B": 2600, "C": 2600}
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 38, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_CORNER", "zone": "base", "x_mm": 2300, "w_mm": 900, "h_mm": 720, "d_mm": 900, "label": "Ugao A desno", "params": {}, "wall_key": "A"},
        {"id": 2, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 1500, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B800 A", "params": {"n_shelves": 1}, "wall_key": "A"},
        {"id": 3, "template_id": "BASE_1DOOR", "zone": "base", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "B600 B", "params": {}, "wall_key": "B"},
        {"id": 4, "template_id": "SINK_BASE", "zone": "base", "x_mm": 600, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "Sudopera B", "params": {"sink_cutout_x_mm": 120, "sink_cutout_width_mm": 500, "sink_cutout_depth_mm": 480}, "wall_key": "B"},
        {"id": 5, "template_id": "BASE_COOKING_UNIT", "zone": "base", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "Rerna C", "params": {"hob_cutout_x_mm": 40, "hob_cutout_width_mm": 510, "hob_cutout_depth_mm": 470}, "wall_key": "C"},
        {"id": 6, "template_id": "BASE_2DOOR", "zone": "base", "x_mm": 600, "w_mm": 800, "h_mm": 720, "d_mm": 560, "label": "B800 C", "params": {"n_shelves": 1}, "wall_key": "C"},
        {"id": 7, "template_id": "WALL_2DOOR", "zone": "wall", "x_mm": 1500, "w_mm": 800, "h_mm": 720, "d_mm": 320, "label": "G800 A", "params": {}, "wall_key": "A"},
        {"id": 8, "template_id": "WALL_1DOOR", "zone": "wall", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 320, "label": "G600 B", "params": {}, "wall_key": "B"},
        {"id": 9, "template_id": "WALL_1DOOR", "zone": "wall", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 320, "label": "G600 C", "params": {}, "wall_key": "C"},
    ]
    return k


def run_export_consistency_check(lang: str = "en") -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    final_ds = get_final_cutlist_dataset(kitchen, lang=lang)
    detail_df = final_ds["summary"].get("summary_detaljna")
    expected_rows = 0 if detail_df is None else len(detail_df)

    csv_bytes = generate_cutlist_csv(kitchen, lang=lang)
    csv_text = csv_bytes.decode("utf-8-sig")
    csv_rows = list(csv.DictReader(StringIO(csv_text)))
    if len(csv_rows) != expected_rows:
        return False, f"csv_rows={len(csv_rows)} expected={expected_rows}"

    if any(str(row.get("part_name", "")).strip() == "" for row in csv_rows):
        return False, "csv_blank_part_name"

    if any(str(row.get("material", "")).strip().lower() in {"nan", "none", "null"} for row in csv_rows):
        return False, "csv_unsanitized_material"

    pdf_bytes = build_cutlist_pdf_bytes(kitchen, final_ds["sections"], project_title="Export consistency", lang=lang)
    if len(pdf_bytes) < 5000:
        return False, f"pdf_too_small={len(pdf_bytes)}"

    xlsx_bytes = generate_cutlist_excel(kitchen, lang=lang)
    if len(xlsx_bytes) < 5000:
        return False, f"xlsx_too_small={len(xlsx_bytes)}"

    return True, f"rows={expected_rows}"


def run_export_sanitization_check() -> tuple[bool, str]:
    df = pd.DataFrame(
        [
            {"Deo": "Valid", "Materijal": "Chipboard", "Duzina [mm]": 500, "Sirina [mm]": 300, "Napomena": None},
            {"Deo": "Zero length", "Materijal": "Chipboard", "Duzina [mm]": 0, "Sirina [mm]": 300, "Napomena": "bad"},
            {"Deo": "", "Materijal": "Chipboard", "Duzina [mm]": 500, "Sirina [mm]": 300, "Napomena": "bad"},
            {"Deo": "NaN row", "Materijal": float("nan"), "Duzina [mm]": 500, "Sirina [mm]": 300, "Napomena": "bad"},
            {"Deo": "None dims", "Materijal": "Chipboard", "Duzina [mm]": None, "Sirina [mm]": 300, "Napomena": "bad"},
        ]
    )
    out = _sanitize_export_df(
        df,
        require_positive_dims=True,
        part_col="Deo",
        material_col="Materijal",
        width_candidates=("Duzina [mm]",),
        height_candidates=("Sirina [mm]",),
    )
    if len(out) != 1:
        return False, f"sanitized_rows={len(out)}"
    row = out.iloc[0].to_dict()
    if row.get("Napomena", "__missing__") != "":
        return False, "note_not_sanitized"
    if row.get("Deo") != "Valid":
        return False, f"wrong_row={row.get('Deo')}"
    return True, "rows=1"


def run_english_summary_translation_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    final_ds = get_final_cutlist_dataset(kitchen, lang="en")
    sum_df = final_ds["summary"].get("summary_all", pd.DataFrame())
    det_df = final_ds["summary"].get("summary_detaljna", pd.DataFrame())
    if sum_df is None or sum_df.empty or det_df is None or det_df.empty:
        return False, "summary_data_missing"

    banned_sum = ("Iverica", "Radna ploča", "Radna ploca")
    for col in ("Materijal", "Deo"):
        if col in sum_df.columns:
            txt = " | ".join(sum_df[col].astype(str).tolist())
            if any(term in txt for term in banned_sum):
                return False, f"summary_sr_leak:{col}"

    banned_det = ("Zid A", "Leva strana", "Desna strana", "Polica (podesiva)")
    for col in ("Zid", "Deo", "Pozicija"):
        if col in det_df.columns:
            txt = " | ".join(det_df[col].astype(str).tolist())
            if any(term in txt for term in banned_det):
                return False, f"detail_sr_leak:{col}"

    return True, "ok"


def run_english_service_translation_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    final_ds = get_final_cutlist_dataset(kitchen, lang="en")
    service_packet = final_ds["service_packet"]

    checks = [
        ("service_cuts", ("Iverica", "Radna ploča", "Radna ploca", "Zid A")),
        ("service_edge", ("Zid A", "Leva strana", "Desna strana")),
        ("service_processing", ("Zid A", "Radna ploča", "Radna ploca")),
        ("shopping_list", ("Sudopera", "Slavina", "Sifon i odvodni set", "Iverica")),
        ("ready_made_items", ("Sudopera", "Slavina", "Ugradna sudopera", "Baterija za sudoperu")),
    ]

    for key, banned_terms in checks:
        df = service_packet.get(key, pd.DataFrame())
        if df is None or df.empty:
            continue
        txt = " | ".join(df.astype(str).fillna("").values.flatten().tolist())
        if any(term in txt for term in banned_terms):
            return False, f"service_sr_leak:{key}"

    return True, "ok"


def run_summary_detail_structure_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    final_ds = get_final_cutlist_dataset(kitchen, lang="sr")
    det_df = final_ds["summary"].get("summary_detaljna", pd.DataFrame())
    if det_df is None or det_df.empty:
        return False, "summary_detaljna_missing"

    required = {"ID", "PartCode", "Kol."}
    missing = [c for c in required if c not in det_df.columns]
    if missing:
        return False, f"missing_cols={','.join(missing)}"

    if det_df["ID"].astype(str).str.strip().eq("").any():
        return False, "blank_id"
    if det_df["PartCode"].astype(str).str.strip().eq("").any():
        return False, "blank_partcode"

    return True, f"rows={len(det_df)}"


def run_validation_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 99,
            "template_id": "MISSING_TEMPLATE_X",
            "zone": "base",
            "x_mm": 2200,
            "w_mm": 600,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Broken",
            "params": {},
            "wall_key": "Z",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (MISSING_TEMPLATE)" not in codes:
        return False, "missing_template_warning_absent"
    return True, f"warnings={len(rows)}"


def run_invalid_dimensions_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["w_mm"] = 0
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (INVALID_DIMENSIONS)" not in codes:
        return False, "invalid_dimensions_warning_absent"
    return True, f"warnings={len(rows)}"


def run_invalid_wall_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["wall_key"] = "Z"
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (INVALID_WALL_REFERENCE)" not in codes:
        return False, "invalid_wall_warning_absent"
    return True, f"warnings={len(rows)}"


def run_missing_template_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "BAD_TEMPLATE_X"
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (MISSING_TEMPLATE)" not in codes:
        return False, "missing_template_warning_absent"
    return True, f"warnings={len(rows)}"


def run_base_alignment_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][1]["h_mm"] = 700
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (BASE_ALIGNMENT_INVALID)" not in codes:
        return False, "base_alignment_warning_absent"
    return True, f"warnings={len(rows)}"


def run_overlap_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 10,
            "template_id": "BASE_1DOOR",
            "zone": "base",
            "x_mm": 500,
            "w_mm": 700,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Overlap test",
            "params": {},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (OVERLAP)" not in codes:
        return False, "overlap_warning_absent"
    return True, f"warnings={len(rows)}"


def run_left_out_of_bounds_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["x_mm"] = -30
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (MODULE_BEFORE_WALL_START)" not in codes:
        return False, "module_before_wall_start_warning_absent"
    return True, f"warnings={len(rows)}"


def run_tall_top_height_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 10,
            "template_id": "TALL_PANTRY",
            "zone": "tall",
            "x_mm": 2200,
            "w_mm": 600,
            "h_mm": 2200,
            "d_mm": 600,
            "label": "Visoki test",
            "params": {},
            "wall_key": "A",
        }
    )
    kitchen["modules"].append(
        {
            "id": 11,
            "template_id": "TALL_TOP_DOORS",
            "zone": "tall_top",
            "x_mm": 2200,
            "w_mm": 600,
            "h_mm": 400,
            "d_mm": 600,
            "label": "Popuna test",
            "params": {},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (TALL_TOP_HEIGHT)" not in codes:
        return False, "tall_top_height_warning_absent"
    return True, f"warnings={len(rows)}"


def run_wall_upper_height_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 10,
            "template_id": "WALL_UPPER_1DOOR",
            "zone": "wall_upper",
            "x_mm": 0,
            "w_mm": 800,
            "h_mm": 1200,
            "d_mm": 320,
            "label": "Drugi red test",
            "params": {},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (WALL_UPPER_HEIGHT)" not in codes:
        return False, "wall_upper_height_warning_absent"
    return True, f"warnings={len(rows)}"


def run_corner_neighbor_warning_check() -> tuple[bool, str]:
    kitchen = build_reference_l_kitchen()
    kitchen["modules"] = [
        m for m in kitchen["modules"]
        if not (int(m.get("id", 0) or 0) == 2 and str(m.get("wall_key", "")).upper() == "A")
    ]
    kitchen["modules"].append(
        {
            "id": 2,
            "template_id": "BASE_1DOOR",
            "zone": "base",
            "x_mm": 1500,
            "w_mm": 600,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Corner neighbor A",
            "params": {"handle_side": "right"},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (CORNER_OPENING_CLEARANCE)" not in codes:
        return False, "corner_opening_clearance_warning_absent"
    if "UPOZORENJE (CORNER_FRONT_COLLISION)" not in codes:
        return False, "corner_front_collision_warning_absent"
    return True, f"warnings={len(rows)}"


def run_corner_door_swing_warning_check() -> tuple[bool, str]:
    kitchen = build_reference_l_kitchen()
    kitchen["modules"] = [
        m for m in kitchen["modules"]
        if not (int(m.get("id", 0) or 0) == 2 and str(m.get("wall_key", "")).upper() == "A")
    ]
    kitchen["modules"].append(
        {
            "id": 2,
            "template_id": "BASE_1DOOR",
            "zone": "base",
            "x_mm": 1650,
            "w_mm": 450,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Swing neighbor A",
            "params": {"handle_side": "left"},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (CORNER_DOOR_SWING)" not in codes:
        return False, "corner_door_swing_warning_absent"
    return True, f"warnings={len(rows)}"


def run_wall_upper_support_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 10,
            "template_id": "WALL_UPPER_1DOOR",
            "zone": "wall_upper",
            "x_mm": 900,
            "w_mm": 800,
            "h_mm": 400,
            "d_mm": 320,
            "label": "Unsupported wall upper",
            "params": {},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (WALL_UPPER_SUPPORT)" not in codes:
        return False, "wall_upper_support_warning_absent"
    return True, f"warnings={len(rows)}"


def run_tall_top_support_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 10,
            "template_id": "TALL_TOP_DOORS",
            "zone": "tall_top",
            "x_mm": 1000,
            "w_mm": 600,
            "h_mm": 300,
            "d_mm": 600,
            "label": "Unsupported tall top",
            "params": {},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (TALL_TOP_SUPPORT)" not in codes:
        return False, "tall_top_support_warning_absent"
    return True, f"warnings={len(rows)}"


def run_side_wall_door_warning_check() -> tuple[bool, str]:
    kitchen_left = build_sample_kitchen()
    kitchen_left["modules"][0]["template_id"] = "BASE_1DOOR"
    kitchen_left["modules"][0]["w_mm"] = 600
    kitchen_left["modules"][0]["x_mm"] = 0
    kitchen_left["modules"][0]["params"] = {"handle_side": "right"}
    rows_left = _manufacturing_warnings(
        kitchen_left["modules"],
        kitchen=kitchen_left,
        wall_h_mm=float(kitchen_left["wall"]["height_mm"]),
        foot_h_mm=float(kitchen_left["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes_left = {str(r.get("Naziv", "")) for r in rows_left}
    if "UPOZORENJE (SIDE_WALL_DOOR)" not in codes_left:
        return False, "side_wall_door_left_warning_absent"

    kitchen_right = build_sample_kitchen()
    kitchen_right["wall"]["length_mm"] = 3000
    kitchen_right["modules"][2]["template_id"] = "BASE_1DOOR"
    kitchen_right["modules"][2]["w_mm"] = 600
    kitchen_right["modules"][2]["x_mm"] = 2400
    kitchen_right["modules"][2]["params"] = {"handle_side": "left"}
    rows_right = _manufacturing_warnings(
        kitchen_right["modules"],
        kitchen=kitchen_right,
        wall_h_mm=float(kitchen_right["wall"]["height_mm"]),
        foot_h_mm=float(kitchen_right["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes_right = {str(r.get("Naziv", "")) for r in rows_right}
    if "UPOZORENJE (SIDE_WALL_DOOR)" not in codes_right:
        return False, "side_wall_door_right_warning_absent"
    return True, f"left={len(rows_left)} right={len(rows_right)}"


def run_module_out_of_bounds_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["wall"]["length_mm"] = 3000
    kitchen["modules"][2]["x_mm"] = 2600
    kitchen["modules"][2]["w_mm"] = 600
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (MODULE_OUT_OF_BOUNDS)" not in codes:
        return False, "module_out_of_bounds_warning_absent"
    return True, f"warnings={len(rows)}"


def run_front_gap_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=0.5,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (FRONT_GAP)" not in codes:
        return False, "front_gap_warning_absent"
    return True, f"warnings={len(rows)}"


def run_single_door_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "BASE_1DOOR"
    kitchen["modules"][0]["w_mm"] = 650
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (SINGLE_DOOR_WIDTH)" not in codes:
        return False, "single_door_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_drawer_stack_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "BASE_DRAWERS_3"
    kitchen["modules"][0]["params"] = {"drawer_heights": [250, 250, 220]}
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (DRAWER_STACK)" not in codes:
        return False, "drawer_stack_warning_absent"
    return True, f"warnings={len(rows)}"


def run_drawer_front_min_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "BASE_DRAWERS_3"
    kitchen["modules"][0]["params"] = {"drawer_heights": [70, 200, 200]}
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (DRAWER_FRONT_MIN)" not in codes:
        return False, "drawer_front_min_warning_absent"
    return True, f"warnings={len(rows)}"


def run_door_drawer_warning_check() -> tuple[bool, str]:
    kitchen_low = build_sample_kitchen()
    kitchen_low["modules"][0]["template_id"] = "BASE_DOOR_DRAWER"
    kitchen_low["modules"][0]["params"] = {"door_height": 150}
    rows_low = _manufacturing_warnings(
        kitchen_low["modules"],
        kitchen=kitchen_low,
        wall_h_mm=float(kitchen_low["wall"]["height_mm"]),
        foot_h_mm=float(kitchen_low["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes_low = {str(r.get("Naziv", "")) for r in rows_low}
    if "UPOZORENJE (DOOR_DRAWER_DOOR_MIN)" not in codes_low:
        return False, "door_drawer_door_min_warning_absent"

    kitchen_high = build_sample_kitchen()
    kitchen_high["modules"][0]["template_id"] = "BASE_DOOR_DRAWER"
    kitchen_high["modules"][0]["params"] = {"door_height": 650}
    rows_high = _manufacturing_warnings(
        kitchen_high["modules"],
        kitchen=kitchen_high,
        wall_h_mm=float(kitchen_high["wall"]["height_mm"]),
        foot_h_mm=float(kitchen_high["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes_high = {str(r.get("Naziv", "")) for r in rows_high}
    if "UPOZORENJE (DOOR_DRAWER_DRAWER_MIN)" not in codes_high:
        return False, "door_drawer_drawer_min_warning_absent"
    return True, f"low={len(rows_low)} high={len(rows_high)}"


def run_wall_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][3]["d_mm"] = 420
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (WALL_DEPTH)" not in codes:
        return False, "wall_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_base_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["d_mm"] = 480
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (BASE_DEPTH)" not in codes:
        return False, "base_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_tall_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "TALL_PANTRY"
    kitchen["modules"][0]["zone"] = "tall"
    kitchen["modules"][0]["d_mm"] = 480
    kitchen["modules"][0]["h_mm"] = 2100
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (TALL_DEPTH)" not in codes:
        return False, "tall_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_dishwasher_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][1]["template_id"] = "BASE_DISHWASHER"
    kitchen["modules"][1]["w_mm"] = 550
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (DISHWASHER_WIDTH)" not in codes:
        return False, "dishwasher_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_fridge_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "TALL_FRIDGE"
    kitchen["modules"][0]["zone"] = "tall"
    kitchen["modules"][0]["w_mm"] = 550
    kitchen["modules"][0]["h_mm"] = 2100
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (FRIDGE_WIDTH)" not in codes:
        return False, "fridge_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_cooking_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][2]["template_id"] = "BASE_COOKING_UNIT"
    kitchen["modules"][2]["w_mm"] = 550
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (COOKING_WIDTH)" not in codes:
        return False, "cooking_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_hob_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][2]["template_id"] = "BASE_HOB"
    kitchen["modules"][2]["w_mm"] = 400
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (HOB_WIDTH)" not in codes:
        return False, "hob_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_sink_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][1]["template_id"] = "SINK_BASE"
    kitchen["modules"][1]["w_mm"] = 550
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (SINK_WIDTH)" not in codes:
        return False, "sink_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_drawer_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "BASE_DRAWERS_3"
    kitchen["modules"][0]["d_mm"] = 430
    kitchen["modules"][0]["params"] = {}
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (DRAWER_DEPTH)" not in codes:
        return False, "drawer_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_liftup_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][3]["template_id"] = "WALL_LIFTUP"
    kitchen["modules"][3]["w_mm"] = 1300
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (LIFTUP_WIDTH)" not in codes:
        return False, "liftup_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_tall_height_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "TALL_PANTRY"
    kitchen["modules"][0]["zone"] = "tall"
    kitchen["modules"][0]["h_mm"] = 2500
    kitchen["modules"][0]["d_mm"] = 560
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (TALL_HEIGHT)" not in codes:
        return False, "tall_height_warning_absent"
    return True, f"warnings={len(rows)}"


def run_wall_appliance_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][3]["template_id"] = "WALL_HOOD"
    kitchen["modules"][3]["w_mm"] = 550
    kitchen["modules"][3]["d_mm"] = 320
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (WALL_APPLIANCE_WIDTH)" not in codes:
        return False, "wall_appliance_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_wall_appliance_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][3]["template_id"] = "WALL_HOOD"
    kitchen["modules"][3]["w_mm"] = 600
    kitchen["modules"][3]["d_mm"] = 280
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (WALL_APPLIANCE_DEPTH)" not in codes:
        return False, "wall_appliance_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_tall_appliance_width_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "TALL_OVEN"
    kitchen["modules"][0]["zone"] = "tall"
    kitchen["modules"][0]["w_mm"] = 550
    kitchen["modules"][0]["h_mm"] = 2100
    kitchen["modules"][0]["d_mm"] = 560
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (TALL_APPLIANCE_WIDTH)" not in codes:
        return False, "tall_appliance_width_warning_absent"
    return True, f"warnings={len(rows)}"


def run_tall_appliance_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "TALL_OVEN"
    kitchen["modules"][0]["zone"] = "tall"
    kitchen["modules"][0]["w_mm"] = 600
    kitchen["modules"][0]["h_mm"] = 2100
    kitchen["modules"][0]["d_mm"] = 540
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (TALL_APPLIANCE_DEPTH)" not in codes:
        return False, "tall_appliance_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_freestanding_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][2]["template_id"] = "BASE_OVEN_HOB_FREESTANDING"
    kitchen["modules"][2]["w_mm"] = 600
    kitchen["modules"][2]["d_mm"] = 570
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (FREESTANDING_DEPTH)" not in codes:
        return False, "freestanding_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_freestanding_fridge_depth_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][0]["template_id"] = "TALL_FRIDGE_FREESTANDING"
    kitchen["modules"][0]["zone"] = "tall"
    kitchen["modules"][0]["w_mm"] = 600
    kitchen["modules"][0]["h_mm"] = 2100
    kitchen["modules"][0]["d_mm"] = 590
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (FREESTANDING_FRIDGE_DEPTH)" not in codes:
        return False, "freestanding_fridge_depth_warning_absent"
    return True, f"warnings={len(rows)}"


def run_worktop_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["worktop"]["standard_lengths_mm"] = [2000]
    kitchen["wall"]["length_mm"] = 2400
    kitchen["wall_lengths_mm"] = {"A": 2400}
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (WORKTOP_TOO_SHORT)" not in codes:
        return False, "worktop_too_short_warning_absent"
    return True, f"warnings={len(rows)}"


def run_cutout_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"][1]["params"]["sink_cutout_x_mm"] = 500
    kitchen["modules"][1]["params"]["sink_cutout_width_mm"] = 400
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (CUTOUT_OUT_OF_BOUNDS)" not in codes:
        return False, "cutout_out_of_bounds_warning_absent"
    return True, f"warnings={len(rows)}"


def run_filler_warning_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 11,
            "template_id": "FILLER_PANEL",
            "zone": "base",
            "x_mm": 2200,
            "w_mm": 160,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Filer",
            "params": {},
            "wall_key": "A",
        }
    )
    rows = _manufacturing_warnings(
        kitchen["modules"],
        kitchen=kitchen,
        wall_h_mm=float(kitchen["wall"]["height_mm"]),
        foot_h_mm=float(kitchen["foot_height_mm"]),
        front_gap_mm=2.0,
    )
    codes = {str(r.get("Naziv", "")) for r in rows}
    if "UPOZORENJE (FILLER_TOO_WIDE)" not in codes:
        return False, "filler_too_wide_warning_absent"
    return True, f"warnings={len(rows)}"


def run_panel_service_processing_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 11,
            "template_id": "FILLER_PANEL",
            "zone": "base",
            "x_mm": 2200,
            "w_mm": 100,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Filer",
            "params": {},
            "wall_key": "A",
        }
    )
    kitchen["modules"].append(
        {
            "id": 12,
            "template_id": "END_PANEL",
            "zone": "base",
            "x_mm": 2300,
            "w_mm": 18,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Završna bočna",
            "params": {},
            "wall_key": "A",
        }
    )
    final_ds = get_final_cutlist_dataset(kitchen, lang="en")
    proc_df = final_ds["service_packet"].get("service_processing", pd.DataFrame())
    if proc_df is None or proc_df.empty:
        return False, "service_processing_missing"
    txt = " | ".join(proc_df.astype(str).fillna("").values.flatten().tolist())
    if "Filler panel" not in txt:
        return False, "filler_service_processing_missing"
    if "End side panel" not in txt:
        return False, "end_panel_service_processing_missing"
    if "Treat as a narrow finishing filler" not in txt:
        return False, "filler_service_processing_note_missing"
    if "Treat as the finished end side" not in txt:
        return False, "end_panel_service_processing_note_missing"
    return True, "ok"


def run_worktop_excel_sheet_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    xlsx_bytes = generate_cutlist_excel(kitchen, lang="en")
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    if "Worktop spec" not in wb.sheetnames:
        return False, "worktop_spec_sheet_missing"
    ws = wb["Worktop spec"]
    values = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        values.extend("" if v is None else str(v) for v in row)
    txt = " | ".join(values)
    if "WORKTOP SPECIFICATION" not in txt:
        return False, "worktop_spec_title_missing"
    if "Worktop" not in txt:
        return False, "worktop_spec_row_missing"

    xlsx_sr = generate_cutlist_excel(kitchen, lang="sr")
    wb_sr = openpyxl.load_workbook(BytesIO(xlsx_sr), data_only=True)
    if "Spec ploce" not in wb_sr.sheetnames:
        return False, "worktop_spec_sheet_missing_sr"
    ws_sr = wb_sr["Spec ploce"]
    sr_values = []
    for row in ws_sr.iter_rows(min_row=1, max_row=min(ws_sr.max_row, 12), values_only=True):
        sr_values.extend("" if v is None else str(v) for v in row)
    sr_txt = " | ".join(sr_values)
    if "Specifikacija radne ploče" not in sr_txt:
        return False, "worktop_spec_title_missing_sr"
    return True, "ok"


def run_worktop_pdf_instruction_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()

    pdf_en = build_cutlist_pdf_bytes(
        kitchen,
        generate_cutlist(kitchen),
        project_title="Worktop clarity check",
        lang="en",
    )
    reader_en = PdfReader(BytesIO(pdf_en))
    txt_en = " ".join("\n".join((page.extract_text() or "") for page in reader_en.pages).split())
    if "Worktop specification" not in txt_en:
        return False, "worktop_pdf_title_missing_en"
    if "Workshop works strictly by CUT dimensions." not in txt_en:
        return False, "worktop_pdf_cut_instruction_missing_en"
    if "Wall requirement / finished dimension" not in txt_en:
        return False, "worktop_pdf_finished_dimension_missing_en"

    pdf_sr = build_cutlist_pdf_bytes(
        kitchen,
        generate_cutlist(kitchen),
        project_title="Provera radne ploce",
        lang="sr",
    )
    reader_sr = PdfReader(BytesIO(pdf_sr))
    txt_sr = " ".join("\n".join((page.extract_text() or "") for page in reader_sr.pages).split())
    if "Specifikacija radne plo" not in txt_sr:
        return False, "worktop_pdf_title_missing_sr"
    if "Servis radi isključivo po CUT merama." not in txt_sr:
        return False, "worktop_pdf_cut_instruction_missing_sr"
    if "Zidna mera / finished dimension" not in txt_sr:
        return False, "worktop_pdf_finished_dimension_missing_sr"
    return True, "ok"


def run_pdf_note_formatting_check() -> tuple[bool, str]:
    sr_note = _format_pdf_table_cell(
        "Opis problema | Preporuka: Pomeri element. Finalni rez na licu mesta.",
        "Napomena",
    )
    if "<br/>" not in sr_note:
        return False, "pdf_note_linebreak_missing_sr"
    if "<br/>Preporuka:" not in sr_note:
        return False, "pdf_note_recommendation_break_missing_sr"
    if ".<br/>Finalni rez" not in sr_note:
        return False, "pdf_note_final_cut_break_missing_sr"

    en_note = _format_pdf_table_cell(
        "Wall requirement / finished dimension: 2400 mm. CUT basis / purchase segment: 3000 mm. Final cut is done on site.",
        "Workshop note",
    )
    if ".<br/>Final cut" not in en_note:
        return False, "pdf_note_final_cut_break_missing_en"
    return True, "ok"


def run_pdf_warning_priority_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()
    kitchen["modules"].append(
        {
            "id": 10,
            "template_id": "BASE_1DOOR",
            "zone": "base",
            "x_mm": 500,
            "w_mm": 700,
            "h_mm": 720,
            "d_mm": 560,
            "label": "Overlap test",
            "params": {},
            "wall_key": "A",
        }
    )
    pdf_en = generate_cutlist_pdf(kitchen, lang="en")
    txt_en = " ".join("\n".join((page.extract_text() or "") for page in PdfReader(BytesIO(pdf_en)).pages).split())
    hardware_idx = txt_en.find("Hardware and consumables")
    warnings_idx = txt_en.find("Critical warnings before production")
    warning_row_idx = txt_en.find("WARNING (OVERLAP)")
    if hardware_idx == -1:
        return False, "pdf_hardware_section_missing"
    if warnings_idx == -1:
        return False, "pdf_warning_header_missing"
    if warning_row_idx == -1:
        return False, "pdf_overlap_warning_missing"
    if not (hardware_idx < warnings_idx < warning_row_idx):
        return False, "pdf_warning_priority_order_invalid"
    return True, "ok"


def run_excel_intro_sheet_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()

    xlsx_sr = generate_cutlist_excel(kitchen, lang="sr")
    wb_sr = openpyxl.load_workbook(BytesIO(xlsx_sr), data_only=True)
    if "Kako koristiti" not in wb_sr.sheetnames:
        return False, "intro_sheet_missing_sr"
    if "Lista kupovine" not in wb_sr.sheetnames:
        return False, "shopping_sheet_missing_sr"
    if "Vodič" not in wb_sr.sheetnames:
        return False, "guide_sheet_missing_sr"
    ws_sr = wb_sr["Kako koristiti"]
    sr_values = []
    for row in ws_sr.iter_rows(min_row=1, max_row=min(ws_sr.max_row, 10), values_only=True):
        sr_values.extend("" if v is None else str(v) for v in row)
    sr_txt = " | ".join(sr_values)
    if "Kako koristiti dokument" not in sr_txt:
        return False, "intro_sheet_title_missing_sr"

    xlsx_en = generate_cutlist_excel(kitchen, lang="en")
    wb_en = openpyxl.load_workbook(BytesIO(xlsx_en), data_only=True)
    if "How to use" not in wb_en.sheetnames:
        return False, "intro_sheet_missing_en"
    if "Shopping" not in wb_en.sheetnames:
        return False, "shopping_sheet_missing_en"
    if "Guide" not in wb_en.sheetnames:
        return False, "guide_sheet_missing_en"
    return True, "ok"


def run_serbian_export_text_check() -> tuple[bool, str]:
    kitchen = build_sample_kitchen()

    pdf_bytes = generate_cutlist_pdf(kitchen, lang="sr")
    reader = PdfReader(BytesIO(pdf_bytes))
    pdf_text = "\n".join((page.extract_text() or "") for page in reader.pages[:4])
    pdf_checks = [
        "Kako koristiš ovaj dokument",
        "Ako ovde vidiš grešku, nemoj naručivati sečenje.",
        "Posebno kupi gotove uređaje, okove i alat",
        "Za sklapanje prati deo 'Za montažu'",
        "Debljina radne ploče",
    ]
    for item in pdf_checks:
        if item not in pdf_text:
            return False, f"pdf_missing_sr_text:{item}"

    xlsx_bytes = generate_cutlist_excel(kitchen, lang="sr")
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Kako koristiti"]
    values = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        values.extend("" if v is None else str(v) for v in row)
    xlsx_text = " | ".join(values)
    xlsx_checks = [
        "Kako koristiti dokument",
        "Ako ovde vidiš grešku, nemoj slati dokument u servis.",
        "Posebno kupi gotove uređaje, okove i alat",
        "završna provera",
    ]
    for item in xlsx_checks:
        if item not in xlsx_text:
            return False, f"xlsx_missing_sr_text:{item}"

    return True, "ok"


def _run_reference_kitchen_check(name: str, kitchen: dict) -> tuple[bool, str]:
    final_ds = get_final_cutlist_dataset(kitchen, lang="en")
    summary = final_ds.get("summary", {}) or {}
    detail_df = summary.get("summary_detaljna", pd.DataFrame())
    if detail_df is None or detail_df.empty:
        return False, f"{name}:empty_detail"

    warnings_df = final_ds.get("sections", {}).get("warnings", pd.DataFrame())
    if warnings_df is not None and not warnings_df.empty and "Naziv" in warnings_df.columns:
        codes = " | ".join(warnings_df["Naziv"].astype(str).tolist())
        banned = (
            "INVALID_DIMENSIONS",
            "INVALID_WALL_REFERENCE",
            "MISSING_TEMPLATE",
            "OVERLAP",
            "WORKTOP_TOO_SHORT",
            "CUTOUT_OUT_OF_BOUNDS",
            "BASE_ALIGNMENT_INVALID",
        )
        if any(code in codes for code in banned):
            return False, f"{name}:critical_warning:{codes}"

    csv_bytes = generate_cutlist_csv(kitchen, lang="en")
    if len(csv_bytes) < 500:
        return False, f"{name}:csv_too_small"
    pdf_bytes = build_cutlist_pdf_bytes(kitchen, final_ds["sections"], project_title=name, lang="en")
    if len(pdf_bytes) < 5000:
        return False, f"{name}:pdf_too_small"
    xlsx_bytes = generate_cutlist_excel(kitchen, lang="en")
    if len(xlsx_bytes) < 5000:
        return False, f"{name}:xlsx_too_small"
    return True, f"{name}:rows={len(detail_df)}"


def run_reference_linear_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("linear_kitchen", build_reference_linear_kitchen())


def run_reference_tall_block_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("tall_block_kitchen", build_reference_tall_block_kitchen())


def run_reference_utility_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("utility_kitchen", build_reference_utility_kitchen())


def run_reference_raised_dishwasher_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("raised_dishwasher_kitchen", build_reference_raised_dishwasher_kitchen())


def run_reference_l_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("l_kitchen", build_reference_l_kitchen())


def run_reference_galley_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("galley_kitchen", build_reference_galley_kitchen())


def run_reference_u_kitchen_check() -> tuple[bool, str]:
    return _run_reference_kitchen_check("u_kitchen", build_reference_u_kitchen())


def run_raised_dishwasher_cutlist_check() -> tuple[bool, str]:
    kitchen = _default_kitchen()
    kitchen["wall"]["length_mm"] = 2400
    kitchen["wall"]["height_mm"] = 2600
    kitchen["foot_height_mm"] = 150
    kitchen["base_korpus_h_mm"] = 812
    kitchen["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 3.8, "depth_mm": 600}
    kitchen["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    kitchen["modules"] = [
        {
            "id": 1,
            "template_id": "BASE_DISHWASHER",
            "zone": "base",
            "x_mm": 0,
            "w_mm": 600,
            "h_mm": 812,
            "d_mm": 560,
            "label": "MZS",
            "params": {},
            "wall_key": "A",
        }
    ]
    sections = generate_cutlist(kitchen)
    fronts = sections.get("fronts", pd.DataFrame())
    carcass = sections.get("carcass", pd.DataFrame())
    if fronts is None or fronts.empty:
        return False, "dishwasher_fronts_missing"
    ftxt = " | ".join(fronts.astype(str).fillna("").values.flatten().tolist())
    if "Donja maska — MZS" not in ftxt:
        return False, "dishwasher_lower_filler_missing"
    if carcass is None or carcass.empty:
        return False, "dishwasher_carcass_missing"
    ctxt = " | ".join(carcass.astype(str).fillna("").values.flatten().tolist())
    if "Postolje / nosač — MZS" not in ctxt:
        return False, "dishwasher_platform_missing"
    return True, "ok"


def run_warning_translation_check() -> tuple[bool, str]:
    sr_txt = "Element koristi nepoznat template_id 'BAD_ID'. | Preporuka: Proveri module_templates.json ili zameni element važećim template-om pre eksporta."
    en_txt = _translate_export_text(sr_txt, lang="en", column="Napomena")
    if "unknown template_id" not in str(en_txt):
        return False, "missing_template_not_translated"

    sr_txt2 = "Element referiše na nepostojeći zid 'Z'. | Preporuka: Postavi wall_key na jedan od važećih zidova: A, B."
    en_txt2 = _translate_export_text(sr_txt2, lang="en", column="Napomena")
    if "non-existent wall" not in str(en_txt2):
        return False, "invalid_wall_not_translated"

    sr_txt3 = "Element ima nevažeće dimenzije w=0, h=720, d=560 mm. | Preporuka: Proveri širinu, visinu i dubinu elementa pre eksporta."
    en_txt3 = _translate_export_text(sr_txt3, lang="en", column="Napomena")
    if "invalid dimensions" not in str(en_txt3):
        return False, "invalid_dimensions_not_translated"

    sr_txt4 = "Potrebna dužina radne ploče za zid A je 2420 mm, a najveća standardna nabavna dužina je 2000 mm. | Preporuka: Dodaj spoj/raskid ploče, koristi duži komercijalni format ili naruči specijalnu radnu ploču."
    en_txt4 = _translate_export_text(sr_txt4, lang="en", column="Napomena")
    if "Required worktop length for wall A" not in str(en_txt4):
        return False, "worktop_too_short_not_translated"

    sr_txt5 = "Izrez za sudoperu izlazi van radne ploče/modula (X=500, W=400, D=480 mm; modul W=800, D=560 mm). | Preporuka: Smanji izrez ili ga pomeri tako da ostane unutar širine i dubine modula."
    en_txt5 = _translate_export_text(sr_txt5, lang="en", column="Napomena")
    if "The sink cut-out extends outside the worktop/unit" not in str(en_txt5):
        return False, "cutout_out_of_bounds_not_translated"

    sr_txt6 = "Filer panel širine 160 mm je širok za standardnu popunu prostora. | Preporuka: Razmotri pravi završni panel ili poseban uski modul umesto širokog filera."
    en_txt6 = _translate_export_text(sr_txt6, lang="en", column="Napomena")
    if "Filler panel width 160 mm" not in str(en_txt6):
        return False, "filler_too_wide_not_translated"

    return True, "ok"


def run_raised_dishwasher_translation_check() -> tuple[bool, str]:
    kitchen = _default_kitchen()
    kitchen["wall"]["length_mm"] = 2400
    kitchen["wall"]["height_mm"] = 2600
    kitchen["foot_height_mm"] = 150
    kitchen["base_korpus_h_mm"] = 812
    kitchen["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 3.8, "depth_mm": 600}
    kitchen["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    kitchen["modules"] = [
        {
            "id": 1,
            "template_id": "BASE_DISHWASHER",
            "zone": "base",
            "x_mm": 0,
            "w_mm": 600,
            "h_mm": 812,
            "d_mm": 560,
            "label": "MZS",
            "params": {},
            "wall_key": "A",
        }
    ]

    final_ds = get_final_cutlist_dataset(kitchen, lang="en")
    det_df = final_ds["summary"].get("summary_detaljna", pd.DataFrame())
    if det_df is None or det_df.empty:
        return False, "summary_detaljna_missing"

    txt = " | ".join(det_df.astype(str).fillna("").values.flatten().tolist())
    if "Donja maska" in txt:
        return False, "dishwasher_lower_filler_not_translated"
    if "Postolje / nosa" in txt or "Postolje / nosac" in txt:
        return False, "dishwasher_platform_not_translated"
    if "Dishwasher lower filler" not in txt:
        return False, "dishwasher_lower_filler_missing_en"
    if "Dishwasher platform / support" not in txt:
        return False, "dishwasher_platform_missing_en"
    return True, "ok"


def run_raised_dishwasher_height_consistency_check() -> tuple[bool, str]:
    kitchen = _default_kitchen()
    kitchen["wall"]["length_mm"] = 2400
    kitchen["wall"]["height_mm"] = 2600
    kitchen["foot_height_mm"] = 150
    kitchen["base_korpus_h_mm"] = 812
    kitchen["worktop"] = {"enabled": True, "material": "Iverica", "thickness": 3.8, "depth_mm": 600}
    kitchen["materials"] = {
        "carcass_material": "Iverica",
        "carcass_thk": 18,
        "front_material": "MDF",
        "front_thk": 19,
        "back_material": "Lesonit",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    kitchen["modules"] = [
        {
            "id": 1,
            "template_id": "BASE_DISHWASHER",
            "zone": "base",
            "x_mm": 0,
            "w_mm": 600,
            "h_mm": 812,
            "d_mm": 560,
            "label": "MZS",
            "params": {},
            "wall_key": "A",
        }
    ]

    final_ds = get_final_cutlist_dataset(kitchen, lang="sr")
    det_df = final_ds["summary"].get("summary_detaljna", pd.DataFrame())
    proc_df = final_ds["service_packet"].get("service_processing", pd.DataFrame())
    if det_df is None or det_df.empty:
        return False, "summary_detaljna_missing"
    if proc_df is None or proc_df.empty:
        return False, "service_processing_missing"

    filler = det_df[det_df["Deo"].astype(str) == "Donja maska — MZS"]
    if filler.empty:
        return False, "dishwasher_lower_filler_missing"
    filler_row = filler.iloc[0]
    if int(float(filler_row.get("Sirina [mm]", 0) or 0)) != 84:
        return False, "dishwasher_lower_filler_fin_mismatch"
    filler_note = str(filler_row.get("Napomena", ""))
    if "gotova visina maske = 84 mm" not in filler_note:
        return False, "dishwasher_lower_filler_fin_note_missing"
    if "CUT visina = 88 mm" not in filler_note:
        return False, "dishwasher_lower_filler_cut_note_missing"

    proc_txt = " | ".join(proc_df.astype(str).fillna("").values.flatten().tolist())
    if "CUT visinu donje maske 88 mm" not in proc_txt:
        return False, "dishwasher_service_cut_height_missing"
    if "donju masku 92 mm" in proc_txt:
        return False, "dishwasher_raw_height_leaked"
    return True, "ok"


def run_drawer_backs_in_detail_summary_check() -> tuple[bool, str]:
    k = _default_kitchen()
    k["wall"]["length_mm"] = 3000
    k["wall"]["height_mm"] = 2600
    k["foot_height_mm"] = 150
    k["worktop"] = {"enabled": True, "material": "Granit", "thickness": 4.0, "depth_mm": 600}
    k["materials"] = {
        "carcass_material": "Iver 18mm",
        "carcass_thk": 18,
        "front_material": "MDF 19mm",
        "front_thk": 19,
        "back_material": "Iver 8mm",
        "back_thk": 8,
        "edge_abs_thk": 2.0,
    }
    k["modules"] = [
        {"id": 1, "template_id": "BASE_DRAWERS_3", "zone": "base", "x_mm": 0, "w_mm": 600, "h_mm": 720, "d_mm": 560, "label": "Fiokar", "params": {"n_drawers": 3}, "wall_key": "A"},
    ]
    final_ds = get_final_cutlist_dataset(k, lang="sr")
    det_df = final_ds["summary"].get("summary_detaljna", pd.DataFrame())
    if det_df is None or det_df.empty:
        return False, "summary_detaljna_missing"
    txt = " | ".join(det_df.astype(str).fillna("").values.flatten().tolist())
    if "Zadnja strana sanduka" not in txt:
        return False, "drawer_back_missing_in_summary_detaljna"
    if "Prednja strana sanduka" not in txt or "Dno sanduka" not in txt or "Bočna ploča" not in txt:
        return False, "drawer_box_family_incomplete_in_summary_detaljna"
    return True, "ok"
